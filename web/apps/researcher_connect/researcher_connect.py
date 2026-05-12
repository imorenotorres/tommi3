"""
Researcher Connect — Networking app for UNINOVIS researchers.

Allows researchers to create profiles with research interests, expertise,
and collaboration preferences. Supports Excel import/export and search.
"""

import io
import json
import os
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Query
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel

STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
DATA_PATH = os.path.join(os.path.dirname(__file__), "data.json")

router = APIRouter(prefix="/researcher_connect", tags=["researcher_connect"])


# -- Auth helpers (shared with tommi server) -----------------------------------

from auth import get_session, ROLES


def _get_token(request: Request) -> str | None:
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        return auth_header[7:]
    return request.query_params.get("token")


def _require_auth(request: Request) -> dict:
    token = _get_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Authentication required")
    session = get_session(token)
    if not session:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    return session


def _require_editor(request: Request) -> dict:
    session = _require_auth(request)
    if ROLES.get(session["role"], 0) < ROLES.get("tester", 99):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return session


# -- Data I/O ------------------------------------------------------------------

def load_data() -> dict:
    with open(DATA_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_data(data: dict):
    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# -- Models --------------------------------------------------------------------

class ResearcherBody(BaseModel):
    first_name: str
    family_name: str
    email: str
    university: str
    department: str = ""
    position: str = ""
    research_areas: list[str] = []
    keywords: list[str] = []
    bio: str = ""
    looking_for: str = ""  # what kind of collaboration they seek
    orcid: str = ""
    website: str = ""


# -- Routes --------------------------------------------------------------------

@router.get("/")
def index():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))


@router.get("/api/auth-check")
def auth_check(session: dict = Depends(_require_auth)):
    can_edit = ROLES.get(session["role"], 0) >= ROLES.get("tester", 99)
    return {"username": session["username"], "role": session["role"], "can_edit": can_edit}


def _find_own_researcher(data: dict, username: str):
    """Return the researcher record linked to this username, or None."""
    for r in data["researchers"]:
        if (r.get("email", "").lower() == username or
                r.get("linked_username", "").lower() == username):
            return r
    return None


@router.get("/api/my-profile")
def get_my_profile(session: dict = Depends(_require_auth)):
    data = load_data()
    r = _find_own_researcher(data, session["username"].lower())
    if r is None:
        raise HTTPException(404, "No researcher profile found for your account")
    return r


@router.post("/api/my-profile")
def create_my_profile(body: ResearcherBody, session: dict = Depends(_require_auth)):
    data = load_data()
    username = session["username"].lower()
    if _find_own_researcher(data, username) is not None:
        raise HTTPException(409, "A profile already exists for your account")
    if body.university not in data["universities"]:
        raise HTTPException(400, f"Unknown university: {body.university}")
    researcher = {
        "id": str(uuid.uuid4())[:8],
        "first_name": body.first_name,
        "family_name": body.family_name,
        "email": body.email,
        "university": body.university,
        "department": body.department,
        "position": body.position,
        "research_areas": body.research_areas,
        "keywords": body.keywords,
        "bio": body.bio,
        "looking_for": body.looking_for,
        "orcid": body.orcid,
        "website": body.website,
        "linked_username": username,
    }
    data["researchers"].append(researcher)
    save_data(data)
    return researcher


@router.put("/api/my-profile")
def update_my_profile(body: ResearcherBody, session: dict = Depends(_require_auth)):
    data = load_data()
    username = session["username"].lower()
    if body.university not in data["universities"]:
        raise HTTPException(400, f"Unknown university: {body.university}")
    r = _find_own_researcher(data, username)
    if r is None:
        raise HTTPException(404, "No researcher profile found for your account")
    r.update({
        "first_name": body.first_name,
        "family_name": body.family_name,
        "email": body.email,
        "university": body.university,
        "department": body.department,
        "position": body.position,
        "research_areas": body.research_areas,
        "keywords": body.keywords,
        "bio": body.bio,
        "looking_for": body.looking_for,
        "orcid": body.orcid,
        "website": body.website,
        "linked_username": username,
    })
    save_data(data)
    return r


@router.get("/api/data")
def get_all_data():
    """Return all researcher data (public read)."""
    return load_data()


# -- Researcher CRUD -----------------------------------------------------------

@router.post("/api/researchers")
def create_researcher(body: ResearcherBody, session: dict = Depends(_require_auth)):
    is_editor = ROLES.get(session["role"], 0) >= ROLES.get("tester", 99)
    data = load_data()
    if not is_editor:
        username = session["username"].lower()
        if body.email.lower() != username:
            raise HTTPException(403, "You can only create a profile for your own email address")
        if any(r.get("email", "").lower() == username for r in data["researchers"]):
            raise HTTPException(409, "A profile already exists for your account")
    if body.university not in data["universities"]:
        raise HTTPException(400, f"Unknown university: {body.university}")
    researcher = {
        "id": str(uuid.uuid4())[:8],
        "first_name": body.first_name,
        "family_name": body.family_name,
        "email": body.email,
        "university": body.university,
        "department": body.department,
        "position": body.position,
        "research_areas": body.research_areas,
        "keywords": body.keywords,
        "bio": body.bio,
        "looking_for": body.looking_for,
        "orcid": body.orcid,
        "website": body.website,
    }
    data["researchers"].append(researcher)
    save_data(data)
    return researcher


@router.put("/api/researchers/{researcher_id}")
def update_researcher(researcher_id: str, body: ResearcherBody, session: dict = Depends(_require_auth)):
    is_editor = ROLES.get(session["role"], 0) >= ROLES.get("tester", 99)
    data = load_data()
    for r in data["researchers"]:
        if r["id"] == researcher_id:
            username = session["username"].lower()
            is_owner = (r.get("email", "").lower() == username or
                        r.get("linked_username", "").lower() == username)
            if not is_editor and not is_owner:
                raise HTTPException(403, "You can only edit your own profile")
            if body.university not in data["universities"]:
                raise HTTPException(400, f"Unknown university: {body.university}")
            r.update({
                "first_name": body.first_name,
                "family_name": body.family_name,
                "email": body.email,
                "university": body.university,
                "department": body.department,
                "position": body.position,
                "research_areas": body.research_areas,
                "keywords": body.keywords,
                "bio": body.bio,
                "looking_for": body.looking_for,
                "orcid": body.orcid,
                "website": body.website,
            })
            save_data(data)
            return r
    raise HTTPException(404, "Researcher not found")


@router.delete("/api/researchers/{researcher_id}")
def delete_researcher(researcher_id: str, session: dict = Depends(_require_editor)):
    data = load_data()
    data["researchers"] = [r for r in data["researchers"] if r["id"] != researcher_id]
    save_data(data)
    return {"ok": True}


# -- Research areas management -------------------------------------------------

class AreasBody(BaseModel):
    research_areas: list[str]


@router.put("/api/research-areas")
def update_research_areas(body: AreasBody, session: dict = Depends(_require_editor)):
    data = load_data()
    data["research_areas"] = body.research_areas
    save_data(data)
    return {"ok": True}


# -- Excel export --------------------------------------------------------------

@router.get("/api/export")
def export_excel():
    """Export all researchers to an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    data = load_data()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Researchers"

    headers = [
        "First Name", "Family Name", "Email", "University",
        "Department", "Position", "Research Areas", "Keywords",
        "Bio", "Looking For", "ORCID", "Website"
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4250B3", end_color="4250B3", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, r in enumerate(data["researchers"], 2):
        ws.cell(row=row_idx, column=1, value=r.get("first_name", ""))
        ws.cell(row=row_idx, column=2, value=r.get("family_name", ""))
        ws.cell(row=row_idx, column=3, value=r.get("email", ""))
        ws.cell(row=row_idx, column=4, value=r.get("university", ""))
        ws.cell(row=row_idx, column=5, value=r.get("department", ""))
        ws.cell(row=row_idx, column=6, value=r.get("position", ""))
        ws.cell(row=row_idx, column=7, value="; ".join(r.get("research_areas", [])))
        ws.cell(row=row_idx, column=8, value="; ".join(r.get("keywords", [])))
        ws.cell(row=row_idx, column=9, value=r.get("bio", ""))
        ws.cell(row=row_idx, column=10, value=r.get("looking_for", ""))
        ws.cell(row=row_idx, column=11, value=r.get("orcid", ""))
        ws.cell(row=row_idx, column=12, value=r.get("website", ""))

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=uninovis_researchers.xlsx"},
    )


# -- Excel import --------------------------------------------------------------

@router.post("/api/import")
async def import_excel(file: UploadFile = File(...), session: dict = Depends(_require_editor)):
    """Import researchers from an Excel file.

    Expected columns (header row): First Name, Family Name, Email, University,
    Department, Position, Research Areas, Keywords, Bio, Looking For, ORCID, Website.
    Research Areas and Keywords are semicolon-separated.
    """
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be .xlsx")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(400, "File must be under 5 MB")

    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty spreadsheet")

    # Map header names to column indices
    raw_headers = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
    col_map = {
        "first_name": None, "family_name": None, "email": None, "university": None,
        "department": None, "position": None, "research_areas": None, "keywords": None,
        "bio": None, "looking_for": None, "orcid": None, "website": None,
    }
    for idx, h in enumerate(raw_headers):
        if h in col_map:
            col_map[h] = idx

    if col_map["first_name"] is None or col_map["email"] is None:
        raise HTTPException(400, "Spreadsheet must have at least 'First Name' and 'Email' columns")

    data = load_data()
    existing_emails = {r["email"].lower() for r in data["researchers"]}
    added = 0
    skipped = 0

    for row in rows[1:]:
        def cell(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        email = cell("email").lower()
        if not email:
            skipped += 1
            continue
        if email in existing_emails:
            skipped += 1
            continue

        university = cell("university").upper()
        if university and university not in data["universities"]:
            university = ""

        areas_raw = cell("research_areas")
        areas = [a.strip() for a in areas_raw.split(";") if a.strip()] if areas_raw else []
        kw_raw = cell("keywords")
        keywords = [k.strip() for k in kw_raw.split(";") if k.strip()] if kw_raw else []

        researcher = {
            "id": str(uuid.uuid4())[:8],
            "first_name": cell("first_name"),
            "family_name": cell("family_name"),
            "email": email,
            "university": university,
            "department": cell("department"),
            "position": cell("position"),
            "research_areas": areas,
            "keywords": keywords,
            "bio": cell("bio"),
            "looking_for": cell("looking_for"),
            "orcid": cell("orcid"),
            "website": cell("website"),
        }
        data["researchers"].append(researcher)
        existing_emails.add(email)
        added += 1

    save_data(data)
    return {"added": added, "skipped": skipped}


# -- Excel template ------------------------------------------------------------

@router.get("/api/template")
def download_template():
    """Download an empty Excel template with the expected columns."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Researchers"

    headers = [
        "First Name", "Family Name", "Email", "University",
        "Department", "Position", "Research Areas", "Keywords",
        "Bio", "Looking For", "ORCID", "Website"
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4250B3", end_color="4250B3", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 20

    # Example row
    example = ["Jane", "Doe", "jane.doe@example.com", "UMA",
               "Computer Science", "Associate Professor",
               "Artificial Intelligence; Data Science", "NLP; machine learning",
               "Researcher in AI and NLP", "Joint publications; PhD co-supervision",
               "0000-0001-2345-6789", "https://example.com"]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=researcher_template.xlsx"},
    )
