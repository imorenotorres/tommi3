"""
Tommi Web Interface - FastAPI server para interactuar con agentes Tommi
"""

# Activar venv automáticamente si no está activo
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / "apps"))
from venv_helper import ensure_venv
ensure_venv()

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import FastAPI, File as FastFile, HTTPException, Query, Request, Depends, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, RedirectResponse
from pydantic import BaseModel

# Cargar .env de web/ con override para asegurar que se usa la configuración correcta
_web_env = Path(__file__).parent / ".env"
load_dotenv(_web_env, override=True)

from agent_runner import AgentRunner
from auth import (
    authenticate, approve_access_request, change_password, create_access_request,
    create_user, create_user_pending, create_invite_token, delete_user,
    ensure_superuser, get_session, has_role, list_access_requests, list_users,
    logout, reject_access_request, send_invite_email, set_password_from_invite,
    update_user_role, user_exists, validate_invite_token, validate_password,
    validate_uninovis_email, ROLES, TOOL_ACCESS, UNINOVIS_DOMAINS,
    can_access_tool, can_edit, user_roles, max_role_level,
)
from error_codes import (
    format_error,
    LLM_OLLAMA_NOT_RUNNING, LLM_MODEL_NOT_FOUND, LLM_OLLAMA_ERROR,
    LLM_OLLAMA_TIMEOUT, LLM_NO_API_KEY, LLM_INVALID_API_KEY,
    LLM_MISTRAL_ERROR, LLM_CONNECTION_ERROR, LLM_UNKNOWN_ERROR,
    LLM_VLLM_NOT_RUNNING, LLM_VLLM_ERROR, LLM_VLLM_MODEL_NOT_FOUND,
    AGENT_NOT_FOUND, SERVER_STREAMING_ERROR
)

# Configuración de logging (activable/desactivable via ENABLE_LOGGING)
ENABLE_LOGGING = os.getenv("ENABLE_LOGGING", "true").lower() in ("true", "1", "yes")

# All logs (conversations + feedback) stored in /logs, one file per agent
LOGS_DIR = Path(__file__).parent / "logs"
LOGS_DIR.mkdir(exist_ok=True)

# Per-agent conversation loggers (one .log file per agent, created on demand)
_agent_loggers: dict = {}


def _get_agent_logger(agent_id: str) -> logging.Logger:
    """Return (or create) a per-agent conversation logger."""
    if agent_id not in _agent_loggers:
        logger = logging.getLogger(f"conversations.{agent_id}")
        logger.setLevel(logging.INFO)
        logger.propagate = False
        handler = logging.FileHandler(
            LOGS_DIR / f"{agent_id}_conversations.log", encoding="utf-8"
        )
        handler.setFormatter(logging.Formatter("%(message)s"))
        logger.addHandler(handler)
        _agent_loggers[agent_id] = logger
    return _agent_loggers[agent_id]


def log_conversation(
    client_ip: str,
    agent_id: str,
    agent_name: str,
    question: str,
    response: str,
    session_id: str = "",
    transparency_level: str = None,
    username: str = None,
    extra: dict = None,
):
    """Registra una conversación en el log (si está habilitado).
    Writes a per-agent .log file and a per-agent .jsonl file."""
    if not ENABLE_LOGGING:
        return

    # Pseudonymize username for privacy (same method as AuditLogger)
    anon_user_id = None
    email_domain = None
    if username:
        import hashlib
        raw = f"tommi-uninovis-2026:{username}"
        anon_user_id = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]
        # Extract country TLD (e.g. "user@uma.es" -> "es")
        if "@" in username:
            email_domain = username.rsplit("@", 1)[1].rsplit(".", 1)[-1].lower()

    entry = {
        "timestamp": datetime.now().isoformat(),
        "client_ip": client_ip,
        "session_id": session_id,
        "user_id": anon_user_id,
        "email_domain": email_domain,
        "transparency_level": transparency_level,
        "agent_id": agent_id,
        "agent_name": agent_name,
        "question": question,
        "response": response[:500] + "..." if len(response) > 500 else response
    }
    if extra:
        entry.update(extra)

    # Per-agent .log (human-readable, pretty-printed)
    try:
        _get_agent_logger(agent_id).info(
            json.dumps(entry, ensure_ascii=False, indent=2) + "\n"
        )
    except Exception:
        pass
    # Per-agent .jsonl (machine-readable, one line per entry)
    try:
        agent_log = LOGS_DIR / f"{agent_id}_conversations.jsonl"
        with open(agent_log, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass

# Configuración
SCRIPT_DIR = Path(__file__).parent
AGENTS_PATH = SCRIPT_DIR.parent / "agents"  # tommi/agents/

# Inicializar runner
runner = AgentRunner(agents_base_path=str(AGENTS_PATH))

# FastAPI app
app = FastAPI(
    title="Tommi Web Interface",
    description="Interfaz web para agentes Tommi",
    version="1.0.0"
)

# Ensure a superuser exists on startup
_su = ensure_superuser()
logger = logging.getLogger("tommi")
logger.info(f"Superuser ready: {_su}")

# Auth middleware — protects /api/* routes (except /api/auth/login)
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import JSONResponse


class AuthMiddleware(BaseHTTPMiddleware):
    # Routes that don't require authentication
    PUBLIC_PATHS = {"/api/auth/login", "/api/auth/invite/validate", "/api/auth/invite/set-password", "/api/auth/request-access", "/api/auth/forgot-password"}

    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        # Only protect /api/ routes (not static files, HTML pages, etc.)
        # PDF endpoints are public (academic documents, not personal data)
        # Study app routes are public (participants don't need TOMMI accounts)
        is_study = path.startswith("/study/api/") or path.startswith("/rag-study/api/") or path.startswith("/sql-study/api/")
        # rag_study2 agents are public (research study participants don't need accounts)
        is_rag_study2 = "rag_study2_" in request.query_params.get("agent_id", "")
        is_public_search = any(path.endswith(s) for s in ("/publications-search", "/topic-search", "/collaboration-search", "/collaboration-map", "/projects-search", "/project-topic-search", "/publications-map", "/pdf-list")) and any(f"/{aid}/" in path for aid in ("responsible_ai3", "health_wellbeing_sistems"))
        is_tutores = path.startswith("/api/tutores/")
        is_lali_public = path in ("/api/eulalia/auth-level", "/api/agent/eulalia/transcripcion-config")
        if path.startswith("/api/") and path not in self.PUBLIC_PATHS and "/pdf/" not in path and "/quickguide" not in path and "/agreements-search" not in path and "/agreements-config" not in path and "/interaction-log" not in path and "/public-tools" not in path and "/public-agent/" not in path and "/api/feedback" != path and not is_study and not is_public_search and not is_rag_study2 and not is_tutores and not is_lali_public:
            token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
            if not token:
                token = request.query_params.get("token", "")
            if not token or not get_session(token):
                # Also accept tutores tokens for eulalia agent
                is_lali_chat = (request.query_params.get("agent_id") == "eulalia"
                                and token and tutores_get_session(token))
                if not is_lali_chat:
                    return JSONResponse(status_code=401, content={"detail": "Authentication required"})
        return await call_next(request)


app.add_middleware(AuthMiddleware)


# ── Rate limiting middleware ──────────────────────────────────────────
import time as _rl_time
from collections import defaultdict

class RateLimiter:
    """In-memory rate limiter per key (IP or user)."""

    def __init__(self):
        self._buckets = defaultdict(list)  # key -> list of timestamps
        self._cleanup_counter = 0

    def is_allowed(self, key: str, max_requests: int, window_seconds: int) -> bool:
        now = _rl_time.time()
        cutoff = now - window_seconds
        # Remove old entries
        self._buckets[key] = [t for t in self._buckets[key] if t > cutoff]
        if len(self._buckets[key]) >= max_requests:
            return False
        self._buckets[key].append(now)
        # Periodic cleanup of stale keys (every 100 checks)
        self._cleanup_counter += 1
        if self._cleanup_counter % 100 == 0:
            stale = [k for k, v in self._buckets.items() if not v or v[-1] < cutoff]
            for k in stale:
                del self._buckets[k]
        return True


_rate_limiter = RateLimiter()

# Rate limit tiers
_RL_CHAT = (15, 60)      # 15 requests per 60 seconds
_RL_CHAT_HOUR = (300, 3600)  # 300 per hour
_RL_DATA = (120, 60)     # 120 per minute
_RL_WRITE = (20, 60)     # 20 per minute
_RL_LOGIN = (10, 60)     # 10 per minute


class RateLimitMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        path = request.url.path
        method = request.method

        # Only rate-limit API calls
        if not path.startswith("/api/"):
            return await call_next(request)

        # Skip rate limiting for localhost (development)
        raw_ip = request.client.host if request.client else ""
        if raw_ip in ("127.0.0.1", "::1", "localhost"):
            return await call_next(request)

        # Identify the caller: prefer username > session_id > IP
        # Use X-Forwarded-For from reverse proxy to get the real client IP
        forwarded = request.headers.get("x-forwarded-for")
        if forwarded:
            client_ip = forwarded.split(",")[0].strip()
        else:
            client_ip = request.client.host if request.client else "unknown"
        token = request.query_params.get("token", "")
        session_id = request.query_params.get("session_id", "")
        caller_key = client_ip

        # 1. Best: authenticated user
        if token:
            try:
                from auth_tutores import get_session as _rl_get_session
                session = _rl_get_session(token)
                if session:
                    caller_key = "user:" + session.get("username", client_ip)
            except Exception:
                pass

        # 2. Fallback: chat session_id (unique per browser tab)
        if caller_key == client_ip and session_id:
            caller_key = "session:" + session_id

        # Determine tier based on path
        if "moodle-login" in path or "login" in path:
            tier = _RL_LOGIN
            tier_key = "login:" + client_ip  # Always by IP for login
        elif "chat/stream" in path:
            # Chat: check both per-minute and per-hour limits
            minute_key = "chat_min:" + caller_key
            hour_key = "chat_hr:" + caller_key
            if not _rate_limiter.is_allowed(minute_key, *_RL_CHAT):
                return JSONResponse(
                    {"error": "Demasiadas consultas al chat. Espera un momento antes de enviar otra."},
                    status_code=429
                )
            if not _rate_limiter.is_allowed(hour_key, *_RL_CHAT_HOUR):
                return JSONResponse(
                    {"error": "Has alcanzado el límite de consultas por hora. Inténtalo más tarde."},
                    status_code=429
                )
            return await call_next(request)
        elif method == "POST" and any(w in path for w in ("consulta", "progreso", "retos", "revision", "registrar", "solicitar")):
            tier = _RL_WRITE
            tier_key = "write:" + caller_key
        else:
            tier = _RL_DATA
            tier_key = "data:" + caller_key

        if not _rate_limiter.is_allowed(tier_key, *tier):
            return JSONResponse(
                {"error": "Demasiadas peticiones. Espera un momento."},
                status_code=429
            )

        return await call_next(request)


app.add_middleware(RateLimitMiddleware)

# Security headers middleware
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        # Allow framing from same origin and trusted UMA domains
        # X-Frame-Options SAMEORIGIN as baseline (no ALLOW-FROM, not supported in modern browsers)
        response.headers["X-Frame-Options"] = "SAMEORIGIN"
        # Prevent MIME type sniffing
        response.headers["X-Content-Type-Options"] = "nosniff"
        # XSS protection (legacy browsers)
        response.headers["X-XSS-Protection"] = "1; mode=block"
        # Referrer policy: don't leak full URL to external sites
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        # HSTS: force HTTPS (only effective on HTTPS, ignored on localhost)
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        # CSP: restrict sources while allowing CDNs used by agents
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com; "
            "style-src 'self' 'unsafe-inline' https://unpkg.com; "
            "img-src 'self' data: https:; "
            "font-src 'self' https:; "
            "connect-src 'self'; "
            "frame-src 'self' https:; "
            "frame-ancestors 'self' https://relacionesi.uma.es https://*.uma.es"
        )
        return response


app.add_middleware(SecurityHeadersMiddleware)

# Servir archivos estáticos
app.mount("/static", StaticFiles(directory=SCRIPT_DIR / "static"), name="static")
app.mount("/img", StaticFiles(directory=SCRIPT_DIR / "img"), name="img")

# Temp directory for generated files (Praat spectrograms, etc.)
_TEMP_DIR = SCRIPT_DIR / "temp"
_TEMP_DIR.mkdir(exist_ok=True)
app.mount("/temp", StaticFiles(directory=_TEMP_DIR), name="temp")

# Mount UNIGRACON app (grade converter)
from apps.unigracon.unigracon import router as unigracon_router
app.include_router(unigracon_router)
app.mount("/unigracon/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "unigracon" / "static"), name="unigracon_static")

# Mount Mobility Planner app
from apps.mobility_planner.mobility_planner import router as mobility_router
app.include_router(mobility_router)
app.mount("/mobility_planner/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "mobility_planner" / "static"), name="mobility_static")

# Mount Directory app
from apps.directory.directory import router as directory_router
app.include_router(directory_router)
app.mount("/directory/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "directory" / "static"), name="directory_static")

# Mount UNINOVIS Admin hub
from apps.uninovis.uninovis import router as uninovis_router
app.include_router(uninovis_router)
app.mount("/uninovis/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "uninovis" / "static"), name="uninovis_static")

# Mount Researcher Connect app
from apps.researcher_connect.researcher_connect import router as researcher_connect_router
app.include_router(researcher_connect_router)
app.mount("/researcher_connect/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "researcher_connect" / "static"), name="researcher_connect_static")

# Mount Transparency Study apps (TODO: create apps/rag_study/)
# from apps.rag_study.study import router as rag_study_router
# app.include_router(rag_study_router)
# app.mount("/rag-study/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "rag_study" / "static"), name="rag_study_static")

# Mount Event Tracker app
from apps.event_tracker.event_tracker import router as event_tracker_router
app.include_router(event_tracker_router)
app.mount("/event-tracker/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "event_tracker" / "static"), name="event_tracker_static")

# Mount Matomo Analytics app
from apps.matomo_analytics.matomo_analytics import router as matomo_analytics_router
app.include_router(matomo_analytics_router)
app.mount("/matomo-analytics/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "matomo_analytics" / "static"), name="matomo_analytics_static")

# Mount Survey DATA FOR L.I.F.E. app
from apps.survey_datalife.survey_datalife import router as survey_datalife_router
app.include_router(survey_datalife_router)
app.mount("/survey-datalife/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "survey_datalife" / "static"), name="survey_datalife_static")

# Mount Research Proposals app
from apps.research_proposals.research_proposals import router as research_proposals_router
app.include_router(research_proposals_router)
app.mount("/research_proposals/static", StaticFiles(directory=SCRIPT_DIR / "apps" / "research_proposals" / "static"), name="research_proposals_static")


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def _get_token(request: Request) -> str | None:
    """Extract bearer token from Authorization header or query param."""
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:]
    return request.query_params.get("token")


def require_auth(request: Request) -> dict:
    """Dependency: require a valid session. Returns session dict."""
    token = _get_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Authentication required")
    session = get_session(token)
    if not session:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    return session


def require_role(minimum_role: str):
    """Dependency factory: require at least the given role level."""
    from auth import max_role_level as _max_level
    def _check(session: dict = Depends(require_auth)):
        if _max_level(session) < ROLES.get(minimum_role, 99):
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return session
    return _check


# ---------------------------------------------------------------------------
# Auth API routes
# ---------------------------------------------------------------------------

class LoginRequest(BaseModel):
    username: str
    password: str


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


class CreateUserRequest(BaseModel):
    username: str
    password: str
    role: str
    roles: list[str] = []


class UpdateRoleRequest(BaseModel):
    role: str
    roles: list[str] = []


@app.post("/api/auth/login")
async def api_login(req: LoginRequest):
    result = authenticate(req.username, req.password)
    if not result:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    return result


@app.post("/api/auth/logout")
async def api_logout(request: Request):
    token = _get_token(request)
    if token:
        logout(token)
    return {"ok": True}


@app.get("/api/auth/me")
async def api_me(session: dict = Depends(require_auth)):
    from auth import _load_users, is_study_mode
    users = _load_users()
    user = users.get(session["username"], {})
    result = {
        "username": session["username"],
        "role": session["role"],
        "roles": session.get("roles", [session["role"]]),
        "provisional_password": user.get("provisional_password", False),
    }
    if is_study_mode():
        result["study_mode"] = True
        result["study_condition"] = user.get("study_condition")
    return result


@app.post("/api/auth/change-password")
async def api_change_password(req: ChangePasswordRequest, session: dict = Depends(require_auth)):
    pwd_error = validate_password(req.new_password)
    if pwd_error:
        raise HTTPException(status_code=400, detail=pwd_error)
    ok = change_password(session["username"], req.old_password, req.new_password)
    if not ok:
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    return {"ok": True}


@app.get("/api/auth/users")
async def api_list_users(session: dict = Depends(require_role("superuser"))):
    return list_users()


@app.post("/api/auth/users")
async def api_create_user(req: CreateUserRequest, session: dict = Depends(require_role("superuser"))):
    all_roles = req.roles if req.roles else [req.role]
    for r in all_roles:
        if r not in ROLES:
            raise HTTPException(status_code=400, detail=f"Invalid role '{r}'. Must be one of: {list(ROLES.keys())}")
    pwd_error = validate_password(req.password)
    if pwd_error:
        raise HTTPException(status_code=400, detail=pwd_error)
    ok = create_user(req.username, req.password, all_roles[0], provisional=True, roles=all_roles)
    if not ok:
        raise HTTPException(status_code=409, detail="Username already exists")
    return {"ok": True, "username": req.username, "roles": all_roles}


@app.post("/api/auth/users/bulk")
async def api_bulk_create_users(
    file: UploadFile,
    session: dict = Depends(require_role("superuser")),
):
    """
    Bulk-create users from a TSV or Excel (.xlsx) file.
    Expected columns: username, password, role
    Header row is optional (auto-detected).
    All users are created with provisional_password=True.
    """
    filename = (file.filename or "").lower()
    content = await file.read()

    rows: list[list[str]] = []

    if filename.endswith((".xlsx", ".xls")):
        import io
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    rows.append(cells)
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel file: {e}")
    elif filename.endswith((".tsv", ".txt", ".csv")):
        import csv, io
        text = content.decode("utf-8-sig")
        # Auto-detect delimiter
        dialect = csv.Sniffer().sniff(text[:2048], delimiters="\t,;")
        reader = csv.reader(io.StringIO(text), dialect)
        for row in reader:
            cells = [c.strip() for c in row]
            if any(cells):
                rows.append(cells)
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Use .xlsx, .tsv, .csv, or .txt",
        )

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    # Auto-detect header: if first row looks like a header, skip it
    first = [c.lower() for c in rows[0]]
    if "username" in first or "user" in first or "nombre" in first:
        rows = rows[1:]

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found after header")

    valid_roles = set(ROLES.keys())
    created = []
    skipped = []
    errors = []

    for i, row in enumerate(rows, start=1):
        if len(row) < 2:
            errors.append(f"Row {i}: not enough columns (need at least username and password)")
            continue

        username = row[0].strip()
        password = row[1].strip()
        role = row[2].strip().lower() if len(row) > 2 and row[2].strip() else "user"

        if not username:
            errors.append(f"Row {i}: empty username")
            continue
        pwd_err = validate_password(password)
        if pwd_err:
            errors.append(f"Row {i} ({username}): {pwd_err}")
            continue
        if role not in valid_roles:
            errors.append(f"Row {i} ({username}): invalid role '{role}' (must be {', '.join(valid_roles)})")
            continue

        ok = create_user(username, password, role, provisional=True)
        if ok:
            created.append({"username": username, "role": role})
        else:
            skipped.append(f"{username} (already exists)")

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "total_created": len(created),
        "total_skipped": len(skipped),
        "total_errors": len(errors),
    }


@app.delete("/api/auth/users/{username}")
async def api_delete_user(username: str, session: dict = Depends(require_role("superuser"))):
    if username == session["username"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    ok = delete_user(username)
    if not ok:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


@app.put("/api/auth/users/{username}/role")
async def api_update_role(username: str, req: UpdateRoleRequest, session: dict = Depends(require_role("superuser"))):
    all_roles = req.roles if req.roles else [req.role]
    for r in all_roles:
        if r not in ROLES:
            raise HTTPException(status_code=400, detail=f"Invalid role '{r}'. Must be one of: {list(ROLES.keys())}")
    if username == session["username"]:
        raise HTTPException(status_code=400, detail="Cannot change your own role")
    ok = update_user_role(username, all_roles[0], all_roles)
    if not ok:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


# ---------------------------------------------------------------------------
# Invitation / email routes
# ---------------------------------------------------------------------------

# SMTP config — re-read from .env on each call so changes don't require restart
def _check_directory_email(email: str) -> str:
    """Check if an email exists in the directory. Returns the person's name if found, empty string if not."""
    try:
        import json
        directory_path = SCRIPT_DIR / "apps" / "directory" / "data.json"
        if not directory_path.exists():
            return ""
        with open(directory_path, encoding="utf-8") as f:
            data = json.load(f)
        email_lower = email.lower()
        for user in data.get("users", []):
            if user.get("email", "").lower() == email_lower:
                return f"{user.get('first_name', '')} {user.get('family_name', '')}".strip()
    except Exception:
        pass
    return ""


def _get_smtp_config() -> dict:
    """Read SMTP settings from .env file each time (not cached)."""
    from dotenv import dotenv_values
    env = dotenv_values(_web_env)
    host = env.get("SMTP_HOST", "")
    user = env.get("SMTP_USER", "")
    password = env.get("SMTP_PASSWORD", "")
    return {
        "host": host,
        "port": int(env.get("SMTP_PORT", "587")),
        "user": user,
        "password": password,
        "from_addr": env.get("SMTP_FROM", "") or user,
        "use_tls": env.get("SMTP_USE_TLS", "true").lower() in ("true", "1", "yes"),
        "configured": bool(host and user and password),
    }


class InviteUserRequest(BaseModel):
    username: str  # email address
    role: str


class SetPasswordRequest(BaseModel):
    token: str
    password: str


@app.get("/api/auth/smtp-status")
async def api_smtp_status(session: dict = Depends(require_role("superuser"))):
    """Check if SMTP is configured."""
    smtp = _get_smtp_config()
    return {"configured": smtp["configured"]}


@app.post("/api/auth/invite")
async def api_invite_user(req: InviteUserRequest, request: Request, session: dict = Depends(require_role("superuser"))):
    """Create a user and send an invitation email to set their password."""
    if req.role not in ROLES:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {list(ROLES.keys())}")
    smtp = _get_smtp_config()
    if not smtp["configured"]:
        raise HTTPException(status_code=503, detail="SMTP is not configured. Add SMTP_HOST, SMTP_USER, and SMTP_PASSWORD to web/.env")

    # Create user without password (pending invite)
    from auth import user_exists
    if user_exists(req.username):
        raise HTTPException(status_code=409, detail="Username already exists")
    create_user_pending(req.username, req.role)

    # Generate invitation token
    invite_token = create_invite_token(req.username)
    if not invite_token:
        raise HTTPException(status_code=500, detail="Failed to create invitation token")

    # Build invitation URL
    base_url = str(request.base_url).rstrip("/").replace("http://", "https://", 1)
    invite_url = f"{base_url}/set-password?token={invite_token}"

    # Send email
    ok = send_invite_email(
        username=req.username,
        invite_url=invite_url,
        smtp_host=smtp["host"],
        smtp_port=smtp["port"],
        smtp_user=smtp["user"],
        smtp_password=smtp["password"],
        smtp_from=smtp["from_addr"],
        smtp_use_tls=smtp["use_tls"],
    )

    if not ok:
        return {"ok": True, "email_sent": False, "warning": "User created but email could not be sent. Check SMTP configuration."}

    return {"ok": True, "email_sent": True, "username": req.username}


@app.post("/api/auth/invite/resend/{username}")
async def api_resend_invite(username: str, request: Request, session: dict = Depends(require_role("superuser"))):
    """Resend invitation email to an existing user."""
    smtp = _get_smtp_config()
    if not smtp["configured"]:
        raise HTTPException(status_code=503, detail="SMTP is not configured")

    from auth import user_exists
    if not user_exists(username):
        raise HTTPException(status_code=404, detail="User not found")

    invite_token = create_invite_token(username)
    if not invite_token:
        raise HTTPException(status_code=500, detail="Failed to create invitation token")

    base_url = str(request.base_url).rstrip("/").replace("http://", "https://", 1)
    invite_url = f"{base_url}/set-password?token={invite_token}"

    ok = send_invite_email(
        username=username,
        invite_url=invite_url,
        smtp_host=smtp["host"],
        smtp_port=smtp["port"],
        smtp_user=smtp["user"],
        smtp_password=smtp["password"],
        smtp_from=smtp["from_addr"],
        smtp_use_tls=smtp["use_tls"],
    )

    if not ok:
        raise HTTPException(status_code=500, detail="Failed to send email. Check SMTP configuration.")
    return {"ok": True, "email_sent": True}


@app.get("/api/auth/invite/validate")
async def api_validate_invite(token: str = Query(...)):
    """Validate an invitation token (public endpoint)."""
    username = validate_invite_token(token)
    if not username:
        return {"valid": False}
    return {"valid": True, "username": username}


@app.post("/api/auth/invite/set-password")
async def api_set_password_from_invite(req: SetPasswordRequest):
    """Set password using an invitation token (public endpoint)."""
    pwd_error = validate_password(req.password)
    if pwd_error:
        raise HTTPException(status_code=400, detail=pwd_error)
    username = set_password_from_invite(req.token, req.password)
    if not username:
        raise HTTPException(status_code=400, detail="Invalid or expired invitation link")
    return {"ok": True, "username": username}


@app.get("/set-password")
async def set_password_page():
    """Serve the set-password page."""
    return FileResponse(SCRIPT_DIR / "static" / "set_password.html")


# ---------------------------------------------------------------------------
# Access request routes (self-registration)
# ---------------------------------------------------------------------------

class AccessRequestBody(BaseModel):
    email: str
    full_name: str
    institution: str
    department: str = ""
    profile_url: str = ""
    reason: str = ""


@app.post("/api/auth/request-access")
async def api_request_access(req: AccessRequestBody, request: Request):
    """Public endpoint: submit an access request (must be UNINOVIS email).
    If the email is found in the directory, an invitation is sent automatically."""
    email = req.email.strip().lower()
    # Validate email domain
    domain_err = validate_uninovis_email(email)
    if domain_err:
        raise HTTPException(status_code=400, detail=domain_err)
    if not req.full_name.strip():
        raise HTTPException(status_code=400, detail="Full name is required")

    # Check if user already exists
    if user_exists(email):
        raise HTTPException(status_code=409, detail="An account with this email already exists. Use 'Forgot password' if you need to reset it.")

    # Check if the email is in the directory
    directory_match = _check_directory_email(email)

    if directory_match:
        # Auto-create user and send invitation
        try:
            create_user(email, role="student", provisional=True)
            invite_token = create_invite_token(email)
            if invite_token:
                smtp = _get_smtp_config()
                if smtp["configured"]:
                    base_url = str(request.base_url).rstrip("/").replace("http://", "https://", 1)
                    invite_url = f"{base_url}/set-password?token={invite_token}"
                    try:
                        send_invite_email(
                            username=email,
                            invite_url=invite_url,
                            smtp_host=smtp["host"],
                            smtp_port=smtp["port"],
                            smtp_user=smtp["user"],
                            smtp_password=smtp["password"],
                            from_addr=smtp["from_addr"],
                        )
                    except Exception:
                        pass
            return {"ok": True, "message": f"Your email was found in the UNINOVIS directory ({directory_match}). An invitation has been sent to {email}."}
        except Exception:
            pass  # Fall through to manual request

    # Not in directory — create a pending access request
    reason = req.reason.strip() if req.reason.strip() else "Sign up request"
    ok = create_access_request(
        email, req.full_name.strip(), req.institution.strip(),
        department=getattr(req, 'department', ''),
        profile_url=getattr(req, 'profile_url', ''),
        reason=reason,
    )
    if not ok:
        raise HTTPException(status_code=409, detail="A request with this email already exists")
    return {"ok": True, "message": "Access request submitted. A UNINOVIS administrator will review your request and you will receive an email when approved."}


@app.post("/api/auth/forgot-password")
async def api_forgot_password(request: Request, body: dict = None):
    """Public endpoint: request a password reset link.
    Checks if the email/username exists and sends an invite token that
    allows setting a new password via the existing set-password flow."""
    if body is None:
        body = await request.json()
    email = body.get("email", "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Email or username is required")

    # Check if user exists
    if not user_exists(email):
        # Don't reveal whether the user exists — always show success
        return {"ok": True, "message": "If this email is registered, a password reset link has been sent."}

    # Generate an invite token (reuses the existing invite mechanism)
    invite_token = create_invite_token(email)
    if not invite_token:
        return {"ok": True, "message": "If this email is registered, a password reset link has been sent."}

    # Try to send the reset email
    smtp = _get_smtp_config()
    if smtp["configured"]:
        base_url = str(request.base_url).rstrip("/").replace("http://", "https://", 1)
        reset_url = f"{base_url}/set-password?token={invite_token}"
        try:
            send_invite_email(
                username=email,
                invite_url=reset_url,
                smtp_host=smtp["host"],
                smtp_port=smtp["port"],
                smtp_user=smtp["user"],
                smtp_password=smtp["password"],
                from_addr=smtp["from_addr"],
                subject="UNINOVIS — Password Reset",
            )
        except Exception:
            pass  # Don't reveal email sending failures

    return {"ok": True, "message": "If this email is registered, a password reset link has been sent."}


@app.get("/api/auth/access-requests")
async def api_list_requests(
    status: Optional[str] = Query(None),
    session: dict = Depends(require_role("superuser")),
):
    """List access requests (superuser only)."""
    return list_access_requests(status)


@app.post("/api/auth/access-requests/{email}/approve")
async def api_approve_request(
    email: str,
    request: Request,
    role: str = Query("user"),
    session: dict = Depends(require_role("superuser")),
):
    """Approve an access request and optionally send invitation email."""
    if role not in ROLES:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {list(ROLES.keys())}")
    ok = approve_access_request(email, role)
    if not ok:
        raise HTTPException(status_code=404, detail="Pending request not found")

    # Try to send invitation email
    email_sent = False
    smtp = _get_smtp_config()
    if smtp["configured"]:
        invite_token = create_invite_token(email)
        if invite_token:
            base_url = str(request.base_url).rstrip("/").replace("http://", "https://", 1)
            invite_url = f"{base_url}/set-password?token={invite_token}"
            email_sent = send_invite_email(
                username=email,
                invite_url=invite_url,
                smtp_host=smtp["host"],
                smtp_port=smtp["port"],
                smtp_user=smtp["user"],
                smtp_password=smtp["password"],
                smtp_from=smtp["from_addr"],
                smtp_use_tls=smtp["use_tls"],
            )

    return {"ok": True, "email_sent": email_sent}


@app.post("/api/auth/access-requests/{email}/reject")
async def api_reject_request(email: str, session: dict = Depends(require_role("superuser"))):
    """Reject an access request."""
    ok = reject_access_request(email)
    if not ok:
        raise HTTPException(status_code=404, detail="Pending request not found")
    return {"ok": True}


# ---------------------------------------------------------------------------
# Study mode endpoints (superuser only)
# ---------------------------------------------------------------------------

@app.get("/api/study/status")
async def study_status(session: dict = Depends(require_role("superuser"))):
    """Get study mode status and condition assignments."""
    from auth import is_study_mode, _load_users, STUDY_CONDITIONS
    users = _load_users()
    counts = {c: 0 for c in STUDY_CONDITIONS}
    assignments = []
    for uname, udata in users.items():
        cond = udata.get("study_condition")
        if cond:
            counts[cond] = counts.get(cond, 0) + 1
            assignments.append({"username": uname, "condition": cond})
    return {
        "enabled": is_study_mode(),
        "conditions": STUDY_CONDITIONS,
        "counts": counts,
        "total_assigned": sum(counts.values()),
        "assignments": assignments,
    }


@app.post("/api/study/enable")
async def study_enable(session: dict = Depends(require_role("superuser"))):
    """Enable study mode (random transparency assignment)."""
    from auth import STUDY_CONFIG_FILE
    config = {"enabled": True, "study_id": "UMA-IA-TOMMI-STUDY-001",
              "conditions": ["black_box", "grey_box", "crystal_box"],
              "description": "Effect of AI Transparency Levels on User Trust"}
    with open(STUDY_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2)
    return {"ok": True, "study_mode": True}


@app.post("/api/study/disable")
async def study_disable(session: dict = Depends(require_role("superuser"))):
    """Disable study mode (users choose transparency freely)."""
    from auth import STUDY_CONFIG_FILE
    config = {"enabled": False}
    with open(STUDY_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2)
    return {"ok": True, "study_mode": False}


@app.post("/api/study/reset")
async def study_reset(session: dict = Depends(require_role("superuser"))):
    """Clear all study condition assignments (for a new study run)."""
    from auth import _load_users, _save_users
    users = _load_users()
    cleared = 0
    for udata in users.values():
        for key in ["study_condition", "study_participant", "study_id",
                     "email_domain", "study_completed"]:
            if key in udata:
                del udata[key]
        cleared += 1
    _save_users(users)
    return {"ok": True, "cleared": cleared}


@app.post("/api/study/enroll/{username}")
async def study_enroll(username: str, session: dict = Depends(require_role("superuser"))):
    """Enroll a specific user as a study participant."""
    from auth import enroll_study_participant
    result = enroll_study_participant(username)
    if not result:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True, **result}


@app.post("/api/study/enroll-bulk")
async def study_enroll_bulk(usernames: list[str], session: dict = Depends(require_role("superuser"))):
    """Enroll multiple users as study participants."""
    from auth import enroll_study_participant
    results = []
    for u in usernames:
        r = enroll_study_participant(u)
        results.append({"username": u, **(r or {"error": "not found"})})
    return {"ok": True, "enrolled": results}


@app.get("/api/study/queries")
async def study_queries(session: dict = Depends(require_auth)):
    """Return the predefined study queries."""
    queries_file = SCRIPT_DIR / "data" / "study_queries.json"
    if not queries_file.exists():
        raise HTTPException(status_code=404, detail="Study queries not configured")
    with open(queries_file, "r", encoding="utf-8") as f:
        return json.load(f)


class StudyQuestionnaireRequest(BaseModel):
    agent_id: str
    query_number: int
    stias_1: int
    stias_2: int
    stias_3: int
    understanding: int
    reliance: int


@app.post("/api/study/questionnaire")
async def study_questionnaire(req: StudyQuestionnaireRequest, session: dict = Depends(require_auth)):
    """Save per-query questionnaire answers to the study log."""
    from auth import get_study_info
    info = get_study_info(session["username"])
    if not info:
        raise HTTPException(status_code=403, detail="Not a study participant")

    # Find the agent's study_log.jsonl
    agent_instance = runner.get_agent_instance(req.agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not found")

    import os
    from sys import path as sys_path
    agent_dir = getattr(agent_instance, '_agent_dir', None)
    if not agent_dir:
        raise HTTPException(status_code=500, detail="Agent directory not found")

    study_log_path = os.path.join(agent_dir, "data", "study_log.jsonl")

    # Import StudyLogger from the agent's base
    sys_path_entry = os.path.join(os.path.dirname(agent_dir), "base")
    if sys_path_entry not in sys_path:
        sys_path.insert(0, sys_path_entry)
    from badges import StudyLogger

    questionnaire = {
        "stias_1": req.stias_1,
        "stias_2": req.stias_2,
        "stias_3": req.stias_3,
        "understanding": req.understanding,
        "reliance": req.reliance,
    }

    updated = StudyLogger.update_questionnaire(
        study_log_path=study_log_path,
        study_id=info["study_id"],
        query_number=req.query_number,
        questionnaire=questionnaire,
    )

    if not updated:
        raise HTTPException(status_code=404, detail="Study log entry not found for this query")
    return {"ok": True}


@app.post("/api/study/comparison")
async def study_comparison(request: Request, session: dict = Depends(require_auth)):
    """Save within-subjects comparison study results to a JSONL log."""
    import os
    data = await request.json()
    log_path = SCRIPT_DIR / "data" / "study_comparison_log.jsonl"
    entry = {
        "timestamp": datetime.utcnow().isoformat(),
        "username": session["username"],
        **data,
    }
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    return {"ok": True}


@app.post("/api/study/complete")
async def study_complete(session: dict = Depends(require_auth)):
    """Mark the current user's study as completed."""
    from auth import mark_study_completed, get_study_info
    info = get_study_info(session["username"])
    if not info:
        raise HTTPException(status_code=403, detail="Not a study participant")
    mark_study_completed(session["username"])
    return {"ok": True, "study_id": info["study_id"]}


@app.get("/study")
async def study_page():
    """Serve the study interface page."""
    return FileResponse(SCRIPT_DIR / "static" / "study.html")


class ChatRequest(BaseModel):
    """Request para enviar un mensaje a un agente"""
    agent_id: str
    message: str
    session_id: Optional[str] = None


class ChatResponse(BaseModel):
    """Respuesta de un agente"""
    response: str
    session_id: str


class AgentResponse(BaseModel):
    """Información de un agente"""
    id: str
    name: str
    agent_type: str
    description: str
    welcome_message: str
    example_queries: list[str]
    rag_approach: str = "context_preserving"
    show_history: bool = True
    show_description: bool = False
    transparency_level: str = "black_box"
    transparency_type: str = ""
    prompt_level: str = "stringent"
    decision_trace: str = "black_box"
    reliability_cues: str = ""


@app.get("/")
async def root():
    """Serve the UNINOVIS intranet as the landing page"""
    return FileResponse(SCRIPT_DIR / "static" / "intranet.html")


@app.get("/intranet")
async def intranet_page():
    """Alias for the intranet landing page"""
    return FileResponse(SCRIPT_DIR / "static" / "intranet.html")


@app.get("/rag-study")
async def rag_study_redirect():
    """Redirect old rag-study URL to rag-study2"""
    from starlette.responses import RedirectResponse
    return RedirectResponse("/rag-study2")


@app.get("/rag-study2")
async def rag_study2_page():
    """RAG Reliability Study — query classification comparison"""
    return FileResponse(SCRIPT_DIR / "static" / "rag_study2.html")


@app.get("/rag-study2/chat")
async def rag_study2_chat_page():
    """RAG Reliability Study — agent chat interface"""
    return FileResponse(SCRIPT_DIR / "static" / "rag_study2_chat.html")


@app.get("/all-agents")
async def all_agents_page():
    """Serve the unified all-agents landing page"""
    return FileResponse(SCRIPT_DIR / "static" / "all_agents.html")


@app.get("/research-explorers")
async def research_explorers_page():
    """Serve the Research Explorers landing page (no auth required)"""
    return FileResponse(SCRIPT_DIR / "static" / "research_explorers.html")


@app.get("/quiron")
async def quiron_public_page():
    """Serve the public Quirón research assistant page (no auth required)"""
    return FileResponse(SCRIPT_DIR / "static" / "quiron.html")

@app.get("/api/public-agent/quiron/fuentes")
async def quiron_list_sources():
    """List all source documents available to Quirón."""
    meta_path = Path(__file__).parent.parent / "agents" / "quiron" / "data" / "papers_metadata.json"
    if meta_path.exists():
        with open(meta_path, "r", encoding="utf-8") as f:
            papers = json.load(f)
        return {"total": len(papers), "papers": papers}
    # Fallback: list files in docs/
    docs_dir = Path(__file__).parent.parent / "agents" / "quiron" / "data" / "docs"
    if docs_dir.exists():
        files = sorted(f for f in docs_dir.iterdir() if f.suffix.lower() == ".pdf")
        papers = [{"filename": f.name, "title": f.stem.replace("_", " ")} for f in files]
        return {"total": len(papers), "papers": papers}
    return {"total": 0, "papers": []}


@app.get("/responsible-ai")
async def responsible_ai_public_page():
    """Serve the public Responsible AI Research Assistant page (no auth required)"""
    return FileResponse(SCRIPT_DIR / "static" / "responsible_ai.html")


@app.get("/health-wellbeing")
async def health_wellbeing_public_page():
    """Serve the public Health & Wellbeing Systems Research Explorer page (no auth required)"""
    return FileResponse(SCRIPT_DIR / "static" / "health_wellbeing.html")


@app.get("/uninovis-uma")
async def uninovis_uma_page():
    """Serve the UNINOVIS-UMA AI Tools landing page"""
    return FileResponse(SCRIPT_DIR / "static" / "uninovis_uma.html")


@app.get("/uninovis-uma/agent")
async def uninovis_uma_agent():
    """Serve an agent within the UNINOVIS-UMA wrapper"""
    return FileResponse(SCRIPT_DIR / "static" / "uma_agent.html")


@app.get("/uninovis-uma/algoria")
async def uninovis_uma_algoria():
    """Serve the Algoria DB agent page within UNINOVIS-UMA"""
    return FileResponse(SCRIPT_DIR / "static" / "uma_algoria.html")


@app.get("/uninovis-uma/proyectos-europeos")
async def uninovis_uma_proyectos():
    """Serve the European Projects agent page within UNINOVIS-UMA"""
    return FileResponse(SCRIPT_DIR / "static" / "uma_proyectos.html")


@app.get("/creative-agents")
async def creative_agents_page():
    """Serve the Creative Agents landing page"""
    return FileResponse(SCRIPT_DIR / "static" / "creative_agents.html")


@app.get("/creative-agents/sonic-composer")
async def sonic_composer_page():
    """Serve the Sonic Composer chat page"""
    return FileResponse(SCRIPT_DIR / "static" / "sonic_composer.html")


@app.get("/creative-agents/sonic-composer/help")
async def sonic_composer_help():
    """Serve the Sonic Composer help page"""
    return FileResponse(SCRIPT_DIR / "static" / "sonic_composer_howto.html")


@app.get("/creative-agents/sonic-composer-v2")
async def sonic_composer2_page():
    """Serve the Sonic Composer v2 chat page"""
    return FileResponse(SCRIPT_DIR / "static" / "sonic_composer2.html")


@app.get("/creative-agents/sonic-composer-v2/help")
async def sonic_composer2_help():
    """Serve the Sonic Composer v2 help page"""
    return FileResponse(SCRIPT_DIR / "static" / "sonic_composer2_howto.html")


@app.get("/tutores-virtuales")
async def tutores_virtuales_page():
    """Serve the Virtual Tutors landing page"""
    return FileResponse(SCRIPT_DIR / "static" / "tutores_virtuales.html")


@app.get("/tutores-virtuales/help/eulalia")
async def help_eulalia():
    """Help page for the LALI tutor"""
    return FileResponse(SCRIPT_DIR / "static" / "help_eulalia.html")


@app.get("/tutores-virtuales/eulalia")
async def tutores_lali(request: Request, moodle_token: str = Query(None)):
    """Serve the LALI tutor page.

    Auth is handled client-side: the frontend checks localStorage for a token
    and redirects to login if needed. The API endpoints verify auth independently.
    If moodle_token is present, it's passed through so the frontend can store it.
    """
    return FileResponse(SCRIPT_DIR / "static" / "eulalia.html")


@app.get("/tutores-virtuales/eulalia/widgets/{widget_name}")
async def lali_widget(widget_name: str):
    """Serve LALI tutor interactive widgets"""
    safe_name = widget_name.replace("/", "").replace("..", "")
    path = SCRIPT_DIR / "static" / "eulalia" / "widgets" / f"{safe_name}.html"
    if not path.is_file():
        return JSONResponse({"error": "Widget not found"}, status_code=404)
    return FileResponse(path)


# ── Tutores Virtuales: auth system ─────────────────────────────────────

from auth_tutores import (
    authenticate as tutores_authenticate,
    get_session as tutores_get_session,
    logout as tutores_logout,
    is_docente as tutores_is_docente,
    list_users as tutores_list_users,
    create_user as tutores_create_user,
    delete_user as tutores_delete_user,
    update_user as tutores_update_user,
    bulk_create as tutores_bulk_create,
    change_password as tutores_change_password,
    create_invite as tutores_create_invite,
    validate_invite as tutores_validate_invite,
    activate_with_invite as tutores_activate_with_invite,
)


def _get_tutores_token(request: Request) -> str:
    """Extract tutores auth token from request."""
    token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
    if not token:
        token = request.query_params.get("token", "")
    return token


@app.get("/tutores-virtuales/login")
async def tutores_login_page():
    return FileResponse(SCRIPT_DIR / "static" / "login_tutores.html")


class TutoresLoginRequest(BaseModel):
    username: str
    password: str


@app.post("/api/tutores/login")
async def tutores_login(req: TutoresLoginRequest):
    result = tutores_authenticate(req.username, req.password)
    if not result:
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
    return result


class TutoresChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@app.post("/api/tutores/change-password")
async def tutores_do_change_password(req: TutoresChangePasswordRequest, request: Request):
    """Change password for the current tutores user."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session:
        raise HTTPException(status_code=401, detail="No autenticado")
    if len(req.new_password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    ok = tutores_change_password(session["username"], req.old_password, req.new_password)
    if not ok:
        raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")
    return {"ok": True}


@app.post("/api/tutores/logout")
async def tutores_do_logout(request: Request):
    token = _get_tutores_token(request)
    if token:
        tutores_logout(token)
    return {"ok": True}


@app.get("/api/tutores/auth-level")
async def tutores_auth_level(request: Request):
    """Check tutores auth level."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session:
        return {"level": "none", "authenticated": False}
    if tutores_is_docente(session):
        return {"level": "docente", "authenticated": True,
                "username": session["username"], "nombre": session.get("nombre", "")}
    return {"level": "estudiante", "authenticated": True,
            "username": session["username"], "nombre": session.get("nombre", "")}


@app.get("/api/tutores/users")
async def tutores_get_users(request: Request):
    """List tutores users (docente only)."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session or not tutores_is_docente(session):
        raise HTTPException(status_code=403, detail="Solo el profesorado puede ver el listado de usuarios")
    return tutores_list_users()


class TutoresCreateUserRequest(BaseModel):
    username: str
    password: str = ""
    role: str = "estudiante"
    nombre: str = ""


@app.post("/api/tutores/users")
async def tutores_add_user(req: TutoresCreateUserRequest, request: Request):
    """Add a tutores user (docente only)."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session or not tutores_is_docente(session):
        raise HTTPException(status_code=403, detail="Solo el profesorado puede crear usuarios")
    try:
        ok = tutores_create_user(req.username, req.password or None, req.role, req.nombre, activo=True)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not ok:
        raise HTTPException(status_code=409, detail=f"El usuario '{req.username}' ya existe")
    return {"ok": True, "username": req.username}


@app.delete("/api/tutores/users/{username}")
async def tutores_remove_user(username: str, request: Request):
    """Delete a tutores user (docente only)."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session or not tutores_is_docente(session):
        raise HTTPException(status_code=403, detail="Solo el profesorado puede eliminar usuarios")
    if not tutores_delete_user(username):
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"ok": True}


class TutoresBulkRequest(BaseModel):
    users: list[dict]


@app.post("/api/tutores/users/bulk")
async def tutores_bulk_add(req: TutoresBulkRequest, request: Request):
    """Bulk create tutores users from a list (docente only)."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session or not tutores_is_docente(session):
        raise HTTPException(status_code=403, detail="Solo el profesorado puede crear usuarios")
    result = tutores_bulk_create(req.users)
    return result


@app.post("/api/tutores/users/{username}/invite")
async def tutores_gen_invite(username: str, request: Request):
    """Generate an invitation link for a user (docente only)."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session or not tutores_is_docente(session):
        raise HTTPException(status_code=403, detail="Solo el profesorado puede generar enlaces")
    token = tutores_create_invite(username)
    if not token:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    # Build full URL
    host = request.headers.get("host", "localhost")
    scheme = request.headers.get("x-forwarded-proto", "https")
    link = f"{scheme}://{host}/tutores-virtuales/activar?token={token}"
    return {"link": link, "token": token, "username": username}


@app.get("/tutores-virtuales/activar")
async def tutores_activation_page():
    """Serve the account activation page."""
    return FileResponse(SCRIPT_DIR / "static" / "activar_tutores.html")


class TutoresActivateRequest(BaseModel):
    token: str
    password: str


@app.post("/api/tutores/activate")
async def tutores_activate(req: TutoresActivateRequest):
    """Activate an account by setting password via invite token."""
    if len(req.password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    username = tutores_activate_with_invite(req.token, req.password)
    if not username:
        raise HTTPException(status_code=400, detail="Enlace inválido o caducado. Solicita un nuevo enlace al profesorado.")
    return {"ok": True, "username": username}


@app.get("/api/tutores/invite/validate")
async def tutores_validate_invite_endpoint(token: str = Query(...)):
    """Validate an invite token and return the username."""
    username = tutores_validate_invite(token)
    if not username:
        raise HTTPException(status_code=400, detail="Enlace inválido o caducado")
    return {"username": username}


# ── LALI tutor: course data API ────────────────────────────────────────

_LALI_DIR = Path(__file__).parent.parent / "agents" / "eulalia"


@app.get("/api/public-agent/eulalia/docentes")
async def lali_get_docentes():
    """Get teaching staff info (public)."""
    path = _LALI_DIR / "docentes.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


@app.get("/api/public-agent/eulalia/progreso-temas")
async def lali_get_progreso_temas(request: Request):
    """Get student progress across all themes (requires auth)."""
    tutores_token = _get_tutores_token(request)
    session = tutores_get_session(tutores_token) if tutores_token else None
    if not session:
        return {"temas": [], "tema_recomendado": 1}
    username = session.get("username", "")
    if not username:
        return {"temas": [], "tema_recomendado": 1}
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "agents"))
        from tutor_fonetica_base.progreso_alumno import (
            progreso_todos_temas, tema_recomendado, set_progress_dir
        )
        set_progress_dir(_LALI_DIR / "progress")
        return {
            "temas": progreso_todos_temas(username),
            "tema_recomendado": tema_recomendado(username),
            "username": username,
        }
    except Exception as e:
        return {"temas": [], "tema_recomendado": 1, "error": str(e)}


@app.post("/api/public-agent/eulalia/solicitar-revision")
async def lali_solicitar_revision(request: Request):
    """A student requests a professor review of an AI-evaluated answer."""
    body = await request.json()
    pregunta = body.get("pregunta", "")
    respuesta_alumno = body.get("respuesta_alumno", "")
    evaluacion_ia = body.get("evaluacion_ia", "")
    widget = body.get("widget", "")
    email = body.get("email", "")

    # Try to get username if authenticated
    tutores_token = _get_tutores_token(request)
    session = tutores_get_session(tutores_token) if tutores_token else None
    username = session.get("username", email or "anónimo") if session else (email or "anónimo")

    revision = {
        "timestamp": datetime.now().isoformat(),
        "username": username,
        "email": email,
        "widget": widget,
        "pregunta": pregunta,
        "respuesta_alumno": respuesta_alumno,
        "evaluacion_ia": evaluacion_ia,
        "estado": "pendiente"
    }

    revision_path = _LALI_DIR / "revisiones_pendientes.json"
    try:
        existing = json.loads(revision_path.read_text(encoding="utf-8")) if revision_path.exists() else []
    except Exception:
        existing = []
    existing.append(revision)
    revision_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"status": "ok", "message": "Solicitud de revisión registrada"}


@app.get("/api/public-agent/eulalia/revisiones")
async def lali_get_revisiones(request: Request):
    """Get all pending review requests (professor only)."""
    revision_path = _LALI_DIR / "revisiones_pendientes.json"
    try:
        revisiones = json.loads(revision_path.read_text(encoding="utf-8")) if revision_path.exists() else []
    except Exception:
        revisiones = []
    return {"revisiones": revisiones}


@app.post("/api/public-agent/eulalia/responder-revision")
async def lali_responder_revision(request: Request):
    """Professor responds to a review request."""
    body = await request.json()
    idx = body.get("index")
    respuesta_profesor = body.get("respuesta", "")
    decision = body.get("decision", "")  # "aceptar" or "mantener"

    revision_path = _LALI_DIR / "revisiones_pendientes.json"
    try:
        revisiones = json.loads(revision_path.read_text(encoding="utf-8")) if revision_path.exists() else []
    except Exception:
        revisiones = []

    if idx is None or idx < 0 or idx >= len(revisiones):
        raise HTTPException(status_code=400, detail="Índice de revisión inválido")

    revisiones[idx]["estado"] = "respondida"
    revisiones[idx]["decision"] = decision
    revisiones[idx]["respuesta_profesor"] = respuesta_profesor
    revisiones[idx]["fecha_respuesta"] = datetime.now().isoformat()

    revision_path.write_text(json.dumps(revisiones, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "ok"}


# ── LALI tutor: student consultations to professor ───────────────────

_LALI_CONSULTAS_PATH = _LALI_DIR / "consultas_profesor.json"


# ── LALI tutor: student practice progress ────────────────────────────

_LALI_PROGRESO_PATH = _LALI_DIR / "data" / "progreso_practica.json"


def _load_progreso():
    try:
        if _LALI_PROGRESO_PATH.exists():
            return json.loads(_LALI_PROGRESO_PATH.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def _save_progreso(data):
    _LALI_PROGRESO_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


@app.post("/api/public-agent/eulalia/progreso-practica")
async def lali_save_progreso(request: Request):
    """Save a student's practice progress for a specific exercise."""
    body = await request.json()
    ejercicio = body.get("ejercicio", "")
    score = body.get("score", 0)
    max_score = body.get("max_score", 0)
    detalles = body.get("detalles", {})

    if not ejercicio:
        return JSONResponse({"error": "Falta el ID del ejercicio"}, status_code=400)

    tutores_token = _get_tutores_token(request)
    session = tutores_get_session(tutores_token) if tutores_token else None
    username = session.get("username", "anónimo") if session else "anónimo"

    all_progress = _load_progreso()
    if username not in all_progress:
        all_progress[username] = {}

    pct = round(score / max_score * 100) if max_score > 0 else 0
    prev = all_progress[username].get(ejercicio, {})
    force = body.get("force", False)

    # Only update if better score, not yet saved, or forced overwrite
    if not prev or pct > prev.get("score", 0) or force:
        all_progress[username][ejercicio] = {
            "score": pct,
            "completed": pct >= 75,
            "raw_score": score,
            "max_score": max_score,
            "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "detalles": detalles,
        }
        _save_progreso(all_progress)

    return {"ok": True, "score": pct}


@app.get("/api/public-agent/eulalia/progreso-practica")
async def lali_get_progreso(request: Request):
    """Get practice progress. Students see only their own; professors see all."""
    tutores_token = _get_tutores_token(request)
    session = tutores_get_session(tutores_token) if tutores_token else None
    username = session.get("username", "anónimo") if session else "anónimo"
    role = session.get("role", "student") if session else "student"

    all_progress = _load_progreso()

    if role in ("admin", "editor", "superuser"):
        # Professor: return all students
        return {"role": "profesor", "progreso": all_progress}
    else:
        # Student: return only their own
        return {"role": "estudiante", "username": username, "progreso": all_progress.get(username, {})}


@app.post("/api/public-agent/eulalia/consulta-profesor")
async def lali_consulta_profesor(request: Request):
    """A student sends a question/consultation to the professor from a widget."""
    body = await request.json()
    consulta = body.get("consulta", "").strip()
    widget = body.get("widget", "")
    contexto = body.get("contexto", "")  # optional: what the student was doing
    email = body.get("email", "")

    if not consulta:
        return JSONResponse({"error": "La consulta no puede estar vacía"}, status_code=400)

    tutores_token = _get_tutores_token(request)
    session = tutores_get_session(tutores_token) if tutores_token else None
    username = session.get("username", email or "anónimo") if session else (email or "anónimo")

    entry = {
        "timestamp": datetime.now().isoformat(),
        "username": username,
        "email": email,
        "widget": widget,
        "contexto": contexto,
        "tipo": body.get("tipo", ""),
        "ubicacion": body.get("ubicacion", ""),
        "consulta": consulta,
        "estado": "pendiente",
    }

    try:
        existing = json.loads(_LALI_CONSULTAS_PATH.read_text(encoding="utf-8")) if _LALI_CONSULTAS_PATH.exists() else []
    except Exception:
        existing = []
    existing.append(entry)
    _LALI_CONSULTAS_PATH.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"status": "ok", "message": "Consulta enviada al profesor"}


@app.get("/api/public-agent/eulalia/consultas-profesor")
async def lali_get_consultas():
    """Get all student consultations (professor view)."""
    try:
        consultas = json.loads(_LALI_CONSULTAS_PATH.read_text(encoding="utf-8")) if _LALI_CONSULTAS_PATH.exists() else []
    except Exception:
        consultas = []
    return {"consultas": consultas}


@app.post("/api/public-agent/eulalia/responder-consulta")
async def lali_responder_consulta(request: Request):
    """Professor responds to a student consultation."""
    body = await request.json()
    idx = body.get("index")
    respuesta = body.get("respuesta", "")

    try:
        consultas = json.loads(_LALI_CONSULTAS_PATH.read_text(encoding="utf-8")) if _LALI_CONSULTAS_PATH.exists() else []
    except Exception:
        consultas = []

    if idx is None or idx < 0 or idx >= len(consultas):
        raise HTTPException(status_code=400, detail="Índice de consulta inválido")

    consultas[idx]["estado"] = "respondida"
    consultas[idx]["respuesta_profesor"] = respuesta
    consultas[idx]["fecha_respuesta"] = datetime.now().isoformat()

    _LALI_CONSULTAS_PATH.write_text(json.dumps(consultas, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "ok"}


@app.post("/api/public-agent/eulalia/registrar-actividad")
async def lali_registrar_actividad(request: Request):
    """Register a student activity in a theme (concept viewed, exercise done, etc.)."""
    tutores_token = _get_tutores_token(request)
    session = tutores_get_session(tutores_token) if tutores_token else None
    if not session:
        raise HTTPException(status_code=401, detail="No autenticado")
    username = session.get("username", "")
    if not username:
        raise HTTPException(status_code=401, detail="No autenticado")
    body = await request.json()
    tema_id = body.get("tema_id")
    tipo = body.get("tipo")
    detalle = body.get("detalle", "")
    if not tema_id or not tipo:
        raise HTTPException(status_code=400, detail="Faltan tema_id o tipo")
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "agents"))
        from tutor_fonetica_base.progreso_alumno import registrar_actividad_tema, set_progress_dir
        set_progress_dir(_LALI_DIR / "progress")
        registrar_actividad_tema(username, int(tema_id), tipo, detalle)
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/public-agent/eulalia/autoeval-progreso")
async def lali_get_autoeval_progreso(
    request: Request,
    tema_id: int = Query(None),
):
    """Get detailed self-assessment progress for a student."""
    tutores_token = _get_tutores_token(request)
    session = tutores_get_session(tutores_token) if tutores_token else None
    if not session:
        return {"error": "No autenticado"}
    username = session.get("username", "")
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "agents"))
        from tutor_fonetica_base.autoevaluacion import (
            set_autoeval_dir, progreso_tema as ae_progreso_tema,
            progreso_global as ae_progreso_global
        )
        set_autoeval_dir(_LALI_DIR / "autoevaluacion")
        if tema_id:
            return ae_progreso_tema(username, tema_id)
        else:
            return ae_progreso_global(username)
    except Exception as e:
        return {"error": str(e)}


# ── LALI tutor: Moodle SSO ─────────────────────────────────────────────

# Secreto compartido para firmar los enlaces desde Moodle.
# Cámbialo por uno único. Se pone también en el bloque HTML de Moodle.
_MOODLE_SSO_SECRET = os.environ.get("LALI_MOODLE_SECRET", "cambiar-este-secreto-compartido")
_MOODLE_SSO_EXPIRY = 300  # 5 minutos de validez


@app.get("/api/public-agent/eulalia/moodle-login")
async def lali_moodle_login(
    request: Request,
    user: str = Query(...),
    ts: str = Query(...),
    sig: str = Query(...),
):
    """SSO desde Moodle. Verifica firma HMAC, crea sesión, redirige al tutor.

    URL generada por el bloque HTML de Moodle:
        /api/public-agent/eulalia/moodle-login?user=EMAIL&ts=TIMESTAMP&sig=HMAC
    """
    import hashlib
    import hmac
    import time as _time

    # 1. Verificar que el enlace no ha caducado
    try:
        timestamp = int(ts)
    except ValueError:
        raise HTTPException(status_code=400, detail="Timestamp inválido")
    if abs(_time.time() - timestamp) > _MOODLE_SSO_EXPIRY:
        raise HTTPException(status_code=403, detail="El enlace ha caducado. Vuelve a Moodle y haz clic de nuevo.")

    # 2. Verificar firma HMAC
    mensaje = f"{user}:{ts}"
    firma_esperada = hmac.new(
        _MOODLE_SSO_SECRET.encode(), mensaje.encode(), hashlib.sha256
    ).hexdigest()
    if not hmac.compare_digest(sig, firma_esperada):
        print(f"[MOODLE-SSO] FIRMA INVÁLIDA user={repr(user)} ts={repr(ts)}")
        raise HTTPException(status_code=403, detail="Firma inválida. Acceso denegado.")

    # 3. Buscar el usuario en tutores_users.json
    username = user.strip().lower()
    from auth_tutores import _load_users as tutores_load_users
    users = tutores_load_users()

    if username not in users:
        raise HTTPException(
            status_code=403,
            detail=f"El usuario '{username}' no está registrado en Eulalia. Contacta con el profesor para que te dé de alta."
        )

    # 4. Crear sesión directamente (sin requerir contraseña)
    import secrets as _secrets
    from auth_tutores import _sessions
    token = _secrets.token_hex(32)
    user_role = users[username].get("role", "estudiante")
    _sessions[token] = {
        "username": username,
        "role": user_role,
        "created": _time.time(),
    }

    # 5. Redirigir al tutor con el token en la URL
    redirect_url = f"/tutores-virtuales/eulalia?moodle_token={token}"
    return RedirectResponse(url=redirect_url, status_code=302)


# ── LALI tutor: Proyecto de investigación (Tema 6) ─────────────────

_LALI_PROYECTOS_DIR = _LALI_DIR / "proyectos"
_LALI_PROYECTOS_DIR.mkdir(exist_ok=True)


def _load_proyecto(project_id: str) -> dict | None:
    path = _LALI_PROYECTOS_DIR / f"{project_id}.json"
    if not path.exists():
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_proyecto(proyecto: dict):
    path = _LALI_PROYECTOS_DIR / f"{proyecto['id']}.json"
    proyecto["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(proyecto, f, ensure_ascii=False, indent=2)


def _get_user_email(request: Request) -> str | None:
    token = _get_tutores_token(request)
    session = tutores_get_session(token) if token else None
    return session.get("username") if session else None


@app.get("/api/public-agent/eulalia/proyectos")
async def lali_list_proyectos(request: Request):
    """List projects for the current user (member or docente)."""
    email = _get_user_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="Autenticación requerida")

    session = tutores_get_session(_get_tutores_token(request))
    is_prof = tutores_is_docente(session) if session else False

    proyectos = []
    for f in _LALI_PROYECTOS_DIR.glob("*.json"):
        try:
            p = json.loads(f.read_text(encoding="utf-8"))
            if is_prof or email in p.get("grupo", []):
                # Don't send full chat in list view
                summary = {k: v for k, v in p.items() if k != "chat"}
                summary["n_chat"] = len(p.get("chat", []))
                proyectos.append(summary)
        except Exception:
            pass

    proyectos.sort(key=lambda x: x.get("updated", ""), reverse=True)
    return {"proyectos": proyectos}


@app.post("/api/public-agent/eulalia/proyectos")
async def lali_create_proyecto(request: Request):
    """Create a new project. The creator is automatically added to the group."""
    email = _get_user_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="Autenticación requerida")
    body = await request.json()
    titulo = body.get("titulo", "").strip()
    if not titulo:
        raise HTTPException(status_code=400, detail="Falta el título")

    import uuid
    project_id = str(uuid.uuid4())[:8]
    proyecto = {
        "id": project_id,
        "titulo": titulo,
        "grupo": [email],
        "created": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "fases": {
            "introduccion": {"texto": "", "ultima_edicion": "", "editado_por": ""},
            "metodologia": {"texto": "", "ultima_edicion": "", "editado_por": ""},
            "resultados": {"texto": "", "ultima_edicion": "", "editado_por": ""},
            "discusion": {"texto": "", "ultima_edicion": "", "editado_por": ""},
        },
        "chat": [],
    }
    _save_proyecto(proyecto)
    return {"ok": True, "id": project_id}


@app.get("/api/public-agent/eulalia/proyectos/{project_id}")
async def lali_get_proyecto(project_id: str, request: Request):
    """Get a project by ID. Only group members and docentes can access."""
    email = _get_user_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="Autenticación requerida")
    proyecto = _load_proyecto(project_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    session = tutores_get_session(_get_tutores_token(request))
    is_prof = tutores_is_docente(session) if session else False
    if not is_prof and email not in proyecto.get("grupo", []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")
    return proyecto


@app.put("/api/public-agent/eulalia/proyectos/{project_id}/fases/{fase_id}")
async def lali_update_fase(project_id: str, fase_id: str, request: Request):
    """Update the text of a project phase."""
    email = _get_user_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="Autenticación requerida")
    proyecto = _load_proyecto(project_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if email not in proyecto.get("grupo", []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")
    if fase_id not in ("introduccion", "metodologia", "resultados", "discusion"):
        raise HTTPException(status_code=400, detail="Fase no válida")

    body = await request.json()
    texto = body.get("texto", "")
    proyecto["fases"][fase_id] = {
        "texto": texto,
        "ultima_edicion": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "editado_por": email,
    }
    _save_proyecto(proyecto)
    return {"ok": True}


@app.post("/api/public-agent/eulalia/proyectos/{project_id}/invitar")
async def lali_invite_member(project_id: str, request: Request):
    """Invite a member to the project."""
    email = _get_user_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="Autenticación requerida")
    proyecto = _load_proyecto(project_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if email not in proyecto.get("grupo", []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")

    body = await request.json()
    new_email = body.get("email", "").strip().lower()
    if not new_email or "@" not in new_email:
        raise HTTPException(status_code=400, detail="Email no válido")
    if new_email in proyecto["grupo"]:
        raise HTTPException(status_code=409, detail="Ya es miembro del proyecto")

    proyecto["grupo"].append(new_email)
    _save_proyecto(proyecto)
    return {"ok": True}


@app.post("/api/public-agent/eulalia/proyectos/{project_id}/chat")
async def lali_project_chat(project_id: str, request: Request):
    """Send a chat message within the project context. Uses LLM with phase context."""
    email = _get_user_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="Autenticación requerida")
    proyecto = _load_proyecto(project_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if email not in proyecto.get("grupo", []):
        raise HTTPException(status_code=403, detail="No tienes acceso a este proyecto")

    body = await request.json()
    mensaje = body.get("mensaje", "").strip()
    fase = body.get("fase", "")
    contexto_fase = body.get("contexto_fase", "")

    if not mensaje:
        raise HTTPException(status_code=400, detail="Mensaje vacío")

    # Build LLM prompt with project context
    agent = runner.get_agent("eulalia")
    if not agent:
        raise HTTPException(status_code=500, detail="Agente no disponible")

    agent_instance = runner._load_agent_module("eulalia")
    model = agent_instance.model

    fase_nombres = {"introduccion": "Introducción", "metodologia": "Metodología",
                    "resultados": "Resultados", "discusion": "Discusión y conclusiones"}
    fase_nombre = fase_nombres.get(fase, fase)

    system_prompt = (
        "Eres Eulalia, tutora de Lingüística Aplicada a la Logopedia. "
        "Estás ayudando a un grupo de estudiantes con su proyecto de investigación fonológica. "
        f"El grupo está trabajando en la fase: {fase_nombre}.\n\n"
        "Tu rol es ser un interlocutor crítico: ayuda al estudiante a mejorar su trabajo, "
        "señala inconsistencias, sugiere mejoras, pero NO hagas el trabajo por ellos.\n\n"
        "IMPORTANTE: Tus respuestas son generadas por IA y pueden contener errores. "
        "Los estudiantes deben contrastar con otras fuentes.\n\n"
        "Responde en español. Sé conciso (máximo 200 palabras)."
    )

    if contexto_fase:
        system_prompt += f"\n\nContenido actual de la fase '{fase_nombre}' escrito por el grupo:\n{contexto_fase}"

    # Recent chat history for context
    recent = [m for m in proyecto.get("chat", []) if m.get("fase") == fase][-6:]
    messages = [{"role": "system", "content": system_prompt}]
    for m in recent:
        messages.append({"role": "user", "content": m["mensaje"]})
        if m.get("respuesta"):
            messages.append({"role": "assistant", "content": m["respuesta"]})
    messages.append({"role": "user", "content": mensaje})

    try:
        response = agent_instance.client.chat.complete(model=model, messages=messages)
        respuesta = response.choices[0].message.content
    except Exception as e:
        respuesta = f"Error al consultar la IA: {str(e)}"

    # Save to project
    chat_entry = {
        "autor": email,
        "mensaje": mensaje,
        "respuesta": respuesta,
        "fase": fase,
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "response_type": "llm",
    }
    proyecto["chat"].append(chat_entry)
    _save_proyecto(proyecto)

    return {"respuesta": respuesta, "response_type": "llm"}


@app.get("/api/public-agent/eulalia/temas")
async def lali_get_temas():
    """Get course topics structure (public)."""
    path = _LALI_DIR / "temas.json"
    if not path.exists():
        return {"temas": []}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


@app.get("/api/public-agent/eulalia/contenido-seccion")
async def lali_get_contenido_seccion(
    doc: str = Query(..., description="Nombre del archivo markdown (ej: Tema1_Introduccion.md)"),
    seccion: str = Query("", description="Título de la sección a extraer (ej: '1. Fonología y fonética')"),
):
    """Devuelve el contenido de una sección de un documento markdown del curso.

    Si seccion está vacío, devuelve todo el documento.
    El contenido se devuelve en markdown crudo (el frontend lo renderiza con marked.js).
    """
    import re as _re

    # Sanitize: no path traversal
    safe_doc = Path(doc).name
    doc_path = _LALI_DIR / "data" / "docs" / safe_doc
    if not doc_path.exists():
        raise HTTPException(status_code=404, detail=f"Documento no encontrado: {safe_doc}")

    with open(doc_path, "r", encoding="utf-8") as f:
        contenido = f.read()

    if not seccion:
        return {"doc": safe_doc, "seccion": "", "contenido": contenido}

    # Buscar la sección por título (## o ### o ####)
    # Los encabezados pueden tener HTML inline como <a id="..."></a>
    titulo_escaped = _re.escape(seccion.strip())
    # Permitir HTML tags opcionales entre los # y el título
    pattern = r'^(#{2,4})\s+(?:<[^>]*>\s*)*' + titulo_escaped + r'\s*$'
    match = _re.search(pattern, contenido, _re.MULTILINE)
    if not match:
        # Intentar búsqueda parcial (sin número de sección)
        titulo_sin_num = _re.sub(r'^\d+[\.\d]*\.?\s*', '', seccion.strip())
        titulo_escaped2 = _re.escape(titulo_sin_num)
        pattern2 = r'^(#{2,4})\s+(?:<[^>]*>\s*)*(?:\d+[\.\d]*\.?\s*)?' + titulo_escaped2 + r'\s*$'
        match = _re.search(pattern2, contenido, _re.MULTILINE | _re.IGNORECASE)

    if not match:
        return {"doc": safe_doc, "seccion": seccion, "contenido": "", "error": "Sección no encontrada"}

    nivel = len(match.group(1))  # 2 para ##, 3 para ###
    inicio = match.start()

    # Buscar el final: siguiente heading del mismo nivel o superior
    rest = contenido[match.end():]
    end_pattern = r'^#{2,' + str(nivel) + r'}\s+'
    end_match = _re.search(end_pattern, rest, _re.MULTILINE)
    if end_match:
        fin = match.end() + end_match.start()
    else:
        fin = len(contenido)

    seccion_texto = contenido[inicio:fin].strip()
    return {"doc": safe_doc, "seccion": seccion, "contenido": seccion_texto}


# ── LALI tutor: transcription config API ──────────────────────────────

_LALI_CONFIG_PATH = _LALI_DIR / "transcripcion_config.json"


@app.post("/api/public-agent/eulalia/evaluar-definicion")
async def lali_evaluar_definicion(request: Request):
    """Evaluate a student's definition of a concept using LLM with a closed rubric."""
    from llm_client import LLMClient

    body = await request.json()
    concepto = body.get("concepto", "")
    definicion_alumno = body.get("definicion", "")
    rubrica = body.get("rubrica", [])
    definicion_referencia = body.get("referencia", "")

    if not concepto or not definicion_alumno or not rubrica:
        return JSONResponse({"error": "Faltan campos obligatorios"}, status_code=400)

    criterios_text = "\n".join(
        f"- Criterio {i+1}: {c['descripcion']}" for i, c in enumerate(rubrica)
    )

    prompt = f"""Eres un tutor de fonología que evalúa definiciones. Dirígete al estudiante de tú, con un tono cercano y constructivo (por ejemplo: "Tu definición...", "Debes repasar...", "Has captado bien...").

DEFINICIÓN DE REFERENCIA (del temario):
{definicion_referencia}

DEFINICIÓN DEL ESTUDIANTE:
{definicion_alumno}

RÚBRICA — Evalúa cada criterio como true (cumplido) o false (no cumplido):
{criterios_text}

INSTRUCCIONES:
- Evalúa si la definición del estudiante cubre cada criterio, no si usa las mismas palabras exactas.
- Sé justo y generoso: si el concepto está expresado con otras palabras, con sinónimos o de forma implícita pero clara, marca true. Por ejemplo, si el criterio pide que asocie una subdisciplina a la "fase de producción" y el estudiante dice "cómo se producen los sonidos", eso CUMPLE el criterio aunque no use la palabra "fase".
- Solo marca false si el concepto realmente falta o es incorrecto.
- IMPORTANTE: Asegúrate de que el valor "cumplido" (true/false) es coherente con tu comentario. Si tu comentario dice que la respuesta es correcta, el valor debe ser true.
- Responde ÚNICAMENTE con un JSON válido, sin texto adicional, con esta estructura:
{{
  "criterios": [
    {{"cumplido": true/false, "comentario": "breve explicación"}},
    ...
  ],
  "comentario_general": "retroalimentación constructiva en 1-2 frases"
}}"""

    try:
        client = LLMClient()
        response = client.chat.complete(
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1024,
        )
        llm_text = response.choices[0].message.content.strip()
        # Extract JSON from response (handle markdown code blocks)
        if "```" in llm_text:
            import re
            match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", llm_text, re.DOTALL)
            if match:
                llm_text = match.group(1)
        result = json.loads(llm_text)
        return result
    except json.JSONDecodeError:
        return JSONResponse({"error": "La LLM no devolvió JSON válido", "raw": llm_text}, status_code=502)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


# ── LALI tutor: syllable frequency data ──────────────────────────────

@app.get("/api/public-agent/eulalia/silabas-frecuencia")
async def lali_get_silabas_frecuencia():
    """Return phonological syllable frequency data."""
    sil_path = _LALI_DIR / "data" / "silabas_frecuencia.json"
    if not sil_path.exists():
        return []
    with open(sil_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ── LALI tutor: transcription exercises (programmatic) ───────────────

@app.get("/api/public-agent/eulalia/ejercicio-transcripcion")
async def lali_ejercicio_transcripcion(nivel: int = Query(1, ge=1, le=5), items: int = Query(5, ge=1, le=15)):
    """Generate a transcription exercise with random words/phrases for levels 1-5."""
    import random
    import sys

    ej_path = _LALI_DIR / "data" / "ejercicios_transcripcion.json"
    if not ej_path.exists():
        raise HTTPException(status_code=500, detail="Banco de ejercicios no encontrado")

    with open(ej_path, "r", encoding="utf-8") as f:
        banco = json.load(f)

    nivel_key = f"nivel{nivel}"
    if nivel_key not in banco:
        raise HTTPException(status_code=400, detail=f"Nivel {nivel} no existe")

    nivel_data = banco[nivel_key]

    if nivel == 5:
        # Level 5: phrases with pre-computed solutions
        frases = nivel_data["frases"]
        seleccion = random.sample(frases, min(items, len(frases)))
        ejercicios = [{"palabra": f["frase"], "solucion": f["solucion"]} for f in seleccion]
    else:
        # Levels 1-4: words, transcribe programmatically
        palabras = nivel_data["palabras"]
        seleccion = random.sample(palabras, min(items, len(palabras)))

        base_dir = Path(__file__).parent.parent / "agents" / "tutor_fonetica_base"
        if str(base_dir) not in sys.path:
            sys.path.insert(0, str(base_dir))
        from transcriptor import transcribir_palabra

        ejercicios = []
        for palabra in seleccion:
            t = transcribir_palabra(palabra)
            if isinstance(t, tuple):
                continue
            ejercicios.append({"palabra": palabra, "solucion": t})

    return {
        "nivel": nivel,
        "nombre": nivel_data["nombre"],
        "descripcion": nivel_data["descripcion"],
        "ejercicios": ejercicios
    }



@app.get("/api/public-agent/eulalia/ejercicio-informe-errores")
async def lali_ejercicio_informe_errores(
    num_errores: int = Query(5, ge=3, le=10),
    seed: int = Query(None),
):
    """Generate a phonological error report exercise dynamically."""
    import random
    import sys

    base_dir = Path(__file__).parent.parent / "agents" / "tutor_fonetica_base"
    if str(base_dir) not in sys.path:
        sys.path.insert(0, str(base_dir))
    from transcriptor import transcribir
    from errores_fonologicos.generador import generar_errores

    rng = random.Random(seed)

    # Bank of sentences (ortographic)
    ORACIONES = [
        "La mariposa bonita vuela por el jardín grande con flores",
        "Mi hermana pequeña compra tres platos blancos en la tienda",
        "El niño pequeño tiene un perro grande que corre mucho por la plaza",
        "El tren sale a las tres de la tarde con mi amigo Pedro",
        "Dame el plato grande de la mesa blanca con la cuchara",
        "Los pájaros cantan en las ramas del árbol grande del parque",
        "La casa de mi abuela está en la calle estrecha del pueblo grande",
        "El zapato rojo de la entrada está roto y sucio como un trapo",
        "La profesora explica la lección a los estudiantes del primer grupo",
        "El gato blanco duerme en la cama grande de mi hermana",
        "Quiero un vaso de agua fresca con un poco de limón",
        "El médico receta unas pastillas para el dolor de cabeza",
        "La chica rubia lleva una chaqueta verde y zapatos negros",
        "Mi padre trabaja en la fábrica grande que está cerca del río",
        "El perro del vecino ladra cuando viene el cartero a las tres",
        "La enfermera atiende con cuidado al paciente que está dormido",
        "Los niños juegan en la plaza grande del barrio con sus amigos",
        "El anciano camina despacio bajo un cielo gris y frío",
        "Mi madre prepara la comida mientras escucha la radio en la cocina",
        "El maestro pone un examen difícil sobre gramática y fonología",
        "Las flores del jardín crecen con la lluvia de primavera",
        "El autobús sale del centro a las cuatro de la tarde",
        "La ventana grande de la clase da a la calle principal del pueblo",
        "El carpintero trabaja con madera de roble para hacer la mesa",
        "Los alumnos practican la transcripción fonológica en clase de lingüística",
        "La biblioteca del campus está abierta los fines de semana por la tarde",
        "El río que pasa por la ciudad tiene un puente de piedra muy antiguo",
        "Mi vecina tiene tres gatos blancos y un perro grande y negro",
        "El director del colegio habla con los padres sobre el plan de estudios",
        "La farmacia de la esquina cierra los domingos por la tarde",
        "El electricista repara el cable roto que está cerca de la puerta",
        "La cocinera prepara un plato especial con verduras frescas del mercado",
        "El pintor trabaja en un cuadro grande con colores muy brillantes",
        "La secretaria escribe una carta urgente para el director del centro",
        "Los bomberos apagan el fuego que empezó en la cocina del restaurante",
        "Mi primo estudia medicina en la universidad de la capital del país",
        "El fontanero arregla el grifo roto del cuarto de baño del segundo piso",
        "La cantante tiene una voz bonita que gusta a todo el público",
        "El cartero reparte las cartas por la mañana en todas las calles del barrio",
        "La policía busca al ladrón que robó en la tienda de la esquina",
    ]

    oracion = rng.choice(ORACIONES)
    correcta = transcribir(oracion)
    if isinstance(correcta, tuple):
        # Fallback if transcription fails
        oracion = "El niño pequeño tiene un perro grande"
        correcta = transcribir(oracion)

    # Apply errors using patient profiles
    palabras_ort = oracion.split()
    palabras_fon = correcta.strip().split(' ')

    # Patient profiles: consistent error tendencies
    PERFILES = {
        'lenicion': {
            'tipos': ['sonorizacion', 'omision_coda', 'simplificacion_ataque',
                      'simplificacion_nucleo', 'omision_silaba_atona', 'lenicion'],
            'desc': 'Perfil de lenición: tendencia a debilitar sonidos'
        },
        'forticion': {
            'tipos': ['ensordecimiento', 'forticion', 'posteriorización',
                      'simplificacion_ataque'],
            'desc': 'Perfil de fortición: tendencia a reforzar sonidos'
        },
        'mixto': {
            'tipos': ['sonorizacion', 'ensordecimiento', 'adelantamiento',
                      'simplificacion_ataque', 'omision_coda', 'omision_silaba_atona',
                      'simplificacion_nucleo'],
            'desc': 'Perfil mixto'
        }
    }

    perfil_nombre = rng.choice(list(PERFILES.keys()))
    perfil = PERFILES[perfil_nombre]
    tipos_preferidos = perfil['tipos']

    # First pass: apply phonological errors to content words
    producidas = []
    todos_errores = []
    errores_ritmicos = []
    palabras_con_error = set()

    # Function words that can be omitted in lenition profile
    _PALABRAS_ATONAS = {'el', 'la', 'los', 'las', 'un', 'una', 'unos', 'unas',
                        'de', 'en', 'con', 'por', 'a', 'y', 'que', 'del', 'al'}

    for idx, pf in enumerate(palabras_fon):
        pf_clean = pf.strip()
        if not pf_clean:
            continue
        n_sil = pf_clean.count('.') + 1
        is_content = n_sil >= 2 or (len(pf_clean.replace("ˈ", "").replace(".", "")) > 3)
        word_ort_lower = palabras_ort[idx].lower() if idx < len(palabras_ort) else ''

        # Lenition profile: occasional omission of function words
        if (perfil_nombre == 'lenicion'
                and word_ort_lower in _PALABRAS_ATONAS
                and not is_content
                and len(todos_errores) < num_errores
                and rng.random() < 0.25):
            word_ort = palabras_ort[idx] if idx < len(palabras_ort) else '?'
            todos_errores.append({
                'tipo': 'omision_palabra_atona',
                'categoria': 'palabra',
                'palabra': word_ort,
                'descripcion': f"Omisión de palabra átona: '{word_ort}'"
            })
            palabras_con_error.add(idx)
            # Don't add this word to producidas (it's omitted)
            continue

        if is_content and len(todos_errores) < num_errores:
            # Words with complex onsets (tɾ, pɾ, bɾ, kɾ, gɾ, pl, bl, fl, etc.) get 2 errors
            has_complex_onset = any(c in pf_clean for c in ['ɾ', 'l'] if pf_clean.find(c) > 0)
            n_err_word = 2 if has_complex_onset and rng.random() < 0.6 else 1
            n_err_word = min(n_err_word, num_errores - len(todos_errores))

            # Use profile-consistent error types
            prod, errs = generar_errores(pf_clean, tipos=tipos_preferidos,
                                          num_errores=n_err_word, seed=rng.randint(0, 100000))

            # Filter out tonic syllable omissions (very rare)
            errs = [e for e in errs if e.get('tipo') != 'omision_silaba_tonica']

            if errs and prod != pf_clean:
                word_ort = palabras_ort[idx] if idx < len(palabras_ort) else '?'
                palabras_con_error.add(idx)
                for e in errs:
                    e['palabra'] = word_ort
                    if e['tipo'] in ('sonorizacion', 'ensordecimiento', 'adelantamiento',
                                     'posteriorización', 'nasalizacion', 'desnasalizacion',
                                     'lenicion', 'forticion'):
                        e['categoria'] = 'sistemico'
                    elif e['tipo'] in ('omision_ataque', 'simplificacion_ataque',
                                       'simplificacion_nucleo', 'omision_coda'):
                        e['categoria'] = 'silaba'
                    elif e['tipo'] in ('asimilacion_regresiva', 'asimilacion_progresiva',
                                       'metátesis', 'omision_silaba_atona'):
                        e['categoria'] = 'palabra'
                    else:
                        e['categoria'] = 'sistemico'
                    todos_errores.append(e)
                producidas.append(prod)
                continue

        producidas.append(pf_clean)

    # Second pass: rhythmic errors on words WITHOUT phonological errors
    palabras_sin_error = [i for i in range(len(producidas))
                          if i not in palabras_con_error
                          and producidas[i].strip()
                          and producidas[i] != '#'
                          and (producidas[i].count('.') + 1) >= 2]
    rng.shuffle(palabras_sin_error)
    max_ritmicos = rng.randint(1, max(1, min(3, len(palabras_sin_error))))

    for idx in palabras_sin_error[:max_ritmicos]:
        pf = producidas[idx]
        if not pf or pf == '#':
            continue
        word_ort = palabras_ort[idx] if idx < len(palabras_ort) else '?'
        n_sil = pf.count('.') + 1
        r = rng.random()

        if r < 0.33 and n_sil >= 3:
            sils = pf.split('.')
            split_at = rng.randint(1, len(sils) - 1)
            producidas[idx] = '.'.join(sils[:split_at]) + ' # ' + '.'.join(sils[split_at:])
            errores_ritmicos.append(f"Pausa indebida dentro de '{word_ort}'")
        elif r < 0.66 and n_sil >= 2:
            sils = pf.split('.')
            first_sil = sils[0].replace('ˈ', '')
            producidas[idx] = first_sil + '.' + pf
            errores_ritmicos.append(f"Titubeo (repetición de sílaba) en '{word_ort}'")
        else:
            producidas.insert(idx + 1, '#')
            errores_ritmicos.append(f"Pausa indebida después de '{word_ort}'")

    producido_str = '/ ' + ' '.join(producidas) + ' /'

    # Calculate PFC and PPC
    fonemas_correcta = [c for c in correcta.replace('/', '').replace(' ', '').replace('.', '').replace('ˈ', '') if c.isalpha() or c in 'θʝʧɲɾ']
    total_fonemas = len(fonemas_correcta)
    fonemas_erroneos = len(todos_errores)
    pfc = round((total_fonemas - fonemas_erroneos) / total_fonemas * 100) if total_fonemas > 0 else 100

    total_palabras = len([p for p in palabras_ort if len(p) > 2])
    palabras_con_error = len(set(e.get('palabra', '') for e in todos_errores))
    ppc = round((total_palabras - palabras_con_error) / total_palabras * 100) if total_palabras > 0 else 100

    ritmicos = '. '.join(errores_ritmicos) if errores_ritmicos else 'Ninguno'

    # Format errors for the frontend
    errores_fmt = []
    for e in todos_errores:
        subtipo_map = {
            'sonorizacion': 'Sonorización', 'ensordecimiento': 'Ensordecimiento',
            'adelantamiento': 'Adelantamiento', 'posteriorización': 'Posteriorización',
            'nasalizacion': 'Nasalización', 'desnasalizacion': 'Desnasalización',
            'lenicion': 'Lenición', 'forticion': 'Fortición',
            'omision_ataque': 'Omisión de ataque', 'simplificacion_ataque': 'Simplificación de ataque',
            'simplificacion_nucleo': 'Simplificación de núcleo', 'omision_coda': 'Omisión de coda',
            'asimilacion_regresiva': 'Asimilación regresiva', 'asimilacion_progresiva': 'Asimilación progresiva',
            'metátesis': 'Metátesis', 'omision_silaba_atona': 'Omisión de sílaba átona',
            'omision_silaba_tonica': 'Omisión de sílaba tónica',
            'omision_palabra_atona': 'Omisión de palabra átona',
        }
        errores_fmt.append({
            'tipo': e.get('categoria', 'sistemico'),
            'subtipo': subtipo_map.get(e['tipo'], e['tipo']),
            'detalle': e.get('descripcion', '') + " en '" + e.get('palabra', '') + "'"
        })

    return {
        'ortografia': oracion,
        'producido': producido_str,
        'correcta': correcta,
        'pfc': pfc,
        'ppc': ppc,
        'errores': errores_fmt,
        'ritmicos': ritmicos
    }


@app.get("/api/public-agent/eulalia/evaluacion-config")
async def lali_evaluacion_config():
    """Get the evaluation configuration (grading weights, criteria)."""
    config_path = _LALI_DIR / "evaluacion_config.json"
    if not config_path.exists():
        return {}
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


@app.get("/api/public-agent/eulalia/ejercicio-transcripcion-fonetica")
async def lali_ejercicio_transcripcion_fonetica(
    nivel: int = Query(1, ge=1, le=7),
    items: int = Query(5, ge=1, le=15),
    exclude: str = Query("", description="Comma-separated list of words to exclude"),
):
    """Generate a phonetic transcription exercise for levels 1-7."""
    import random
    import sys

    ej_path = _LALI_DIR / "data" / "ejercicios_transcripcion_fonetica.json"
    if not ej_path.exists():
        raise HTTPException(status_code=500, detail="Banco de ejercicios fonéticos no encontrado")

    with open(ej_path, "r", encoding="utf-8") as f:
        banco = json.load(f)

    nivel_key = f"nivel{nivel}"
    if nivel_key not in banco:
        raise HTTPException(status_code=400, detail=f"Nivel {nivel} no existe")

    nivel_data = banco[nivel_key]

    # All levels: words/phrases, transcribe programmatically
    palabras = nivel_data["palabras"]

    # Exclude already-used words in this session
    if exclude:
        excl_set = set(e.strip() for e in exclude.split(",") if e.strip())
        disponibles = [p for p in palabras if p not in excl_set]
        # If all exhausted, reset
        if len(disponibles) < items:
            disponibles = palabras
    else:
        disponibles = palabras

    seleccion = random.sample(disponibles, min(items, len(disponibles)))

    base_dir = Path(__file__).parent.parent / "agents" / "tutor_fonetica_base"
    if str(base_dir) not in sys.path:
        sys.path.insert(0, str(base_dir))
    from transcriptor import transcripcion_fonetica_palabra, transcribir_palabra

    ejercicios = []
    for texto in seleccion:
        palabras_txt = texto.split()
        partes_fon = []
        partes_fonet = []
        error = False
        for i, p in enumerate(palabras_txt):
            is_start = (i == 0)
            prev_last = None
            next_first = None
            if i > 0:
                prev_t = transcribir_palabra(palabras_txt[i - 1])
                if isinstance(prev_t, str) and prev_t:
                    prev_last = prev_t.replace("ˈ", "").replace(".", "")[-1]
            if i + 1 < len(palabras_txt):
                next_t = transcribir_palabra(palabras_txt[i + 1])
                if isinstance(next_t, str) and next_t:
                    next_first = next_t.replace("ˈ", "").replace(".", "")[0]

            tf = transcripcion_fonetica_palabra(p, is_utterance_start=is_start,
                                                 prev_word_last_fonema=prev_last,
                                                 next_word_first_fonema=next_first)
            tl = transcribir_palabra(p)
            if isinstance(tf, tuple) or isinstance(tl, tuple):
                error = True
                break
            partes_fon.append(tl)
            partes_fonet.append(tf)

        if not error:
            ejercicios.append({
                "palabra": texto,
                "fonologica": "/ " + ".".join(partes_fon) + " /",
                "solucion": "[" + ".".join(partes_fonet) + "]"
            })

    return {
        "nivel": nivel,
        "nombre": nivel_data["nombre"],
        "descripcion": nivel_data["descripcion"],
        "pista": nivel_data.get("pista", ""),
        "ejercicios": ejercicios
    }


@app.get("/api/public-agent/eulalia/transcribir")
async def lali_transcribir(texto: str = Query(..., max_length=200)):
    """Transcribe a word or short phrase phonologically (programmatic, no LLM)."""
    import sys
    base_dir = Path(__file__).parent.parent / "agents" / "tutor_fonetica_base"
    if str(base_dir) not in sys.path:
        sys.path.insert(0, str(base_dir))
    from transcriptor import transcribir, transcribir_palabra

    texto = texto.strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Texto vacío")

    # Check for inappropriate content
    from base_tutor import _contiene_palabra_inapropiada
    if _contiene_palabra_inapropiada(texto):
        raise HTTPException(status_code=400, detail="Contenido no permitido")

    resultado = transcribir(texto)
    if isinstance(resultado, tuple):
        return {"ok": False, "error": resultado[1]}

    # Per-word detail
    import re
    palabras = re.findall(r"[a-záéíóúüñ]+", texto.lower())
    detalle = []
    for p in palabras:
        t = transcribir_palabra(p)
        if isinstance(t, tuple):
            detalle.append({"palabra": p, "error": t[1]})
        elif t:
            detalle.append({"palabra": p, "transcripcion": t})

    return {"ok": True, "texto": texto, "transcripcion": resultado, "detalle": detalle}


# ── LALI tutor: empathy challenge (shared contributions) ─────────────

_LALI_EMPATIA_PATH = _LALI_DIR / "data" / "retos_empatia.json"


@app.get("/api/public-agent/eulalia/retos-empatia")
async def lali_get_retos_empatia():
    """Return all shared empathy contributions."""
    if not _LALI_EMPATIA_PATH.exists():
        return []
    with open(_LALI_EMPATIA_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


@app.post("/api/public-agent/eulalia/retos-empatia")
async def lali_post_reto_empatia(request: Request):
    """Save a new empathy contribution (shared across students)."""
    body = await request.json()
    perfil = body.get("perfil", "").strip()
    situacion = body.get("situacion", "").strip()
    tiene_dificultad = body.get("tiene_dificultad")  # True / False
    explicacion = body.get("explicacion", "").strip()
    autor = body.get("autor", "").strip() or "Anónimo"

    if not perfil or not situacion or tiene_dificultad is None or not explicacion:
        return JSONResponse({"error": "Faltan campos obligatorios"}, status_code=400)

    entry = {
        "perfil": perfil,
        "situacion": situacion,
        "tiene_dificultad": bool(tiene_dificultad),
        "explicacion": explicacion,
        "autor": autor,
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    # Load existing
    entries = []
    if _LALI_EMPATIA_PATH.exists():
        with open(_LALI_EMPATIA_PATH, "r", encoding="utf-8") as f:
            entries = json.load(f)

    entries.append(entry)

    with open(_LALI_EMPATIA_PATH, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    return {"ok": True, "total": len(entries)}


@app.put("/api/public-agent/eulalia/retos-empatia/{idx}")
async def lali_put_reto_empatia(idx: int, request: Request):
    """Edit an existing empathy contribution by index."""
    body = await request.json()

    entries = []
    if _LALI_EMPATIA_PATH.exists():
        with open(_LALI_EMPATIA_PATH, "r", encoding="utf-8") as f:
            entries = json.load(f)

    if idx < 0 or idx >= len(entries):
        raise HTTPException(status_code=400, detail="Índice inválido")

    # Only allow editing if same author
    autor_req = body.get("autor", "").strip()
    if entries[idx].get("autor", "") != autor_req:
        raise HTTPException(status_code=403, detail="Solo puedes editar tus propias contribuciones")

    if body.get("situacion"):
        entries[idx]["situacion"] = body["situacion"].strip()
    if body.get("explicacion"):
        entries[idx]["explicacion"] = body["explicacion"].strip()
    if body.get("tiene_dificultad") is not None:
        entries[idx]["tiene_dificultad"] = bool(body["tiene_dificultad"])
    entries[idx]["editado"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    with open(_LALI_EMPATIA_PATH, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    return {"ok": True}


@app.get("/api/public-agent/eulalia/transcripcion-config")
async def lali_get_transcripcion_config():
    """Get current transcription config (public, read-only)."""
    if not _LALI_CONFIG_PATH.exists():
        return {}
    with open(_LALI_CONFIG_PATH, "r", encoding="utf-8") as f:
        raw = json.load(f)
    return {
        "sibilantes": raw.get("sibilantes", {}).get("valor", "distinguidor"),
        "nasalizacion_vocalica": raw.get("nasalizacion_vocalica", {}).get("valor", False),
        "s_coda": raw.get("s_coda", {}).get("valor", "sibilante"),
    }


class TranscripcionConfigUpdate(BaseModel):
    sibilantes: Optional[str] = None
    nasalizacion_vocalica: Optional[bool] = None
    s_coda: Optional[str] = None


@app.put("/api/agent/eulalia/transcripcion-config")
async def lali_update_transcripcion_config(body: TranscripcionConfigUpdate, request: Request):
    """Update transcription config (requires docente role)."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session:
        raise HTTPException(status_code=401, detail="No autenticado. Inicia sesión para modificar la configuración.")
    if not tutores_is_docente(session):
        raise HTTPException(status_code=403, detail="No tienes permisos de edición. Solo el profesorado puede modificar la configuración.")

    # Validate values
    valid_sibilantes = {"distinguidor", "seseo", "ceceo"}
    valid_s_coda = {"sibilante", "aspiracion", "omision", "alargamiento"}

    if body.sibilantes and body.sibilantes not in valid_sibilantes:
        raise HTTPException(status_code=400, detail=f"Valor inválido para sibilantes. Opciones: {', '.join(valid_sibilantes)}")
    if body.s_coda and body.s_coda not in valid_s_coda:
        raise HTTPException(status_code=400, detail=f"Valor inválido para s_coda. Opciones: {', '.join(valid_s_coda)}")

    # Read current config
    with open(_LALI_CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    # Update only provided fields
    if body.sibilantes is not None:
        cfg["sibilantes"]["valor"] = body.sibilantes
    if body.nasalizacion_vocalica is not None:
        cfg["nasalizacion_vocalica"]["valor"] = body.nasalizacion_vocalica
    if body.s_coda is not None:
        cfg["s_coda"]["valor"] = body.s_coda

    with open(_LALI_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
        f.write("\n")

    # Reload config in transcriptor (if loaded)
    try:
        agent = runner.get_agent("eulalia")
        if agent and hasattr(agent, '_instance'):
            import importlib
            transcriptor = importlib.import_module("transcriptor")
            transcriptor.recargar_config()
    except Exception:
        pass  # Non-critical

    username = session.get("username", "unknown")
    if ENABLE_LOGGING:
        logger = _get_agent_logger("eulalia")
        logger.info(f"CONFIG_UPDATE by {username}: sibilantes={body.sibilantes}, nasalizacion={body.nasalizacion_vocalica}, s_coda={body.s_coda}")

    return {"status": "ok", "config": {
        "sibilantes": cfg["sibilantes"]["valor"],
        "nasalizacion_vocalica": cfg["nasalizacion_vocalica"]["valor"],
        "s_coda": cfg["s_coda"]["valor"],
    }}


@app.get("/api/eulalia/auth-level")
async def lali_auth_level(request: Request):
    """Check if current user has editor access to LALI tutor config (tutores auth)."""
    session = tutores_get_session(_get_tutores_token(request))
    if not session:
        return {"level": "none", "authenticated": False}
    if tutores_is_docente(session):
        return {"level": "editor", "authenticated": True, "username": session.get("username")}
    return {"level": "student", "authenticated": True, "username": session.get("username")}


# ── LALI tutor / Fonética: análisis acústico con Praat ─────────────────

_PRAAT_OUTPUT_DIR = SCRIPT_DIR / "temp" / "praat"
_PRAAT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Max audio file size: 10 MB
_MAX_AUDIO_SIZE = 10 * 1024 * 1024
_ALLOWED_AUDIO_TYPES = {
    "audio/wav", "audio/wave", "audio/x-wav",
    "audio/mpeg", "audio/mp3",
    "audio/ogg", "audio/flac",
    "audio/webm",
}


@app.post("/api/public-agent/analyze-audio")
async def analyze_audio(
    request: Request,
    file: UploadFile = FastFile(...),
    mostrar_formantes: bool = Query(True),
    mostrar_pitch: bool = Query(False),
):
    """Analiza un archivo de audio y genera espectrograma + datos acústicos.

    Público (no requiere autenticación). Usado por tutores de fonética.

    Returns:
        JSON con URLs de imágenes y datos de análisis.
    """
    # Validate file type
    content_type = file.content_type or ""
    if content_type not in _ALLOWED_AUDIO_TYPES and not file.filename.endswith((".wav", ".mp3", ".ogg", ".flac", ".webm")):
        raise HTTPException(status_code=400, detail="Formato de audio no soportado. Usa WAV, MP3, OGG o FLAC.")

    # Read and validate size
    audio_data = await file.read()
    if len(audio_data) > _MAX_AUDIO_SIZE:
        raise HTTPException(status_code=400, detail="El archivo es demasiado grande (máximo 10 MB).")
    if len(audio_data) == 0:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")

    # Save to temp file
    import tempfile
    import uuid
    file_id = uuid.uuid4().hex[:12]
    safe_name = "".join(c for c in Path(file.filename or "audio").stem if c.isalnum() or c in "-_")[:30]
    temp_wav = _PRAAT_OUTPUT_DIR / f"{file_id}_{safe_name}.wav"

    # If not WAV, convert using parselmouth (it handles multiple formats)
    temp_input = _PRAAT_OUTPUT_DIR / f"{file_id}_input{Path(file.filename or '.wav').suffix}"
    with open(temp_input, "wb") as f:
        f.write(audio_data)

    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "agents"))
        from tutor_fonetica_base.praat_tools import (
            disponible, generar_espectrograma, generar_oscilograma,
            analizar_pitch as praat_pitch, resumen_formantes, formatear_analisis,
        )

        if not disponible():
            raise HTTPException(status_code=503, detail="parselmouth no está disponible en el servidor.")

        import parselmouth
        snd = parselmouth.Sound(str(temp_input))

        # Save as WAV for consistent processing
        snd.save(str(temp_wav), "WAV")

        # Generate spectrogram
        spec_path = generar_espectrograma(
            str(temp_wav), str(_PRAAT_OUTPUT_DIR),
            mostrar_formantes=mostrar_formantes,
            mostrar_pitch=mostrar_pitch,
            titulo=safe_name or "Audio",
        )

        # Generate oscillogram
        osc_path = generar_oscilograma(str(temp_wav), str(_PRAAT_OUTPUT_DIR))

        # Acoustic analysis
        pitch_data = praat_pitch(str(temp_wav))
        formant_data = resumen_formantes(str(temp_wav))

        # Build response with relative URLs
        spec_url = "/temp/praat/" + Path(spec_path).name
        osc_url = "/temp/praat/" + Path(osc_path).name
        duracion = round(snd.xmax - snd.xmin, 3)

        resultado = {
            "espectrograma_url": spec_url,
            "oscilograma_url": osc_url,
            "duracion": duracion,
            "pitch": pitch_data,
            "formantes": formant_data,
            "archivo": file.filename,
        }

        # Format analysis as markdown for the chat
        analisis_md = formatear_analisis({
            "duracion": duracion,
            "pitch": pitch_data,
            "formantes": formant_data,
        })
        resultado["analisis_md"] = analisis_md

        return resultado

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al analizar el audio: {str(e)}")
    finally:
        # Clean up input file (keep generated outputs for serving)
        if temp_input.exists() and temp_input != temp_wav:
            temp_input.unlink(missing_ok=True)


@app.get("/uninovis-uma/help/agoria-db")
async def help_agoria_db():
    """Help page for the Agoria DB Assistant"""
    return FileResponse(SCRIPT_DIR / "static" / "help_agoria_db.html")


@app.get("/uninovis-uma/help/european-projects")
async def help_european_projects():
    """Help page for the European Projects Assistant"""
    return FileResponse(SCRIPT_DIR / "static" / "help_european_projects.html")


@app.get("/european-projects")
async def european_projects_public_page():
    """Serve the public European Projects Assistant page (no auth required)"""
    return FileResponse(SCRIPT_DIR / "static" / "european_projects.html")


@app.get("/agents")
async def agents_page():
    """Serve the TOMMI AI Agents interface"""
    return FileResponse(SCRIPT_DIR / "static" / "index.html")


@app.get("/agents/{agent_id}")
async def agents_page_with_agent(agent_id: str):
    """Serve the TOMMI AI Agents interface with a specific agent pre-selected"""
    return FileResponse(SCRIPT_DIR / "static" / "index.html")


@app.get("/login")
async def login_page():
    """Sirve la página de login"""
    return FileResponse(SCRIPT_DIR / "static" / "login.html")


@app.get("/testing")
async def testing():
    """Sirve la interfaz de tester-developer"""
    return FileResponse(SCRIPT_DIR / "static" / "testing.html")


@app.get("/favicon.ico")
async def favicon():
    """Serve favicon"""
    return FileResponse(SCRIPT_DIR / "static" / "favicon.svg", media_type="image/svg+xml")


@app.get("/api/config")
async def get_config():
    """Devuelve la configuración pública del servidor"""
    return {"logging_enabled": ENABLE_LOGGING}


async def get_ollama_model_info(base_url: str, model_name: str) -> dict:
    """Consulta la API de Ollama para obtener información detallada del modelo."""
    import httpx
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.post(f"{base_url}/api/show", json={"name": model_name})
            if response.status_code == 200:
                data = response.json()
                # Extraer información relevante
                details = data.get("details", {})
                model_info = {
                    "family": details.get("family", ""),
                    "parameter_size": details.get("parameter_size", ""),
                    "quantization_level": details.get("quantization_level", ""),
                }
                # Construir nombre descriptivo
                parts = [model_name]
                if model_info["parameter_size"]:
                    parts.append(model_info["parameter_size"])
                if model_info["quantization_level"]:
                    parts.append(model_info["quantization_level"])
                return {
                    "full_name": " ".join(parts) if len(parts) > 1 else model_name,
                    "details": model_info
                }
    except Exception:
        pass
    return {"full_name": model_name, "details": {}}


async def check_ollama_health(base_url: str, model: str) -> dict:
    """Check if Ollama is running and if the model is available."""
    import httpx

    result = {
        "ollama_running": False,
        "model_available": False,
        "error": None,
        "error_code": None,
        "error_type": None,
        "instructions": None
    }

    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            # 1. Check if Ollama is running
            try:
                response = await client.get(f"{base_url}/api/tags")
                if response.status_code == 200:
                    result["ollama_running"] = True
                    models_data = response.json()
                    available_models = [m.get("name", "").split(":")[0] for m in models_data.get("models", [])]

                    # 2. Check if model is available
                    model_base = model.split(":")[0]
                    if model_base in available_models or model in [m.get("name", "") for m in models_data.get("models", [])]:
                        result["model_available"] = True
                    else:
                        err = format_error(LLM_MODEL_NOT_FOUND, model=model)
                        result.update(err)
                        result["available_models"] = available_models[:5]
                else:
                    result.update(format_error(LLM_OLLAMA_ERROR))
            except httpx.ConnectError:
                result.update(format_error(LLM_OLLAMA_NOT_RUNNING))
            except httpx.TimeoutException:
                result.update(format_error(LLM_OLLAMA_TIMEOUT))

    except Exception as e:
        result.update(format_error(LLM_UNKNOWN_ERROR, details=str(e)))

    return result


async def check_mistral_health(api_key: str) -> dict:
    """Check if the Mistral API key is valid."""
    import httpx

    result = {
        "api_key_configured": False,
        "api_key_valid": False,
        "error": None,
        "error_code": None,
        "error_type": None,
        "instructions": None
    }

    if not api_key:
        result.update(format_error(LLM_NO_API_KEY))
        return result

    result["api_key_configured"] = True

    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(
                "https://api.mistral.ai/v1/models",
                headers={"Authorization": f"Bearer {api_key}"}
            )
            if response.status_code == 200:
                result["api_key_valid"] = True
            elif response.status_code == 401:
                result.update(format_error(LLM_INVALID_API_KEY))
            else:
                result.update(format_error(LLM_MISTRAL_ERROR, status_code=response.status_code))
    except Exception as e:
        result.update(format_error(LLM_CONNECTION_ERROR, details=str(e)))

    return result


async def check_vllm_health(base_url: str, model: str) -> dict:
    """Check if vLLM is running and if the model is available."""
    import httpx

    result = {
        "vllm_running": False,
        "model_available": False,
        "error": None,
        "error_code": None,
        "error_type": None,
        "instructions": None
    }

    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            # 1. Check if vLLM is running (endpoint /v1/models)
            try:
                response = await client.get(f"{base_url}/models")
                if response.status_code == 200:
                    result["vllm_running"] = True
                    models_data = response.json()
                    available_models = [m.get("id", "") for m in models_data.get("data", [])]

                    # 2. Check if model is available
                    if model in available_models:
                        result["model_available"] = True
                    else:
                        err = format_error(LLM_VLLM_MODEL_NOT_FOUND, model=model)
                        result.update(err)
                        result["available_models"] = available_models[:5]
                else:
                    result.update(format_error(LLM_VLLM_ERROR, status_code=response.status_code))
            except httpx.ConnectError:
                result.update(format_error(LLM_VLLM_NOT_RUNNING))
            except httpx.TimeoutException:
                result.update(format_error(LLM_VLLM_NOT_RUNNING))

    except Exception as e:
        result.update(format_error(LLM_UNKNOWN_ERROR, details=str(e)))

    return result


# ---------------------------------------------------------------------------
# LLM provider helpers & agent visibility
# ---------------------------------------------------------------------------

LOCAL_PROVIDERS = {"ollama", "vllm"}

# Agent visibility: {agent_id: {level: "hidden"|"restricted"|"open", allowed_users: [...]}}
# Stored in data/agent_visibility.json
# Levels: hidden = superuser only, restricted (default) = testers+, open = any user
_VISIBILITY_FILE = Path(__file__).parent / "data" / "agent_visibility.json"

VISIBILITY_LEVELS = {"hidden", "restricted", "open"}


def _load_visibility() -> dict:
    if not _VISIBILITY_FILE.exists():
        return {}
    with open(_VISIBILITY_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Migrate old boolean format to new format
    migrated = False
    for k, v in list(data.items()):
        if isinstance(v, bool):
            data[k] = {"level": "restricted" if v else "hidden", "allowed_users": []}
            migrated = True
        elif isinstance(v, str):
            data[k] = {"level": v if v in VISIBILITY_LEVELS else "restricted", "allowed_users": []}
            migrated = True
    if migrated:
        _save_visibility(data)
    return data


def _save_visibility(data: dict) -> None:
    with open(_VISIBILITY_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _get_agent_visibility(agent_id: str) -> dict:
    """Get visibility config for an agent. Default is restricted."""
    vis = _load_visibility()
    entry = vis.get(agent_id, {})
    if isinstance(entry, dict):
        return {"level": entry.get("level", "restricted"), "allowed_users": entry.get("allowed_users", [])}
    return {"level": "restricted", "allowed_users": []}


def _can_user_see_agent(agent_id: str, username: str, role: str) -> bool:
    """Check if a user can see an agent based on visibility level and allowed_users."""
    if role == "superuser":
        return True
    cfg = _get_agent_visibility(agent_id)
    # If allowed_users is set, only those users can see it (overrides level)
    if cfg["allowed_users"]:
        return username.lower() in [u.lower() for u in cfg["allowed_users"]]
    level = cfg["level"]
    if level == "hidden":
        return False
    if level == "open":
        return True
    # restricted: testers and above
    return role in ("tester", "superuser")


def _get_agent_provider(agent_id: str) -> str:
    """Determine the LLM provider for a given agent. Returns provider name (lowercase)."""
    from dotenv import dotenv_values

    # Check agent-specific .env
    agent_info = runner.get_agent(agent_id)
    if agent_info:
        agent_env_path = Path(agent_info.path) / ".env"
        if agent_env_path.exists():
            agent_env = dotenv_values(agent_env_path)
            if agent_env.get("LLM_PROVIDER"):
                return agent_env["LLM_PROVIDER"].lower()

    # Fallback to global web/.env
    web_env = dotenv_values(Path(__file__).parent / ".env")
    return web_env.get("LLM_PROVIDER", "mistral").lower()


def _is_local_provider(agent_id: str) -> bool:
    """Check if an agent uses a local (on-premise) LLM provider."""
    return _get_agent_provider(agent_id) in LOCAL_PROVIDERS


@app.get("/api/llm-status")
async def get_llm_status(agent_id: Optional[str] = Query(None, description="ID del agente (opcional)")):
    """Devuelve información sobre el proveedor de LLM del agente especificado o el global"""
    from dotenv import dotenv_values

    # Determinar configuración a usar
    provider = None
    model = None
    base_url = None
    api_key = None

    # Si se especifica un agente, verificar si tiene su propia config LLM
    # Solo usa la config del agente si define LLM_PROVIDER explícitamente
    if agent_id:
        agent_info = runner.get_agent(agent_id)
        if agent_info:
            agent_env_path = Path(agent_info.path) / ".env"
            if agent_env_path.exists():
                agent_env = dotenv_values(agent_env_path)

                # Solo usar config del agente si define LLM_PROVIDER
                if agent_env.get("LLM_PROVIDER"):
                    provider = agent_env.get("LLM_PROVIDER").lower()

                    if provider == "ollama":
                        model = agent_env.get("OLLAMA_MODEL", "mistral")
                        base_url = agent_env.get("OLLAMA_BASE_URL", "http://localhost:11434")
                    elif provider == "vllm":
                        model = agent_env.get("VLLM_MODEL")
                        base_url = agent_env.get("VLLM_BASE_URL", "http://localhost:8000/v1")
                    else:
                        model = agent_env.get("MISTRAL_MODEL", "mistral-small-latest")
                        api_key = agent_env.get("MISTRAL_API_KEY")

    # Configuración global (leer directamente de web/.env, no de os.environ
    # porque agent_runner puede haberlo sobrescrito al cargar otro agente)
    if not provider:
        web_env = dotenv_values(Path(__file__).parent / ".env")
        provider = web_env.get("LLM_PROVIDER", "mistral").lower()
        if provider == "ollama":
            model = web_env.get("OLLAMA_MODEL", "mistral")
            base_url = web_env.get("OLLAMA_BASE_URL", "http://localhost:11434")
        elif provider == "vllm":
            model = web_env.get("VLLM_MODEL")
            base_url = web_env.get("VLLM_BASE_URL", "http://localhost:8000/v1")
        else:
            model = web_env.get("MISTRAL_MODEL", "mistral-small-latest")
            api_key = web_env.get("MISTRAL_API_KEY")

    # Build main LLM response
    response = {}

    if provider == "ollama":
        health = await check_ollama_health(base_url, model)

        if health["error"]:
            response = {
                "provider": "ollama",
                "is_local": True,
                "model": model,
                "base_url": base_url,
                "status": "error",
                "error_code": health["error_code"],
                "error_type": health["error_type"],
                "error": health["error"],
                "instructions": health["instructions"],
                "available_models": health.get("available_models", [])
            }
        else:
            model_info = await get_ollama_model_info(base_url, model)
            # Fetch all available Ollama models for cycling
            ollama_models = []
            try:
                import httpx
                async with httpx.AsyncClient(timeout=5.0) as hclient:
                    resp = await hclient.get(f"{base_url}/api/tags")
                    if resp.status_code == 200:
                        for m in resp.json().get("models", []):
                            name = m.get("name", "")
                            size_bytes = m.get("size", 0)
                            size_gb = round(size_bytes / (1024**3), 1)
                            ollama_models.append({"name": name, "size_gb": size_gb})
            except Exception:
                pass
            # Filter out models larger than the cycling threshold (default 20 GB).
            # This prevents users from cycling to very large models that would
            # cause memory pressure and slow model-swapping on shared instances.
            max_cycling_gb = float(os.getenv("OLLAMA_MAX_CYCLING_GB", "20"))
            cycling_models = [m for m in ollama_models if m["size_gb"] <= max_cycling_gb]
            # Build sizes dict with both full name and base name (without tag)
            model_sizes = {}
            for m in ollama_models:
                model_sizes[m["name"]] = m["size_gb"]
                base_name = m["name"].split(":")[0]
                if base_name not in model_sizes:
                    model_sizes[base_name] = m["size_gb"]
            response = {
                "provider": "ollama",
                "is_local": True,
                "model": model,
                "display_name": f"Ollama: {model_info['full_name']}",
                "base_url": base_url,
                "model_details": model_info["details"],
                "status": "ok",
                "available_models": [m["name"] for m in cycling_models],
                "model_sizes": model_sizes,
            }
    elif provider == "vllm":
        health = await check_vllm_health(base_url, model)

        if health["error"]:
            response = {
                "provider": "vllm",
                "is_local": True,
                "model": model,
                "base_url": base_url,
                "status": "error",
                "error_code": health["error_code"],
                "error_type": health["error_type"],
                "error": health["error"],
                "instructions": health["instructions"],
                "available_models": health.get("available_models", [])
            }
        else:
            response = {
                "provider": "vllm",
                "is_local": True,
                "model": model,
                "display_name": f"vLLM: {model}",
                "base_url": base_url,
                "status": "ok"
            }
    else:
        health = await check_mistral_health(api_key)

        if health["error"]:
            response = {
                "provider": "mistral",
                "is_local": False,
                "model": model,
                "status": "error",
                "error_code": health["error_code"],
                "error_type": health["error_type"],
                "error": health["error"],
                "instructions": health["instructions"]
            }
        else:
            response = {
                "provider": "mistral",
                "is_local": False,
                "model": model,
                "display_name": f"Mistral: {model}",
                "base_url": None,
                "status": "ok"
            }

    # Add available models list (for model switching UI)
    # Only apply AVAILABLE_MODELS from .env for cloud providers;
    # local (Ollama/vLLM) already have their own list from the API.
    if not response.get("is_local"):
        from dotenv import dotenv_values as _dv
        _web_env = _dv(Path(__file__).parent / ".env")
        available_raw = _web_env.get("AVAILABLE_MODELS", "")
        if available_raw:
            response["available_models"] = [m.strip() for m in available_raw.split(",") if m.strip()]

    return response


@app.get("/api/llm-models")
async def get_llm_models(provider: str = Query(...), session: dict = Depends(require_role("tester"))):
    """Return available models for a given LLM provider."""
    from dotenv import dotenv_values
    web_env = dotenv_values(Path(__file__).parent / ".env")
    provider = provider.lower()

    if provider == "ollama":
        base_url = web_env.get("OLLAMA_BASE_URL", "http://localhost:11434")
        models = []
        try:
            import httpx
            async with httpx.AsyncClient(timeout=5.0) as hclient:
                resp = await hclient.get(f"{base_url}/api/tags")
                if resp.status_code == 200:
                    for m in resp.json().get("models", []):
                        name = m.get("name", "")
                        size_gb = round(m.get("size", 0) / (1024**3), 1)
                        models.append({"name": name, "size_gb": size_gb})
        except Exception:
            pass
        max_gb = float(os.getenv("OLLAMA_MAX_CYCLING_GB", "20"))
        return {"provider": "ollama", "models": [m["name"] for m in models if m["size_gb"] <= max_gb]}
    elif provider == "mistral":
        available_raw = web_env.get("AVAILABLE_MODELS", "")
        models = [m.strip() for m in available_raw.split(",") if m.strip()] if available_raw else ["mistral-small-latest", "mistral-large-latest"]
        return {"provider": "mistral", "models": models}
    elif provider == "openai":
        return {"provider": "openai", "models": ["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo"]}
    elif provider == "anthropic":
        return {"provider": "anthropic", "models": ["claude-sonnet-4-20250514", "claude-haiku-4-5-20251001"]}
    else:
        return {"provider": provider, "models": []}


@app.put("/api/agents/{agent_id}/llm-provider")
async def set_agent_llm_provider(agent_id: str, request: Request, session: dict = Depends(require_role("superuser"))):
    """Update the LLM provider and model for an agent by writing its .env file."""
    agent = runner.get_agent(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    body = await request.json()
    provider = body.get("provider", "").lower()
    model = body.get("model", "")

    if not provider or not model:
        raise HTTPException(status_code=400, detail="provider and model are required")

    agent_env_path = Path(agent.path) / ".env"

    # Read existing .env lines, preserving non-LLM entries
    existing_lines = []
    llm_keys = {"LLM_PROVIDER", "OLLAMA_MODEL", "OLLAMA_BASE_URL", "MISTRAL_MODEL",
                "MISTRAL_API_KEY", "VLLM_MODEL", "VLLM_BASE_URL", "OPENAI_MODEL",
                "OPENAI_API_KEY", "ANTHROPIC_MODEL", "ANTHROPIC_API_KEY"}
    if agent_env_path.exists():
        with open(agent_env_path, "r", encoding="utf-8") as f:
            for line in f:
                key = line.split("=", 1)[0].strip()
                if key not in llm_keys:
                    existing_lines.append(line)

    # Build new LLM lines
    from dotenv import dotenv_values
    web_env = dotenv_values(Path(__file__).parent / ".env")
    new_lines = [f"LLM_PROVIDER={provider}\n"]

    if provider == "ollama":
        base_url = web_env.get("OLLAMA_BASE_URL", "http://localhost:11434")
        new_lines.append(f"OLLAMA_MODEL={model}\n")
        new_lines.append(f"OLLAMA_BASE_URL={base_url}\n")
    elif provider == "mistral":
        api_key = web_env.get("MISTRAL_API_KEY", "")
        new_lines.append(f"MISTRAL_MODEL={model}\n")
        if api_key:
            new_lines.append(f"MISTRAL_API_KEY={api_key}\n")
    elif provider == "openai":
        api_key = web_env.get("OPENAI_API_KEY", "")
        new_lines.append(f"OPENAI_MODEL={model}\n")
        if api_key:
            new_lines.append(f"OPENAI_API_KEY={api_key}\n")
    elif provider == "anthropic":
        api_key = web_env.get("ANTHROPIC_API_KEY", "")
        new_lines.append(f"ANTHROPIC_MODEL={model}\n")
        if api_key:
            new_lines.append(f"ANTHROPIC_API_KEY={api_key}\n")

    with open(agent_env_path, "w", encoding="utf-8") as f:
        f.writelines(existing_lines)
        f.writelines(new_lines)

    return {"ok": True, "provider": provider, "model": model}


@app.get("/api/history")
async def get_history(
    agent_id: str = Query(..., description="ID del agente"),
    session_id: Optional[str] = Query(None, description="ID de sesión")
):
    """Obtiene el historial de consultas de un agente para una sesión"""
    history = runner.get_agent_history(agent_id, session_id)
    return {"history": history}


@app.get("/api/agents", response_model=list[AgentResponse])
async def list_agents(request: Request, mode: Optional[str] = Query(None), prefix: Optional[str] = Query(None)):
    """Lista todos los agentes disponibles (filtered by visibility, mode, role, and prefix)"""
    agents = runner.discover_agents()

    # Determine user role and username from session
    token = _get_token(request)
    session = get_session(token) if token else None
    user_role = session["role"] if session else "user"
    username = session["username"] if session else ""

    # Optional prefix filter (e.g., ?prefix=rag_study_ to show only study agents)
    if prefix:
        # When filtering by prefix, skip visibility checks (study agents are hidden from main list)
        agents = [a for a in agents if a.id.startswith(prefix)]
    else:
        # Normal list: filter by visibility and exclude study agents from main list
        agents = [a for a in agents if _can_user_see_agent(a.id, username, user_role)]
        agents = [a for a in agents if not a.id.startswith("rag_study_") and not a.id.startswith("rag_study2_")]

    return [
        AgentResponse(
            id=a.id,
            name=a.name,
            agent_type=a.agent_type,
            description=a.description,
            welcome_message=a.welcome_message,
            example_queries=a.example_queries,
            rag_approach=a.rag_approach,
            show_history=a.show_history,
            show_description=a.show_description,
            transparency_level=a.transparency_level,
            transparency_type=a.transparency_type,
            prompt_level=a.prompt_level,
            decision_trace=a.decision_trace,
            reliability_cues=a.reliability_cues
        )
        for a in agents
    ]


@app.get("/api/agents/visibility")
async def get_agents_visibility(session: dict = Depends(require_role("superuser"))):
    """Get visibility config of all agents (superuser only)."""
    all_agents = runner.discover_agents()
    vis = _load_visibility()
    return [
        {
            "id": a.id,
            "name": a.name,
            "level": vis.get(a.id, {}).get("level", "restricted") if isinstance(vis.get(a.id), dict) else "restricted",
            "allowed_users": vis.get(a.id, {}).get("allowed_users", []) if isinstance(vis.get(a.id), dict) else [],
            "provider": _get_agent_provider(a.id),
            "is_local": _is_local_provider(a.id),
        }
        for a in all_agents
    ]


class AgentVisibilityBody(BaseModel):
    level: str = "restricted"
    allowed_users: list[str] = []


@app.put("/api/agents/{agent_id}/visibility")
async def set_agent_visibility(
    agent_id: str,
    body: AgentVisibilityBody,
    session: dict = Depends(require_role("superuser")),
):
    """Set visibility level and allowed users for an agent (superuser only)."""
    if body.level not in VISIBILITY_LEVELS:
        raise HTTPException(status_code=400, detail=f"Invalid level. Must be one of: {', '.join(VISIBILITY_LEVELS)}")
    agent = runner.get_agent(agent_id)
    if not agent:
        all_agents = runner.discover_agents()
        if not any(a.id == agent_id for a in all_agents):
            raise HTTPException(status_code=404, detail="Agent not found")
    vis = _load_visibility()
    vis[agent_id] = {"level": body.level, "allowed_users": body.allowed_users}
    _save_visibility(vis)
    return {"ok": True, "agent_id": agent_id, "level": body.level, "allowed_users": body.allowed_users}


# ---------------------------------------------------------------------------
# Tool access management (superuser only)
# ---------------------------------------------------------------------------

@app.get("/api/tool-access")
async def get_tool_access(session: dict = Depends(require_auth)):
    """Get current tool access configuration (readable by all authenticated users)."""
    return {"tool_access": TOOL_ACCESS}


@app.put("/api/tool-access")
async def set_tool_access(request: Request, session: dict = Depends(require_role("superuser"))):
    """Update tool access configuration (superuser only)."""
    from auth import save_tool_access
    body = await request.json()
    new_access = body.get("tool_access")
    if not isinstance(new_access, dict):
        raise HTTPException(status_code=400, detail="tool_access must be a dict")
    # Validate: values must be lists of known role strings
    valid_roles = {"public", "student", "admin_staff", "teaching_staff", "tester", "superuser"}
    for tool_id, roles in new_access.items():
        if not isinstance(roles, list):
            raise HTTPException(status_code=400, detail=f"Roles for {tool_id} must be a list")
        for r in roles:
            if r not in valid_roles:
                raise HTTPException(status_code=400, detail=f"Unknown role: {r}")
    # Update in-memory and persist
    TOOL_ACCESS.clear()
    TOOL_ACCESS.update(new_access)
    save_tool_access(new_access)
    return {"ok": True}


@app.get("/api/public-tools")
async def get_public_tools():
    """Return tools marked as public (no auth required)."""
    public = [tool_id for tool_id, roles in TOOL_ACCESS.items()
              if "public" in roles]
    return {"public_tools": public}


# ---------------------------------------------------------------------------
# Public agent endpoints (no auth required) — for standalone public pages
# ---------------------------------------------------------------------------

# Set of agent IDs that are allowed to be accessed publicly
_PUBLIC_AGENT_IDS = {"responsible_ai3", "health_wellbeing_sistems", "proyectoseuopeos", "pisha5", "algoria_map", "eulalia", "sonic_composer", "sonic_composer2", "quiron"}


@app.get("/api/public-agent/{agent_id}/config")
async def get_public_agent_config(agent_id: str):
    """Get basic config for a public agent (no auth required)."""
    if agent_id not in _PUBLIC_AGENT_IDS:
        raise HTTPException(status_code=404, detail="Agent not found")
    agent = runner.get_agent(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    config_path = Path(agent.path) / "config.json"
    if not config_path.exists():
        return {}
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    # Only expose safe fields
    return {
        "agent_name": cfg.get("agent_name", agent.name),
        "description": cfg.get("description", ""),
        "welcome_message": cfg.get("welcome_message", ""),
        "example_queries": cfg.get("example_queries", []),
    }


@app.get("/api/public-agent/{agent_id}/info")
async def get_public_agent_info(agent_id: str):
    """Get agent type, LLM provider, and transparency info for a public agent (no auth)."""
    if agent_id not in _PUBLIC_AGENT_IDS:
        raise HTTPException(status_code=404, detail="Agent not found")
    agent = runner.get_agent(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    config_path = Path(agent.path) / "config.json"
    cfg = {}
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)

    # Agent type label
    type_labels = {
        'rag': 'RAG', 'rag_metadata': 'RAG + Metadata',
        'rag_metadata_vectorless': 'RAG + Metadata (Vectorless)',
        'oneshot': 'One-shot', 'custom': 'Custom',
    }
    agent_type = agent.agent_type or 'oneshot'

    # LLM provider info
    provider = _get_agent_provider(agent_id)
    is_local = provider in LOCAL_PROVIDERS

    # Model name
    from dotenv import dotenv_values
    agent_env_path = Path(agent.path) / ".env"
    model = ""
    if agent_env_path.exists():
        env = dotenv_values(agent_env_path)
        model = env.get("MISTRAL_MODEL") or env.get("OLLAMA_MODEL") or env.get("OPENAI_MODEL") or ""
    if not model:
        web_env = dotenv_values(Path(__file__).parent / ".env")
        model = web_env.get("MISTRAL_MODEL") or web_env.get("OLLAMA_MODEL") or ""

    return {
        "agent_name": cfg.get("agent_name", agent.name),
        "agent_type": agent_type,
        "agent_type_label": type_labels.get(agent_type, agent_type),
        "is_local": is_local,
        "model": model,
        "prompt_level": cfg.get("prompt_level", ""),
        "decision_trace": cfg.get("decision_trace") or cfg.get("transparency_level", ""),
        "reliability_cues": cfg.get("reliability_cues", ""),
    }


@app.get("/api/public-agent/{agent_id}/chat/stream")
async def public_agent_chat_stream(
    request: Request,
    agent_id: str,
    message: str = Query(..., description="Message to send"),
    session_id: Optional[str] = Query(None, description="Session ID"),
):
    """Public streaming chat endpoint for open agents (no auth required)."""
    if agent_id not in _PUBLIC_AGENT_IDS:
        raise HTTPException(status_code=404, detail="Agent not found")
    agent = runner.get_agent(agent_id)
    if not agent:
        err = format_error(AGENT_NOT_FOUND, agent_id=agent_id)
        raise HTTPException(status_code=404, detail=f"Error {err['error_code']}: {err['error']}")

    client_ip = request.client.host if request.client else "unknown"

    # If user is authenticated, pass username to agent (for progress tracking)
    # Try tutores auth first, then TOMMI auth
    tutores_token = _get_tutores_token(request)
    tutores_session = tutores_get_session(tutores_token) if tutores_token else None
    if tutores_session:
        req_username = tutores_session.get("username")
        req_role = tutores_session.get("role")
    else:
        tommi_token = _get_token(request)
        tommi_session = get_session(tommi_token) if tommi_token else None
        req_username = tommi_session.get("username") if tommi_session else None
        req_role = tommi_session.get("role") if tommi_session else None

    async def event_generator():
        new_session_id = None
        full_response = ""
        response_metadata = {}
        try:
            async for event_type, content, returned_session_id in runner.run_query_stream(
                agent_id=agent_id,
                message=message,
                session_id=session_id,
                username=req_username,
                role=req_role,
            ):
                if returned_session_id and not new_session_id:
                    new_session_id = returned_session_id
                    yield f"event: session\ndata: {new_session_id}\n\n"

                if event_type == "metadata":
                    # Capture metadata (not sent to client)
                    if isinstance(content, dict):
                        response_metadata.update(content)
                elif event_type == "status":
                    yield f"event: status\ndata: {content}\n\n"
                elif event_type == "badge":
                    escaped = content.replace("\n", "\\n")
                    yield f"event: badge\ndata: {escaped}\n\n"
                elif event_type == "trace":
                    escaped = content.replace("\n", "\\n")
                    yield f"event: trace\ndata: {escaped}\n\n"
                elif event_type == "procedural_banner":
                    escaped = content.replace("\n", "\\n")
                    yield f"event: procedural_banner\ndata: {escaped}\n\n"
                elif event_type == "replace_banner":
                    escaped = content.replace("\n", "\\n")
                    yield f"event: replace_banner\ndata: {escaped}\n\n"
                elif event_type == "replace":
                    full_response = content
                    escaped = content.replace("\n", "\\n")
                    yield f"event: replace\ndata: {escaped}\n\n"
                else:
                    full_response += content
                    escaped = content.replace("\n", "\\n")
                    yield f"data: {escaped}\n\n"

            log_conversation(
                client_ip=client_ip,
                agent_id=agent_id,
                agent_name=agent.name,
                question=message,
                response=full_response,
                session_id=new_session_id or session_id or "",
                transparency_level=None,
                username=None,
                extra=response_metadata if response_metadata else None,
            )
            yield "event: done\ndata: complete\n\n"
        except Exception as e:
            err = format_error(SERVER_STREAMING_ERROR, details=str(e))
            yield f"event: error\ndata: {json.dumps(err)}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ---------------------------------------------------------------------------
# Agent config editing (tester+ only)
# ---------------------------------------------------------------------------

@app.get("/api/agents/{agent_id}/config")
async def get_agent_config(agent_id: str, session: dict = Depends(require_role("tester"))):
    """Get config.json and prompts.json for an agent (tester+).

    Includes the caller's role so the frontend can render a role-appropriate
    editor (testers see basic settings + prompts + scope terms, superusers
    see everything).
    """
    agent = runner.get_agent(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    agent_path = Path(agent.path)
    config, prompts = {}, {}
    config_path = agent_path / "config.json"
    prompts_path = agent_path / "prompts.json"
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    if prompts_path.exists():
        with open(prompts_path, "r", encoding="utf-8") as f:
            prompts = json.load(f)

    # Include current LLM provider info
    from dotenv import dotenv_values
    agent_env_path = agent_path / ".env"
    llm_provider = None
    llm_model = None
    if agent_env_path.exists():
        agent_env = dotenv_values(agent_env_path)
        llm_provider = agent_env.get("LLM_PROVIDER")
        if llm_provider:
            llm_provider = llm_provider.lower()
            if llm_provider == "ollama":
                llm_model = agent_env.get("OLLAMA_MODEL")
            elif llm_provider == "mistral":
                llm_model = agent_env.get("MISTRAL_MODEL")
            elif llm_provider == "openai":
                llm_model = agent_env.get("OPENAI_MODEL")
            elif llm_provider == "anthropic":
                llm_model = agent_env.get("ANTHROPIC_MODEL")
    if not llm_provider:
        web_env = dotenv_values(Path(__file__).parent / ".env")
        llm_provider = web_env.get("LLM_PROVIDER", "mistral").lower()
        if llm_provider == "ollama":
            llm_model = web_env.get("OLLAMA_MODEL")
        elif llm_provider == "mistral":
            llm_model = web_env.get("MISTRAL_MODEL")
        elif llm_provider == "openai":
            llm_model = web_env.get("OPENAI_MODEL")
        elif llm_provider == "anthropic":
            llm_model = web_env.get("ANTHROPIC_MODEL")

    return {"config": config, "prompts": prompts, "role": session.get("role", "tester"),
            "llm_provider": llm_provider, "llm_model": llm_model}


@app.put("/api/agents/{agent_id}/config")
async def set_agent_config(agent_id: str, request: Request, session: dict = Depends(require_role("tester"))):
    """Update config.json and/or prompts.json for an agent (tester+).

    Role-based write protection:
    - tester: can update basic settings, prompts, scope terms, examples
    - superuser: can update everything (including agent_id, universities, etc.)
    """
    from auth import max_role_level as _max_level
    agent = runner.get_agent(agent_id)
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    body = await request.json()
    agent_path = Path(agent.path)
    role_level = _max_level(session)

    if "config" in body and isinstance(body["config"], dict):
        config_path = agent_path / "config.json"
        current = {}
        if config_path.exists():
            with open(config_path, "r", encoding="utf-8") as f:
                current = json.load(f)

        new_config = body["config"]

        if role_level >= 4:  # superuser — full access
            merged = new_config
        else:  # tester — only immediate-effect settings (no restart needed)
            TESTER_FIELDS = {"prompt_level", "transparency_level", "reliability_cues", "humility_prompt", "humility_postprocessing", "decision_trace"}
            merged = {**current}
            for key in TESTER_FIELDS:
                if key in new_config:
                    merged[key] = new_config[key]

        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(merged, f, indent=2, ensure_ascii=False)
            f.write("\n")

    if "prompts" in body and isinstance(body["prompts"], dict) and role_level >= 4:
        with open(agent_path / "prompts.json", "w", encoding="utf-8") as f:
            json.dump(body["prompts"], f, indent=2, ensure_ascii=False)
            f.write("\n")

    return {"ok": True, "agent_id": agent_id}


# ---------------------------------------------------------------------------
# Data export & review endpoints (for tester verification)
# ---------------------------------------------------------------------------

@app.get("/api/agents/{agent_id}/export/researchers")
async def export_researchers(
    agent_id: str,
    university: Optional[str] = Query(None),
    session: dict = Depends(require_role("tester")),
):
    """Export researchers as Excel with a Review column for tester annotations."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        raise HTTPException(status_code=404, detail="Agent not found")

    researchers_path = Path(agent_info.path) / "data" / "researchers.json"
    if not researchers_path.exists():
        raise HTTPException(status_code=404, detail="No researchers data found for this agent")

    with open(researchers_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Researchers"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    review_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers
    headers = ["University", "Researcher", "Paper count", "Topics", "Papers", "Review: Correct? (Yes/No)", "Review: Comments"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data
    row = 2
    unis = [university.upper()] if university else sorted(data.keys())
    for uni in unis:
        if uni not in data:
            continue
        for researcher in data[uni]:
            papers_str = "; ".join([f"{p.get('title', '')} ({p.get('year', '')})" for p in researcher.get("papers", [])])
            topics_str = ", ".join(researcher.get("topics", []))

            ws.cell(row=row, column=1, value=uni).border = thin_border
            ws.cell(row=row, column=2, value=researcher.get("name", "")).border = thin_border
            ws.cell(row=row, column=3, value=researcher.get("paper_count", 0)).border = thin_border
            ws.cell(row=row, column=4, value=topics_str).border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=5, value=papers_str).border = thin_border
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
            # Review columns (yellow background)
            for col in [6, 7]:
                cell = ws.cell(row=row, column=col, value="")
                cell.fill = review_fill
                cell.border = thin_border
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 35

    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "RESEARCHER DATA REVIEW",
        "",
        "Please review the researcher data for your university:",
        "",
        "1. Check each researcher in the list.",
        "2. In the 'Review: Correct?' column, write 'Yes' if the data is correct, or 'No' if there is an error.",
        "3. In the 'Review: Comments' column, describe any issues found:",
        "   - Researcher should not be listed (not affiliated with this university)",
        "   - Wrong topics assigned",
        "   - Missing papers",
        "   - Wrong paper attribution",
        "",
        "4. If researchers are MISSING from the list, add them at the bottom with:",
        "   - University code, Name, and in Comments write 'MISSING - should be included'",
        "",
        "5. Save the file and send it back to the TOMMI team.",
        "",
        f"Data source: OpenAlex (https://openalex.org/)",
        f"Agent: {agent_id}",
        f"Export date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws_inst.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    ws_inst.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    uni_suffix = f"_{university.upper()}" if university else ""
    filename = f"{agent_id}_researchers{uni_suffix}_review.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/agents/{agent_id}/export/papers")
async def export_papers(
    agent_id: str,
    university: Optional[str] = Query(None),
    session: dict = Depends(require_role("tester")),
):
    """Export publications as Excel with a Review column for tester annotations."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        raise HTTPException(status_code=404, detail="Agent not found")

    papers_path = Path(agent_info.path) / "data" / "papers.json"
    if not papers_path.exists():
        raise HTTPException(status_code=404, detail="No papers data found for this agent")

    with open(papers_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    universities = data.get("universities", data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    review_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["University", "ID", "Title", "Authors", "Year", "DOI", "Type", "Cited by", "Review: Correct? (Yes/No)", "Review: Comments"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row = 2
    unis = [university.upper()] if university else sorted(universities.keys())
    for uni in unis:
        uni_data = universities.get(uni, {})
        papers = uni_data.get("papers", []) if isinstance(uni_data, dict) else uni_data
        for paper in papers:
            authors_str = ", ".join([a.get("name", "") for a in paper.get("authors", [])])
            doi = paper.get("doi", "") or ""

            ws.cell(row=row, column=1, value=uni).border = thin_border
            ws.cell(row=row, column=2, value=paper.get("id", "")).border = thin_border
            ws.cell(row=row, column=3, value=paper.get("title", "")).border = thin_border
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=4, value=authors_str).border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=5, value=paper.get("publication_year", "")).border = thin_border
            ws.cell(row=row, column=6, value=doi).border = thin_border
            ws.cell(row=row, column=7, value=paper.get("type", "")).border = thin_border
            ws.cell(row=row, column=8, value=paper.get("cited_by_count", 0)).border = thin_border
            for col in [9, 10]:
                cell = ws.cell(row=row, column=col, value="")
                cell.fill = review_fill
                cell.border = thin_border
            row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 35

    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "PUBLICATION DATA REVIEW",
        "",
        "Please review the publications for your university:",
        "",
        "1. Check each publication in the list.",
        "2. In the 'Review: Correct?' column, write 'Yes' if correct, or 'No' if there is an error.",
        "3. In the 'Review: Comments' column, describe any issues:",
        "   - Paper should not be attributed to this university",
        "   - Wrong authors listed",
        "   - Wrong year or title",
        "   - Paper is a duplicate",
        "",
        "4. If publications are MISSING from the list, add them at the bottom with:",
        "   - University code, Title, Authors, Year, and in Comments write 'MISSING'",
        "",
        "5. Save the file and send it back to the TOMMI team.",
        "",
        f"Data source: OpenAlex (https://openalex.org/)",
        f"Agent: {agent_id}",
        f"Export date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws_inst.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    ws_inst.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    uni_suffix = f"_{university.upper()}" if university else ""
    filename = f"{agent_id}_papers{uni_suffix}_review.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/agents/{agent_id}/export/projects")
async def export_projects(
    agent_id: str,
    university: Optional[str] = Query(None),
    session: dict = Depends(require_role("tester")),
):
    """Export projects as Excel with a Review column for tester annotations."""
    import io, re
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        raise HTTPException(status_code=404, detail="Agent not found")

    project_dir = Path(agent_info.path) / "data" / "project_docs"
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail="No project data found for this agent")

    # Parse project markdown files
    projects = []
    for md_file in sorted(project_dir.glob("*.md")):
        content = md_file.read_text(encoding="utf-8")
        title_match = re.search(r"^# (.+)", content, re.MULTILINE)
        grant_match = re.search(r"\*\*Grant ID:\*\*\s*(\S+)", content)
        programme_match = re.search(r"\*\*Programme:\*\*\s*(.+)", content)
        period_match = re.search(r"\*\*Period:\*\*\s*(.+)", content)
        status_match = re.search(r"\*\*Status:\*\*\s*(\S+)", content)
        keywords_match = re.search(r"\*\*Keywords:\*\*\s*(.+)", content)

        # Extract participants
        participants = []
        in_participants = False
        for line in content.split("\n"):
            if line.strip().startswith("## Participants"):
                in_participants = True
                continue
            if in_participants:
                if line.strip().startswith("## ") or line.strip().startswith("# "):
                    break
                if line.strip().startswith("- "):
                    participants.append(line.strip()[2:])

        project = {
            "file": md_file.name,
            "title": title_match.group(1).split(":")[1].strip() if title_match and ":" in title_match.group(1) else (title_match.group(1) if title_match else md_file.stem),
            "grant_id": grant_match.group(1) if grant_match else "",
            "programme": programme_match.group(1).strip() if programme_match else "",
            "period": period_match.group(1).strip() if period_match else "",
            "status": status_match.group(1).strip() if status_match else "",
            "keywords": keywords_match.group(1).strip() if keywords_match else "",
            "participants": participants,
        }

        # Filter by university if specified
        if university:
            # Check config for university full names
            config_path = Path(agent_info.path) / "config.json"
            uni_names = {}
            if config_path.exists():
                with open(config_path) as f:
                    cfg = json.load(f)
                uni_names = {k: v.get("name", "") for k, v in cfg.get("universities", {}).items()}
            uni_full = uni_names.get(university.upper(), university)
            if not any(uni_full.lower() in p.lower() or university.upper() in p.upper() for p in participants):
                continue

        projects.append(project)

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    review_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Grant ID", "Title", "Programme", "Period", "Status", "Keywords", "Participants", "Review: Correct? (Yes/No)", "Review: Comments"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r, proj in enumerate(projects, 2):
        ws.cell(row=r, column=1, value=proj["grant_id"]).border = thin_border
        ws.cell(row=r, column=2, value=proj["title"]).border = thin_border
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=proj["programme"]).border = thin_border
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4, value=proj["period"]).border = thin_border
        ws.cell(row=r, column=5, value=proj["status"]).border = thin_border
        ws.cell(row=r, column=6, value=proj["keywords"]).border = thin_border
        ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=7, value="\n".join(proj["participants"])).border = thin_border
        ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True)
        for col in [8, 9]:
            cell = ws.cell(row=r, column=col, value="")
            cell.fill = review_fill
            cell.border = thin_border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 35

    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "PROJECT DATA REVIEW",
        "",
        "Please review the project data:",
        "",
        "1. Check each project in the list.",
        "2. In the 'Review: Correct?' column, write 'Yes' if correct, or 'No' if there is an error.",
        "3. In the 'Review: Comments' column, describe any issues:",
        "   - Your university is not actually a participant",
        "   - Wrong project details (title, period, programme)",
        "   - Missing participants",
        "",
        "4. If projects are MISSING from the list, add them at the bottom with:",
        "   - Grant ID, Title, and in Comments write 'MISSING'",
        "",
        "5. Save the file and send it back to the TOMMI team.",
        "",
        f"Data source: CORDIS (https://cordis.europa.eu/)",
        f"Agent: {agent_id}",
        f"Export date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws_inst.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    ws_inst.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    uni_suffix = f"_{university.upper()}" if university else ""
    filename = f"{agent_id}_projects{uni_suffix}_review.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/agents/{agent_id}/reset-logs")
async def reset_agent_logs(agent_id: str, session: dict = Depends(require_role("superuser"))):
    """
    Archive and reset all logs for an agent (superuser only).
    Feedback logs and audit logs are renamed with a timestamp suffix.
    """
    from datetime import datetime as dt
    timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
    archived = []

    # Feedback logs in web/logs/
    for suffix in ["_feedback_tester.jsonl", "_feedback_user.jsonl", "_conversations.jsonl", "_conversations.log"]:
        log_path = LOGS_DIR / f"{agent_id}{suffix}"
        if log_path.exists() and log_path.stat().st_size > 0:
            ext = suffix.rsplit('.', 1)[-1]
            archive_name = f"{agent_id}{suffix.rsplit('.', 1)[0]}_{timestamp}.{ext}"
            archive_path = LOGS_DIR / archive_name
            log_path.rename(archive_path)
            archived.append(archive_name)

    # Reset cached logger so a new file handler is created for the fresh log
    if agent_id in _agent_loggers:
        logger = _agent_loggers.pop(agent_id)
        for h in logger.handlers[:]:
            h.close()
            logger.removeHandler(h)

    # Audit log in agents/{agent_id}/data/
    agent_info = runner.get_agent(agent_id)
    if agent_info:
        audit_path = Path(agent_info.path) / "data" / "audit_log.jsonl"
        if audit_path.exists() and audit_path.stat().st_size > 0:
            archive_name = f"audit_log_{timestamp}.jsonl"
            archive_path = audit_path.parent / archive_name
            audit_path.rename(archive_path)
            archived.append(f"agents/{agent_id}/data/{archive_name}")

    if not archived:
        return {"ok": True, "message": "No logs to reset", "archived": []}

    return {"ok": True, "message": f"Archived {len(archived)} log file(s)", "archived": archived}


# ---------------------------------------------------------------------------
# Log Analytics API (superuser only)
# ---------------------------------------------------------------------------

@app.get("/api/logs/summary")
async def logs_summary(
    period: str = "day",
    start: str | None = None,
    end: str | None = None,
    session: dict = Depends(require_role("superuser")),
):
    """
    Return aggregated request/visitor stats across all agents.
    period: 'hour', 'day', 'week' (ignored when start/end are given).
    start/end: ISO date(time) strings for a custom interval.
    """
    from datetime import datetime as dt, timedelta

    now = dt.now()
    if start and end:
        try:
            t_start = dt.fromisoformat(start)
            t_end = dt.fromisoformat(end)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use ISO 8601.")
    else:
        if period == "hour":
            t_start = now - timedelta(hours=1)
        elif period == "3days":
            t_start = now - timedelta(days=3)
        elif period == "week":
            t_start = now - timedelta(weeks=1)
        else:
            t_start = now - timedelta(days=1)
        t_end = now

    # Scan all JSONL conversation logs
    totals = {"requests": 0, "sessions": set(), "users": set()}
    per_agent: dict[str, dict] = {}

    for log_file in sorted(LOGS_DIR.glob("*_conversations.jsonl")):
        agent_id = log_file.name.rsplit("_conversations.jsonl", 1)[0]
        agent_stats = {"requests": 0, "sessions": set(), "users": set()}

        try:
            with open(log_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entry = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    try:
                        ts = dt.fromisoformat(entry["timestamp"])
                    except (KeyError, ValueError):
                        continue
                    if ts < t_start or ts > t_end:
                        continue

                    agent_stats["requests"] += 1
                    if entry.get("session_id"):
                        agent_stats["sessions"].add(entry["session_id"])
                    if entry.get("user_id"):
                        agent_stats["users"].add(entry["user_id"])
        except Exception:
            continue

        if agent_stats["requests"] > 0:
            per_agent[agent_id] = {
                "requests": agent_stats["requests"],
                "sessions": len(agent_stats["sessions"]),
                "unique_users": len(agent_stats["users"]),
            }
            totals["requests"] += agent_stats["requests"]
            totals["sessions"] |= agent_stats["sessions"]
            totals["users"] |= agent_stats["users"]

    return {
        "period": period if not (start and end) else "custom",
        "start": t_start.isoformat(),
        "end": t_end.isoformat(),
        "totals": {
            "requests": totals["requests"],
            "sessions": len(totals["sessions"]),
            "unique_users": len(totals["users"]),
        },
        "per_agent": per_agent,
    }


@app.get("/api/logs/files")
async def list_log_files(session: dict = Depends(require_role("superuser"))):
    """List all log files with size and modification time."""
    files = []
    for f in sorted(LOGS_DIR.iterdir()):
        if f.is_file() and f.name != ".DS_Store":
            stat = f.stat()
            files.append({
                "name": f.name,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    return files


@app.get("/api/logs/content")
async def get_log_content(filename: str, session: dict = Depends(require_role("superuser"))):
    """Return parsed content of a log file (superuser only)."""
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    target = LOGS_DIR / filename
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    entries = []
    try:
        with open(target, "r", encoding="utf-8") as f:
            if filename.endswith(".jsonl"):
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entries.append(json.loads(line))
                    except json.JSONDecodeError:
                        entries.append({"_raw": line})
            elif filename.endswith(".log"):
                # .log files contain pretty-printed JSON blocks separated by blank lines
                buf = []
                for line in f:
                    stripped = line.strip()
                    if stripped:
                        buf.append(stripped)
                    else:
                        if buf:
                            try:
                                entries.append(json.loads(" ".join(buf)))
                            except json.JSONDecodeError:
                                entries.append({"_raw": "\n".join(buf)})
                            buf = []
                if buf:
                    try:
                        entries.append(json.loads(" ".join(buf)))
                    except json.JSONDecodeError:
                        entries.append({"_raw": "\n".join(buf)})
            else:
                text = f.read()
                return {"filename": filename, "format": "text", "text": text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading file: {e}")

    return {"filename": filename, "format": "entries", "count": len(entries), "entries": entries}


@app.delete("/api/logs/{filename}")
async def delete_log_file(filename: str, session: dict = Depends(require_role("superuser"))):
    """Delete a specific log file (superuser only)."""
    # Sanitize: prevent directory traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    target = LOGS_DIR / filename
    if not target.exists():
        raise HTTPException(status_code=404, detail="File not found")
    if not target.is_file():
        raise HTTPException(status_code=400, detail="Not a file")

    # If it's an active conversation log, also clear the cached logger
    for agent_id in list(_agent_loggers.keys()):
        if filename == f"{agent_id}_conversations.log":
            logger = _agent_loggers.pop(agent_id)
            for h in logger.handlers[:]:
                h.close()
                logger.removeHandler(h)
            break

    target.unlink()
    return {"ok": True, "message": f"Deleted {filename}"}


@app.post("/api/agents/{agent_id}/init")
async def init_agent(agent_id: str):
    """
    Initialize an agent, forcing ChromaDB indexing for RAG agents.
    Call this when selecting a RAG agent to ensure the database is ready.
    """
    result = runner.init_agent(agent_id)
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "Unknown error"))
    return result


@app.get("/api/agents/{agent_id}/init-stream")
async def init_agent_stream(agent_id: str):
    """
    Initialize a RAG agent with SSE progress reporting.
    Streams progress events during document indexing.
    """
    import asyncio
    import threading

    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        err = format_error(AGENT_NOT_FOUND, agent_id=agent_id)
        raise HTTPException(status_code=404, detail=f"Error {err['error_code']}: {err['error']}")

    async def event_generator():
        loop = asyncio.get_event_loop()
        aq = asyncio.Queue()

        def progress_callback(current, total, filename):
            loop.call_soon_threadsafe(aq.put_nowait, {
                "event": "progress",
                "current": current,
                "total": total,
                "filename": filename,
            })

        def run_init():
            try:
                result = runner.init_agent_with_callback(agent_id, progress_callback)
                loop.call_soon_threadsafe(aq.put_nowait, {"event": "done", "result": result})
            except Exception as e:
                loop.call_soon_threadsafe(aq.put_nowait, {
                    "event": "done",
                    "result": {"success": False, "error": str(e)},
                })

        thread = threading.Thread(target=run_init, daemon=True)
        thread.start()

        while True:
            event = await aq.get()
            if event["event"] == "progress":
                yield f"event: progress\ndata: {json.dumps(event)}\n\n"
            elif event["event"] == "done":
                yield f"event: done\ndata: {json.dumps(event['result'])}\n\n"
                break

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/agents/{agent_id}/reindex")
async def reindex_agent(agent_id: str):
    """
    Force reindex of a RAG agent's documents.
    Use this after adding, removing, or modifying documents in data/docs/.
    """
    result = runner.reindex_agent(agent_id)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "Unknown error"))
    return result


@app.get("/api/agents/{agent_id}/token-usage")
async def get_token_usage(agent_id: str):
    """
    Get token usage statistics for an agent.
    Returns prompt tokens, completion tokens, and total tokens for the session.
    """
    result = runner.get_agent_token_usage(agent_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Agent not loaded or doesn't support token tracking")
    return result


@app.post("/api/agents/{agent_id}/reset-token-usage")
async def reset_token_usage(agent_id: str):
    """
    Reset token usage counters for an agent.
    """
    result = runner.reset_agent_token_usage(agent_id)
    if not result:
        raise HTTPException(status_code=404, detail="Agent not loaded or doesn't support token tracking")
    return {"success": True, "message": "Token usage counters reset"}


def _disable_web_search_for_user(agent_id: str, session: dict | None) -> tuple:
    """
    If the user has role 'user', temporarily disable web search on the agent
    to prevent data from leaving UMA infrastructure.
    Returns (agent_instance, original_web_search_config) for later restoration.
    """
    if not session or session["role"] != "user":
        return None, None
    agent_instance = runner._load_agent_module(agent_id)
    if not agent_instance:
        return None, None
    config = getattr(agent_instance, '_config', None)
    if not config or "web_search" not in config:
        return None, None
    original = config["web_search"].copy()
    config["web_search"]["google_api_key"] = ""
    config["web_search"]["google_cx"] = ""
    return agent_instance, original


def _restore_web_search(agent_instance, original_config):
    """Restore web search config after request."""
    if agent_instance and original_config:
        agent_instance._config["web_search"] = original_config


@app.post("/api/chat", response_model=ChatResponse)
async def chat(request: ChatRequest, http_request: Request):
    """Envía un mensaje a un agente y obtiene respuesta completa"""
    # Block users from accessing cloud-provider agents
    token = _get_token(http_request)
    session = get_session(token) if token else None
    if session and session["role"] == "user" and not _is_local_provider(request.agent_id):
        raise HTTPException(status_code=403, detail="Access restricted: this agent uses a cloud LLM provider")
    # Disable web search for end users (data sovereignty)
    agent_inst, orig_ws = _disable_web_search_for_user(request.agent_id, session)
    try:
        result = await runner.run_query(
            agent_id=request.agent_id,
            message=request.message,
            session_id=request.session_id  # None en primera llamada
        )
        return ChatResponse(response=result.response, session_id=result.session_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _restore_web_search(agent_inst, orig_ws)


@app.get("/api/chat/stream")
async def chat_stream(
    request: Request,
    agent_id: str = Query(..., description="ID del agente"),
    message: str = Query(..., description="Mensaje a enviar"),
    session_id: Optional[str] = Query(None, description="ID de sesión (opcional, se crea automáticamente)"),
    model: Optional[str] = Query(None, description="LLM model override (client preference)"),
    transparency: Optional[str] = Query(None, description="Transparency level override (client preference)"),
    prompt_level: Optional[str] = Query(None, description="Prompt level override (client preference)"),
    study_query_number: Optional[int] = Query(None, description="Study query number (1-8) for study participants"),
):
    """Envía un mensaje y hace streaming de la respuesta via SSE"""
    # Block users from accessing cloud-provider agents
    token = _get_token(request)
    session = get_session(token) if token else None
    if session and session["role"] == "user" and not _is_local_provider(agent_id):
        raise HTTPException(status_code=403, detail="Access restricted: this agent uses a cloud LLM provider")
    # Disable web search for end users (data sovereignty)
    agent_inst, orig_ws = _disable_web_search_for_user(agent_id, session)

    agent = runner.get_agent(agent_id)
    if not agent:
        err = format_error(AGENT_NOT_FOUND, agent_id=agent_id)
        raise HTTPException(status_code=404, detail=f"Error {err['error_code']}: {err['error']}")

    client_ip = request.client.host if request.client else "unknown"

    async def event_generator():
        new_session_id = None
        full_response = ""

        try:
            # Build study_info if this is a study participant query
            _study_info = None
            if session and study_query_number:
                from auth import get_study_info as _get_si
                _si = _get_si(session["username"])
                if _si:
                    _study_info = {**_si, "query_number": study_query_number}

            async for event_type, content, returned_session_id in runner.run_query_stream(
                agent_id=agent_id,
                message=message,
                session_id=session_id,  # None en primera llamada
                model_override=model,
                transparency_override=transparency,
                prompt_level_override=prompt_level,
                username=session["username"] if session else None,
                role=session["role"] if session else None,
                study_info=_study_info,
            ):
                # Enviar session_id cuando lo recibimos (primera iteración)
                if returned_session_id and not new_session_id:
                    new_session_id = returned_session_id
                    yield f"event: session\ndata: {new_session_id}\n\n"

                if event_type == "status":
                    # Enviar evento de estado
                    yield f"event: status\ndata: {content}\n\n"
                elif event_type == "badge":
                    # Enviar badge de fiabilidad como evento separado (no se acumula en el texto)
                    escaped = content.replace("\n", "\\n")
                    yield f"event: badge\ndata: {escaped}\n\n"
                elif event_type == "trace":
                    # Decision trace — raw HTML, bypasses markdown
                    escaped = content.replace("\n", "\\n")
                    yield f"event: trace\ndata: {escaped}\n\n"
                elif event_type == "editor":
                    # Raw HTML editor widget — sent as-is, bypasses markdown
                    escaped = content.replace("\n", "\\n")
                    yield f"event: editor\ndata: {escaped}\n\n"
                elif event_type == "procedural_banner":
                    # Procedural banner — separate event, persists through replace events
                    escaped = content.replace("\n", "\\n")
                    yield f"event: procedural_banner\ndata: {escaped}\n\n"
                elif event_type == "replace_banner":
                    escaped = content.replace("\n", "\\n")
                    yield f"event: replace_banner\ndata: {escaped}\n\n"
                elif event_type == "replace":
                    # Replace full response (e.g. after stripping map links)
                    full_response = content
                    escaped = content.replace("\n", "\\n")
                    yield f"event: replace\ndata: {escaped}\n\n"
                else:
                    # Enviar contenido
                    full_response += content
                    escaped = content.replace("\n", "\\n")
                    yield f"data: {escaped}\n\n"

            # Registrar la conversación en el log
            log_conversation(
                client_ip=client_ip,
                agent_id=agent_id,
                agent_name=agent.name,
                question=message,
                response=full_response,
                session_id=new_session_id or session_id or "",
                transparency_level=transparency,
                username=session["username"] if session else None,
            )

            yield "event: done\ndata: complete\n\n"
        except Exception as e:
            err = format_error(SERVER_STREAMING_ERROR, details=str(e))
            import json
            yield f"event: error\ndata: {json.dumps(err)}\n\n"
        finally:
            _restore_web_search(agent_inst, orig_ws)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


# ============================================================================
# Topic Map Endpoint (for agents with search_papers_by_topic)
# ============================================================================

@app.get("/api/agents/{agent_id}/publications-search")
async def agent_publications_search(agent_id: str, year: Optional[int] = Query(None)):
    """Return all papers grouped by university as JSON data."""
    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "get_all_papers_by_university"):
        raise HTTPException(status_code=400, detail="Agent does not support publications search")
    results = agent_instance.get_all_papers_by_university(year=year)
    title = f"Publications ({year})" if year else "All Publications"
    return {"topic": title, "universities": results}


@app.get("/api/agents/{agent_id}/publications-map")
async def agent_publications_map(agent_id: str, year: Optional[int] = Query(None)):
    """Returns an interactive Leaflet map showing publications per university, optionally filtered by year."""
    from fastapi.responses import HTMLResponse
    import json as json_module

    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "get_all_papers_by_university"):
        raise HTTPException(status_code=400, detail="Agent does not support publications search")

    results = agent_instance.get_all_papers_by_university(year=year)
    results_json = json_module.dumps(results)
    title = f"Publications ({year})" if year else "All Publications"

    if hasattr(agent_instance, "build_topic_map_html"):
        html = agent_instance.build_topic_map_html(results_json, title)
        return HTMLResponse(content=html)

    return HTMLResponse(content=f"<html><body><h1>All Publications</h1><pre>{results_json}</pre></body></html>")


@app.get("/api/agents/{agent_id}/collaboration-search")
async def agent_collaboration_search(agent_id: str, topic: str = Query(None),
                                     year: int = Query(None)):
    """Return collaboration data (universities + connections) as JSON."""
    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "get_collaboration_map_data"):
        raise HTTPException(status_code=400, detail="Agent does not support collaboration search")
    return agent_instance.get_collaboration_map_data(topic=topic, year=year)


@app.get("/api/agents/{agent_id}/topic-search")
async def agent_topic_search(agent_id: str, topic: str = Query(...)):
    """Search papers by topic across universities. Returns JSON data."""
    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "search_papers_by_topic"):
        raise HTTPException(status_code=400, detail="Agent does not support topic search")
    results = agent_instance.search_papers_by_topic(topic)
    return {"topic": topic, "universities": results}


@app.get("/api/agents/{agent_id}/topic-map")
async def agent_topic_map(agent_id: str, topic: str = Query(...)):
    """Returns an interactive Leaflet map for a topic across UNINOVIS universities."""
    from fastapi.responses import HTMLResponse
    import json as json_module

    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "search_papers_by_topic"):
        raise HTTPException(status_code=400, detail="Agent does not support topic search")

    results = agent_instance.search_papers_by_topic(topic)
    results_json = json_module.dumps(results)
    topic_escaped = topic.replace("&", "&amp;").replace("<", "&lt;").replace('"', "&quot;")

    # Use the agent's build_topic_map_html if available
    if hasattr(agent_instance, "build_topic_map_html"):
        html = agent_instance.build_topic_map_html(results_json, topic_escaped)
        return HTMLResponse(content=html)

    # Fallback: simple response
    return HTMLResponse(content=f"<html><body><h1>Topic map for {topic_escaped}</h1><pre>{results_json}</pre></body></html>")


@app.get("/api/agents/{agent_id}/projects-search")
async def agent_projects_search(agent_id: str, year: Optional[int] = Query(None)):
    """Return all projects grouped by university as JSON data, optionally filtered by year."""
    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "get_all_projects_by_university"):
        raise HTTPException(status_code=400, detail="Agent does not support projects search")
    results = agent_instance.get_all_projects_by_university(year=year)
    return {"topic": "All Projects", "universities": results}


@app.get("/api/agents/{agent_id}/project-topic-search")
async def agent_project_topic_search(agent_id: str, topic: str = Query(...),
                                     year: Optional[int] = Query(None)):
    """Search projects by topic across universities, optionally filtered by year. Returns JSON data."""
    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "search_projects_by_topic"):
        raise HTTPException(status_code=400, detail="Agent does not support project topic search")
    results = agent_instance.search_projects_by_topic(topic, year=year)
    return {"topic": topic, "universities": results}


@app.get("/api/agents/{agent_id}/projects-map")
async def agent_projects_map(agent_id: str, year: Optional[int] = Query(None)):
    """Returns an interactive Leaflet map showing projects per university, optionally filtered by year."""
    from fastapi.responses import HTMLResponse
    import json as json_module

    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "get_all_projects_by_university"):
        raise HTTPException(status_code=400, detail="Agent does not support projects search")

    results = agent_instance.get_all_projects_by_university(year=year)
    results_json = json_module.dumps(results)
    title = f"Projects ({year})" if year else "All Projects"

    if hasattr(agent_instance, "build_project_map_html"):
        html = agent_instance.build_project_map_html(results_json, title)
        return HTMLResponse(content=html)

    return HTMLResponse(content=f"<html><body><h1>{title}</h1><pre>{results_json}</pre></body></html>")


@app.get("/api/agents/{agent_id}/project-topic-map")
async def agent_project_topic_map(agent_id: str, topic: str = Query(...),
                                  year: Optional[int] = Query(None)):
    """Returns an interactive Leaflet map for projects on a topic, optionally filtered by year."""
    from fastapi.responses import HTMLResponse
    import json as json_module

    agent_instance = runner.get_agent_instance(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not loaded")
    if not hasattr(agent_instance, "search_projects_by_topic"):
        raise HTTPException(status_code=400, detail="Agent does not support project topic search")

    results = agent_instance.search_projects_by_topic(topic, year=year)
    results_json = json_module.dumps(results)
    topic_escaped = topic.replace("&", "&amp;").replace("<", "&lt;").replace('"', "&quot;")
    title = f"{topic_escaped} ({year})" if year else topic_escaped

    if hasattr(agent_instance, "build_project_map_html"):
        html = agent_instance.build_project_map_html(results_json, title)
        return HTMLResponse(content=html)

    return HTMLResponse(content=f"<html><body><h1>Projects on {title}</h1><pre>{results_json}</pre></body></html>")


# ============================================================================
# Algoria Map — Agreements map endpoint
# ============================================================================

@app.get("/api/agents/{agent_id}/agreements-config")
async def agent_agreements_config(agent_id: str):
    """Returns public config (language, translations) for the agreements map page."""
    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        raise HTTPException(status_code=404, detail="Agent not found")
    config_path = Path(agent_info.path) / "config.json"
    if not config_path.exists():
        return {}
    import json as json_module
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = json_module.load(f)
    # Only expose language and translations (no sensitive data)
    return {"language": cfg.get("language", "en"), "translations": cfg.get("translations", {})}


@app.get("/api/agents/{agent_id}/agreements-search")
async def agent_agreements_search(
    request: Request,
    agent_id: str,
    q: Optional[str] = Query(None),
    continent: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    faculty: Optional[str] = Query(None),
    language: Optional[str] = Query(None),
    language_level: Optional[str] = Query(None),
    mobility_program: Optional[str] = Query(None),
    degree_level: Optional[str] = Query(None),
    university: Optional[str] = Query(None),
    uninovis: Optional[bool] = Query(None),
    session_id: Optional[str] = Query(None),
):
    """Returns JSON data for the agreements map (markers, center, zoom)."""
    agent_instance = runner.get_agent_instance(agent_id) or runner._load_agent_module(agent_id)
    if not agent_instance:
        raise HTTPException(status_code=404, detail="Agent not found")
    if not hasattr(agent_instance, "get_map_data"):
        raise HTTPException(status_code=400, detail="Agent does not support agreements search")

    filters = {}
    if continent: filters["continent"] = continent
    if country: filters["country"] = country
    if faculty: filters["faculty"] = faculty
    if language: filters["language"] = language
    if language_level: filters["language_level"] = language_level
    if mobility_program: filters["mobility_program"] = mobility_program
    if degree_level: filters["degree_level"] = degree_level
    if university: filters["university"] = university
    if uninovis: filters["uninovis"] = True

    result = agent_instance.get_map_data(filters)

    # Log the interaction
    client_ip = request.client.host if request.client else "unknown"
    n_markers = len(result.get("markers", []))
    total = sum(m.get("count", 0) for m in result.get("markers", []))
    if q:
        question = f"[NL] {q} → {json.dumps(filters, ensure_ascii=False)}"
    elif filters:
        question = f"[filters] {json.dumps(filters, ensure_ascii=False)}"
    else:
        question = "[init] all agreements"
    response = f"{total} agreements, {n_markers} markers"
    log_conversation(
        client_ip=client_ip,
        agent_id=agent_id,
        agent_name="Algoria Map",
        question=question,
        response=response,
        session_id=session_id or "",
    )

    return result


@app.post("/api/agents/{agent_id}/interaction-log")
async def agent_interaction_log(request: Request, agent_id: str):
    """Log a UI interaction (click on marker, cluster, detail) without a search."""
    if not ENABLE_LOGGING:
        return {"ok": True}
    body = await request.json()
    action = str(body.get("action", "unknown"))[:100]
    details = str(body.get("details", ""))[:500]
    session_id = str(body.get("session_id", ""))[:64]

    # Referer: keep only the domain
    referer = request.headers.get("referer", "")
    try:
        from urllib.parse import urlparse
        referer_domain = urlparse(referer).hostname or "directo"
    except Exception:
        referer_domain = "directo"

    # User-Agent: classify into category
    ua = (request.headers.get("user-agent") or "").lower()
    if any(b in ua for b in ("googlebot", "bingbot", "yandex", "baidu", "crawler", "spider")):
        client_type = "bot"
    elif any(s in ua for s in ("python", "curl", "wget", "httpie", "postman", "scrapy")):
        client_type = "script"
    elif any(br in ua for br in ("mozilla", "chrome", "safari", "firefox", "edge", "opera")):
        client_type = "browser"
    else:
        client_type = "unknown"

    log_conversation(
        client_ip="proxy",
        agent_id=agent_id,
        agent_name="Algoria Map",
        question=f"[interaction] {action}",
        response=str(details)[:500],
        session_id=session_id,
        extra={"referer_domain": referer_domain, "client_type": client_type},
    )
    return {"ok": True}


# ============================================================================
# PDF Document Endpoint
# ============================================================================

# ============================================================================
# Responsible AI Research Explorer — Filter & Search endpoints
# ============================================================================

_RAI_DIR = AGENTS_PATH / "responsible_ai3"

@app.get("/api/public-agent/responsible_ai3/filters")
async def rai_get_filters():
    """Get filterable fields for the Responsible AI research explorer."""
    meta_path = _RAI_DIR / "data" / "metadata.json"
    if not meta_path.exists():
        return {"universities": [], "years": [], "types": [], "topics": []}

    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)

    universities = []
    years = set()
    types = set()
    topic_counts = {}

    # Topics to exclude (generic/irrelevant OpenAlex concepts)
    _EXCLUDE_TOPICS = {
        # Misclassified
        "key (lock)", "context (archaeology)", "feature (linguistics)",
        "pattern recognition (psychology)", "transparency (behavior)",
        "process (computing)", "generalizability theory", "work (physics)",
        "field (mathematics)", "generative grammar", "perception",
        # Generic academic disciplines
        "computer science", "psychology", "business", "medicine",
        "sociology", "political science", "marketing", "engineering",
        "mathematics", "philosophy", "economics", "biology", "chemistry",
        "physics", "law", "education", "history", "geography", "linguistics",
        "public relations", "management science", "medical education",
        "algorithm",
        # Generic tech terms (too broad)
        "artificial intelligence", "machine learning", "data science",
        "data mining", "big data", "deep learning", "random forest",
        "support vector machine", "software deployment",
        "applications of artificial intelligence",
        "knowledge management", "health care", "health informatics",
        "higher education", "corporate governance",
        # More noise
        "affect (linguistics)", "classifier (uml)", "curriculum",
        "discriminative model", "domain (mathematical analysis)",
        "feature selection", "informatics", "internal medicine",
        "oncology", "perspective (graphical)", "pipeline (software)",
        "tourism", "yearbook", "risk analysis (engineering)",
        "artificial neural network", "inertial measurement unit",
        "information retrieval", "narrative", "pedagogy",
        "relevance (law)", "wearable computer", "epistemology",
        "equity (law)", "confidentiality",
        "computational biology", "distributed computing",
        "electroencephalography", "ensemble learning",
        "identification (biology)", "population",
        "qualitative research", "set (abstract data type)",
        "task (project management)", "workflow",
        "european union", "robotics",
    }

    for uid, udata in meta.get("universities", {}).items():
        universities.append({"id": uid, "name": udata.get("name", uid)})
        for paper in udata.get("papers", []):
            if paper.get("publication_year"):
                years.add(paper["publication_year"])
            if paper.get("type"):
                types.add(paper["type"])
            for concept in paper.get("concepts", []):
                cname = concept.get("name", "")
                if cname and concept.get("score", 0) > 0.3 and cname.lower() not in _EXCLUDE_TOPICS:
                    topic_counts[cname] = topic_counts.get(cname, 0) + 1

    # Top topics by frequency (from OpenAlex concepts, filtered, min 4 papers)
    top_topics = [(n, c) for n, c in sorted(topic_counts.items(), key=lambda x: -x[1]) if c >= 4][:15]

    # Also add curated Responsible AI topics from config
    config_path = _RAI_DIR / "config.json"
    curated_terms = set()
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        curated_terms = set(t.lower() for t in cfg.get("extra_scope_terms", []))

    # Terms to skip from curated list (generic, duplicates, internal)
    _SKIP_CURATED = {
        "uninovis", "llm", "language model", "large language model",
        # Duplicates (keep the better variant)
        "data minimisation",        # keep "data minimization"
        "ai standards",             # keep "ai standardization"
        "ai existential",           # keep "existential risk"
        "lethal autonomous",        # keep "autonomous weapons"
        "red teaming",              # keep "ai red-teaming"
        "ai and labor",             # keep "ai and employment"
        "algorithmic decision",     # keep "ai decision-making"
        "ai colonialism",           # keep "data colonialism"
        "ai and copyright",         # keep "ai and intellectual property"
        "value alignment",          # keep "ai alignment"
        "ai sentience",             # keep "ai consciousness"
        "agi safety",               # keep "artificial general intelligence"
        "algorithmic impact",       # keep "ai impact assessment"
        "energy consumption of ai", # keep "carbon footprint of ai"
        "ai legislation",           # keep "ai regulation"
        "ai policy",                # keep "ai regulation"
    }

    # Merge: OpenAlex topics + curated terms (dedup, capitalize)
    seen = set(t[0].lower() for t in top_topics)
    for term in sorted(curated_terms):
        if term not in seen and term not in _SKIP_CURATED:
            top_topics.append((term.title(), 0))
            seen.add(term)

    return {
        "universities": sorted(universities, key=lambda u: u["name"]),
        "years": sorted(years, reverse=True),
        "types": sorted(types),
        "topics": [{"name": t[0], "count": t[1]} for t in top_topics]
    }


@app.get("/api/public-agent/responsible_ai3/search")
async def rai_search(
    request: Request,
    university: str = Query("", description="University ID filter"),
    year: str = Query("", description="Publication year filter"),
    year_op: str = Query("=", description="Year operator: =, ≥, ≤"),
    pub_type: str = Query("", description="Publication type filter"),
    topic: str = Query("", description="Topic/concept filter"),
    author: str = Query("", description="Author name filter"),
    nl_query: str = Query("", description="Original natural language query (if routed from NL input)"),
):
    """Search research papers with filters."""
    meta_path = _RAI_DIR / "data" / "metadata.json"
    if not meta_path.exists():
        return {"results": [], "total": 0}

    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)

    # Build affiliation-to-university mapping
    _uni_name_map = {}
    for _uid, _udata in meta.get("universities", {}).items():
        _uni_name_map[_uid] = _udata.get("name", _uid)
    _aff_keywords = {
        "UMA": ["universidad de málaga", "universidad de malaga", "university of malaga"],
        "USPN": ["sorbonne paris nord"],
        "UDCLV": ["campania \"luigi vanvitelli\"", "campania luigi vanvitelli", "vanvitelli"],
        "KK": ["kauno kolegija"],
        "UT": ["university of tirana"],
        "THWS": ["technische hochschule würzburg", "technische hochschule wurzburg", "thws", "technical university of applied sciences würzburg", "technical university of applied sciences wurzburg"],
        "TAMK": ["tampere university of applied sciences"],
        "THUAS": ["hague university of applied sciences", "haagse hogeschool"],
    }

    def _match_affiliations(affiliations):
        """Match affiliations to UNINOVIS university IDs."""
        matched = set()
        for aff in affiliations:
            aff_lower = aff.lower()
            for uid, keywords in _aff_keywords.items():
                if any(kw in aff_lower for kw in keywords):
                    matched.add(uid)
        return matched

    results = []
    for uid, udata in meta.get("universities", {}).items():
        uni_name = udata.get("name", uid)
        for paper in udata.get("papers", []):
            # Year filter with operator
            if year:
                try:
                    y_filter = int(year)
                    y_paper = int(paper.get("publication_year", 0))
                    if year_op == "≥" and y_paper < y_filter:
                        continue
                    elif year_op == "≤" and y_paper > y_filter:
                        continue
                    elif year_op == "=" and y_paper != y_filter:
                        continue
                except (ValueError, TypeError):
                    pass
            # Type filter
            if pub_type and paper.get("type", "") != pub_type:
                continue
            # Topic filter (substring match on concept names and paper title/abstract)
            if topic:
                topic_lower = topic.lower()
                concepts = " ".join(c.get("name", "") for c in paper.get("concepts", [])).lower()
                title = paper.get("title", "").lower()
                abstract = paper.get("abstract", "").lower()
                if topic_lower not in concepts and topic_lower not in title and topic_lower not in abstract:
                    continue
            # Author filter
            if author:
                author_lower = author.lower()
                author_names = [a.get("name", "").lower() for a in paper.get("authors", [])]
                if not any(author_lower in name for name in author_names):
                    continue

            # Build result
            authors = paper.get("authors", [])
            author_names = ", ".join(a.get("name", "") for a in authors[:4])
            if len(authors) > 4:
                author_names += f" et al. ({len(authors)} authors)"

            concepts_list = [c.get("name", "") for c in paper.get("concepts", [])
                            if c.get("score", 0) > 0.3]

            # Detect all affiliated UNINOVIS universities
            affiliated_unis = {uid}
            affiliated_unis.update(_match_affiliations(paper.get("affiliations", [])))

            # University filter: check if the selected university is among the affiliated ones
            if university and university not in affiliated_unis:
                continue

            results.append({
                "id": paper.get("id", ""),
                "title": paper.get("title", ""),
                "authors": author_names,
                "university": ", ".join(_uni_name_map.get(u, u) for u in sorted(affiliated_unis)),
                "university_id": ", ".join(sorted(affiliated_unis)),
                "year": paper.get("publication_year", ""),
                "type": paper.get("type", ""),
                "doi": paper.get("doi", ""),
                "cited_by": paper.get("cited_by_count", 0),
                "topics": concepts_list[:5],
            })

    # Deduplicate by paper ID (same paper may appear under multiple universities)
    seen = {}
    deduped = []
    for r in results:
        pid = r.get("id", "")
        if pid in seen:
            # Merge university info
            existing = deduped[seen[pid]]
            if r["university_id"] not in existing["university_id"]:
                existing["university_id"] += ", " + r["university_id"]
                existing["university"] += ", " + r["university"]
        else:
            seen[pid] = len(deduped)
            deduped.append(r)

    # Sort by year desc, then citations desc
    deduped.sort(key=lambda r: (-r.get("year", 0), -r.get("cited_by", 0)))

    # Log the filter search
    filters_used = []
    if university: filters_used.append(f"university={university}")
    if year: filters_used.append(f"year{year_op}{year}")
    if pub_type: filters_used.append(f"type={pub_type}")
    if topic: filters_used.append(f"topic={topic}")
    if author: filters_used.append(f"author={author}")
    search_type = "nl" if nl_query else "filter"
    query_str = "[" + search_type + "] " + ", ".join(filters_used) if filters_used else "[" + search_type + "] all"
    if nl_query:
        query_str = f"[nl] \"{nl_query}\" → {', '.join(filters_used)}"

    client_ip = request.headers.get("x-forwarded-for", request.client.host if request.client else "proxy")
    log_conversation(
        client_ip=client_ip,
        agent_id="responsible_ai3",
        agent_name="EH: Responsible AI",
        question=query_str,
        response=f"{len(deduped)} results",
        extra={
            "search_type": search_type,
            "result_count": len(deduped),
            "nl_query": nl_query or None,
            "resolved_as": ", ".join(filters_used) if nl_query else None,
        },
    )

    return {"results": deduped, "total": len(deduped)}


@app.get("/api/agents/{agent_id}/pdf-list")
async def agent_pdf_list(agent_id: str):
    """List available PDFs for an agent."""
    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        raise HTTPException(status_code=404, detail="Agent not found")
    docs_dir = Path(agent_info.path) / "data" / "docs"
    if not docs_dir.exists():
        return {"pdfs": []}
    pdfs = [f.stem for f in docs_dir.glob("*.pdf")]
    return {"pdfs": pdfs}


@app.get("/api/agents/{agent_id}/pdf/{filename}")
async def agent_pdf(agent_id: str, filename: str):
    """Serve a PDF from an agent's data/docs/ directory."""
    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        raise HTTPException(status_code=404, detail="Agent not found")

    # Security: only allow .pdf files, no path traversal
    if not filename.endswith(".pdf") or "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    pdf_path = Path(agent_info.path) / "data" / "docs" / filename
    if not pdf_path.exists():
        raise HTTPException(status_code=404, detail="PDF not found")

    return FileResponse(pdf_path, media_type="application/pdf", filename=filename)


# ============================================================================
# Agent Quick Guide Endpoint
# ============================================================================

@app.head("/api/agents/{agent_id}/quickguide")
@app.get("/api/agents/{agent_id}/quickguide")
async def agent_quickguide(agent_id: str):
    """Serve the quick guide HTML for an agent (if available)."""
    agent_info = runner.get_agent(agent_id)
    if not agent_info:
        raise HTTPException(status_code=404, detail="Agent not found")

    # Look for agent-specific quickguide, then fall back to the legacy name
    agent_dir = Path(agent_info.path)
    guide_path = None
    for candidate in sorted(agent_dir.glob("*quickguide*.html")):
        guide_path = candidate
        break
    if not guide_path or not guide_path.exists():
        raise HTTPException(status_code=404, detail="Quick guide not found")

    return FileResponse(guide_path, media_type="text/html")


# ============================================================================
# Agent Creation Endpoints
# ============================================================================

@app.get("/create-agent")
async def create_agent_page():
    """Serve the create agent page"""
    return FileResponse(SCRIPT_DIR / "static" / "create_agent.html")


@app.get("/api/prompt-templates")
async def get_prompt_templates(agent_type: str = Query(...)):
    """List available prompt templates for an agent type"""
    # Import from crear_agente
    apps_dir = SCRIPT_DIR.parent / "apps"
    sys.path.insert(0, str(apps_dir))

    try:
        from crear_agente import list_prompt_templates
        templates = list_prompt_templates(agent_type)
        return [{"name": name, "path": path} for name, path in templates]
    except Exception as e:
        return []


@app.get("/api/prompt-template")
async def get_prompt_template(path: str = Query(...)):
    """Get content of a prompt template"""
    try:
        # Security: ensure path is within prompts directory
        prompts_dir = SCRIPT_DIR.parent / "prompts"
        template_path = Path(path).resolve()

        if not str(template_path).startswith(str(prompts_dir.resolve())):
            raise HTTPException(status_code=403, detail="Access denied")

        with open(template_path, "r", encoding="utf-8") as f:
            content = f.read()
        return {"content": content}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Template not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


from fastapi import Form, UploadFile, File
from typing import List, Optional

@app.post("/api/create-agent")
async def create_agent(
    agent_type: str = Form(...),
    agent_id: str = Form(...),
    agent_name: str = Form(...),
    description: str = Form(""),
    welcome_message: str = Form(""),
    examples: str = Form("[]"),  # JSON array
    system_prompt: str = Form(...),
    llm_provider: str = Form("default"),
    mistral_model: str = Form("mistral-large-latest"),
    mistral_api_key: str = Form(""),
    ollama_url: str = Form("http://localhost:11434"),
    ollama_model: str = Form(""),
    context_preserving: bool = Form(True),  # RAG chunking approach
    reliability_green_max_llm: int = Form(20),
    reliability_red_min_llm: int = Form(50),
    database_schema: str = Form(""),
    data_file: Optional[UploadFile] = File(None),
    schema_file: Optional[UploadFile] = File(None),
    rag_documents: List[UploadFile] = File(None),
    rag_vectorless_documents: List[UploadFile] = File(None),
    rag_metadata_documents: List[UploadFile] = File(None),
    metadata_file: Optional[UploadFile] = File(None),
    database_file: Optional[UploadFile] = File(None),
):
    """Create a new agent"""
    import json as json_module
    import shutil

    try:
        # Parse examples
        example_list = json_module.loads(examples) if examples else []

        # Validate agent_id
        if not agent_id or " " in agent_id or not agent_id.replace("_", "").replace("-", "").isalnum():
            raise HTTPException(status_code=400, detail="Invalid agent ID. Use only lowercase letters, numbers, hyphens or underscores.")

        # Check if agent already exists
        output_dir = AGENTS_PATH / agent_id
        if output_dir.exists():
            raise HTTPException(status_code=400, detail=f"Agent '{agent_id}' already exists")

        # Import crear_agente functions
        apps_dir = SCRIPT_DIR.parent / "apps"
        sys.path.insert(0, str(apps_dir))
        from crear_agente import (
            create_agent_structure,
            get_agents_dir
        )

        # Prepare configuration
        config = {
            "agent_type": agent_type,
            "agent_id": agent_id,
            "output_dir": str(output_dir),
            "agent_name": agent_name,
            "description": description or f"{agent_name} Assistant",
            "welcome": welcome_message or f"Hello! I'm {agent_name}. How can I help you?",
            "examples": example_list,
            "system_prompt": system_prompt,
            "llm_provider": llm_provider,
            "model": mistral_model if llm_provider == "mistral" else ollama_model,
            "api_key": mistral_api_key if llm_provider == "mistral" else "",
            "ollama_url": ollama_url if llm_provider == "ollama" else "",
            "ollama_model": ollama_model if llm_provider == "ollama" else "",
            "rag_approach": "context_preserving" if context_preserving else "basic",
            "reliability_green_max_llm": reliability_green_max_llm,
            "reliability_red_min_llm": reliability_red_min_llm,
        }

        # Create agent structure
        create_agent_structure(config)

        # Create data directory
        data_dir = output_dir / "data"
        data_dir.mkdir(exist_ok=True)

        # Handle data based on agent type
        if agent_type == "oneshot" and data_file and data_file.filename:
            # Save uploaded data.md file
            content = await data_file.read()
            with open(data_dir / "data.md", "wb") as f:
                f.write(content)

        elif agent_type == "rag" and rag_documents:
            # Create docs subfolder for RAG
            docs_dir = data_dir / "docs"
            docs_dir.mkdir(exist_ok=True)
            # Save uploaded documents (from folder selection)
            for doc in rag_documents:
                if doc.filename:
                    # Handle folder structure - get just the filename
                    filename = Path(doc.filename).name
                    # Skip hidden files and non-document files
                    if filename.startswith('.'):
                        continue
                    ext = Path(filename).suffix.lower()
                    if ext in ['.pdf', '.txt', '.md', '.docx', '.doc']:
                        doc_path = docs_dir / filename
                        content = await doc.read()
                        with open(doc_path, "wb") as f:
                            f.write(content)

        elif agent_type == "rag_vectorless" and rag_vectorless_documents:
            # Create docs subfolder for RAG Vectorless
            docs_dir = data_dir / "docs"
            docs_dir.mkdir(exist_ok=True)
            for doc in rag_vectorless_documents:
                if doc.filename:
                    filename = Path(doc.filename).name
                    if filename.startswith('.'):
                        continue
                    ext = Path(filename).suffix.lower()
                    if ext in ['.pdf', '.txt', '.md', '.docx', '.doc']:
                        doc_path = docs_dir / filename
                        content = await doc.read()
                        with open(doc_path, "wb") as f:
                            f.write(content)

        elif agent_type == "rag_metadata":
            # Create docs subfolder for RAG+Metadata
            docs_dir = data_dir / "docs"
            docs_dir.mkdir(exist_ok=True)
            # Save uploaded documents (from folder selection)
            if rag_metadata_documents:
                for doc in rag_metadata_documents:
                    if doc.filename:
                        filename = Path(doc.filename).name
                        if filename.startswith('.'):
                            continue
                        ext = Path(filename).suffix.lower()
                        if ext in ['.pdf', '.txt', '.md', '.docx', '.doc']:
                            doc_path = docs_dir / filename
                            content = await doc.read()
                            with open(doc_path, "wb") as f:
                                f.write(content)
            # Save metadata file if provided
            if metadata_file and metadata_file.filename:
                content = await metadata_file.read()
                with open(data_dir / "metadata.json", "wb") as f:
                    f.write(content)

        elif agent_type == "consultabd_sql":
            # Handle database schema - from file or textarea
            if schema_file and schema_file.filename:
                content = await schema_file.read()
                with open(data_dir / "database_schema.md", "wb") as f:
                    f.write(content)
            elif database_schema:
                with open(data_dir / "database_schema.md", "w", encoding="utf-8") as f:
                    f.write(database_schema)

            # Save database file
            if database_file and database_file.filename:
                db_path = data_dir / "database.db"
                content = await database_file.read()
                with open(db_path, "wb") as f:
                    f.write(content)

        # Reload agents
        runner.discover_agents()

        return {
            "success": True,
            "agent_id": agent_id,
            "path": str(output_dir),
            "message": f"Agent '{agent_name}' created successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ── User feedback on agent responses ──────────────────────────────

class FeedbackRequest(BaseModel):
    agent_id: str
    session_id: str = ""
    message_index: int = 0
    mode: str = "user"               # "user" or "tester"
    rating: str                      # "up" or "down"
    error_code: Optional[str] = None # "1.1", "1.2.1", "2.1", "3.1", etc. (tester mode)
    severity: Optional[str] = None   # "minor", "major", "critical" (tester mode)
    notes: Optional[str] = None      # comments (both modes)
    user_question: str = ""
    full_response: str = ""


@app.post("/api/feedback")
async def submit_feedback(fb: FeedbackRequest):
    """Log user feedback on an agent response.
    User and tester feedback are stored in separate per-agent files."""
    is_positive = fb.rating == "up"
    entry = {
        "timestamp": datetime.now().isoformat(),
        "feedback_type": "positive" if is_positive else "negative",
        "agent_id": fb.agent_id,
        "session_id": fb.session_id,
        "message_index": fb.message_index,
        "rating": fb.rating,
        "user_question": fb.user_question,
        "full_response": fb.full_response,
    }
    if not is_positive:
        if fb.mode == "tester":
            # Tester: structured error report with full response
            entry["error_code"] = fb.error_code
            entry["severity"] = fb.severity
        entry["notes"] = (fb.notes or "")[:1000]

    if fb.mode == "tester":
        log_file = LOGS_DIR / f"{fb.agent_id}_feedback_tester.jsonl"
    else:
        log_file = LOGS_DIR / f"{fb.agent_id}_feedback_user.jsonl"

    with open(log_file, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn
    import sys
    import socket

    print(f"Agents path: {AGENTS_PATH}")

    # Descubrir agentes al inicio
    agents = runner.discover_agents()
    print(f"Found {len(agents)} agents:")
    for agent in agents:
        print(f"  - {agent.name} ({agent.id})")

    # Verificar si el puerto está disponible antes de iniciar uvicorn
    host, port = "0.0.0.0", 8000
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind((host, port))
        sock.close()
    except OSError as e:
        if e.errno == 48:  # Address already in use
            msg = (
                f"\n*** ERROR: Port {port} is already in use. ***\n"
                f"Another server (or a previous instance of TOMMI) is using this port.\n\n"
                f"To fix this, either:\n"
                f"  1. Stop the other server, or\n"
                f"  2. Run: ./liberar-puerto.sh {port}\n"
            )
            print(msg)
            print(msg, file=sys.stderr)
            sys.exit(1)
        else:
            raise

    uvicorn.run(app, host=host, port=port)
