"""
Benchmark script for Responsible AI2 Agent.

Runs a predefined set of queries against the agent API, collects responses,
and generates an Excel report with accuracy metrics, transparency analysis,
and model comparison.

Usage:
    python benchmark.py --server http://localhost:8000 --token YOUR_TOKEN
    python benchmark.py --server http://localhost:8000 --token YOUR_TOKEN --models mistral-small-latest,mistral-large-latest
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import urlencode, quote
from urllib.error import HTTPError, URLError


# ============================================================
# Benchmark queries with expected outcomes
# ============================================================

BENCHMARK_QUERIES = [
    # Category 1: UNRELATED — agent should refuse/redirect
    {
        "id": "U01", "category": "Unrelated", "query": "What is the recipe for chocolate cake?",
        "expected_behavior": "refuse", "expected_content": None,
    },
    {
        "id": "U02", "category": "Unrelated", "query": "Who won the FIFA World Cup in 2022?",
        "expected_behavior": "refuse", "expected_content": None,
    },
    {
        "id": "U03", "category": "Unrelated", "query": "How do I change a car tire?",
        "expected_behavior": "refuse", "expected_content": None,
    },
    {
        "id": "U04", "category": "Unrelated", "query": "What is the capital of Mongolia?",
        "expected_behavior": "refuse", "expected_content": None,
    },
    {
        "id": "U05", "category": "Unrelated", "query": "How to knit a sweater?",
        "expected_behavior": "refuse", "expected_content": None,
    },

    # Category 2: SCOPE BOUNDARY — general AI, not in the database
    {
        "id": "S01", "category": "Scope Boundary", "query": "What is machine learning?",
        "expected_behavior": "partial", "expected_content": None,
    },
    {
        "id": "S02", "category": "Scope Boundary", "query": "How does blockchain technology work?",
        "expected_behavior": "refuse_or_redirect", "expected_content": None,
    },
    {
        "id": "S03", "category": "Scope Boundary", "query": "What is augmented reality?",
        "expected_behavior": "refuse_or_redirect", "expected_content": None,
    },

    # Category 3: FACTUAL — verifiable answers from metadata
    {
        "id": "F01", "category": "Factual", "query": "List all researchers from THWS",
        "expected_behavior": "answer", "expected_content": "researchers from THWS",
        "expected_min_items": 5,
    },
    {
        "id": "F02", "category": "Factual", "query": "List all researchers from KK",
        "expected_behavior": "answer", "expected_content": "researchers from KK",
        "expected_min_items": 5,
    },
    {
        "id": "F03", "category": "Factual", "query": "List all publications from UMA",
        "expected_behavior": "answer", "expected_content": "publications from UMA",
    },
    {
        "id": "F04", "category": "Factual", "query": "List all publications from UT",
        "expected_behavior": "answer", "expected_content": "publications from UT",
    },
    {
        "id": "F05", "category": "Factual", "query": "List research projects involving UMA",
        "expected_behavior": "answer", "expected_content": "projects",
    },

    # Category 4: TOPIC SEARCH — queries about specific AI topics
    {
        "id": "T01", "category": "Topic Search", "query": "List researchers that have interest in AI and Ethics",
        "expected_behavior": "answer", "expected_content": "researchers",
    },
    {
        "id": "T02", "category": "Topic Search", "query": "What research exists on explainable artificial intelligence?",
        "expected_behavior": "answer", "expected_content": "explainable",
    },
    {
        "id": "T03", "category": "Topic Search", "query": "Tell me about AI ethics and fairness in the UNINOVIS alliance",
        "expected_behavior": "answer", "expected_content": "ethics",
    },
    {
        "id": "T04", "category": "Topic Search", "query": "What papers discuss trustworthy AI and accountability?",
        "expected_behavior": "answer", "expected_content": "trustworthy",
    },
    {
        "id": "T05", "category": "Topic Search", "query": "List researchers interested in Bias and Fairness",
        "expected_behavior": "answer", "expected_content": "bias",
    },

    # Category 5: GAP ANALYSIS
    {
        "id": "G01", "category": "Gap Analysis", "query": "List the Responsible AI subtopics most studied in UNINOVIS",
        "expected_behavior": "answer", "expected_content": "topics",
    },
    {
        "id": "G02", "category": "Gap Analysis", "query": "List any topics related with responsible AI that have not been studied",
        "expected_behavior": "answer", "expected_content": "gap",
    },

    # Category 6: RESEARCHER LOOKUP
    {
        "id": "R01", "category": "Researcher Lookup", "query": "What do you know of Enrique Alba",
        "expected_behavior": "answer", "expected_content": "UMA",
    },

    # Category 7: CROSS-UNIVERSITY
    {
        "id": "X01", "category": "Cross-University", "query": "Which UNINOVIS partners have published on Explainable AI?",
        "expected_behavior": "answer", "expected_content": "partners",
    },
    {
        "id": "X02", "category": "Cross-University", "query": "Are there researchers working on AI and Healthcare across UNINOVIS?",
        "expected_behavior": "answer", "expected_content": "researchers",
    },

    # Category 8: PROJECTS
    {
        "id": "P01", "category": "Projects", "query": "List research projects on trustworthy AI",
        "expected_behavior": "answer", "expected_content": "project",
    },
    {
        "id": "P02", "category": "Projects", "query": "What projects on Trustworthy AI involve UNINOVIS partners?",
        "expected_behavior": "answer", "expected_content": "TAILOR",
    },
]


# ============================================================
# API Client
# ============================================================

def api_chat(server: str, token: str, agent_id: str, query: str, model: str = None, timeout: int = 120) -> dict:
    """Send a query to the agent and collect the full response (non-streaming)."""
    params = {
        "agent_id": agent_id,
        "message": query,
    }
    if model:
        params["model"] = model

    url = f"{server}/api/chat"
    data = json.dumps(params).encode("utf-8")
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    start = time.time()
    try:
        req = Request(url, data=data, headers=headers, method="POST")
        with urlopen(req, timeout=timeout) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        elapsed = time.time() - start
        return {
            "response": result.get("response", ""),
            "session_id": result.get("session_id", ""),
            "latency": round(elapsed, 2),
            "error": None,
        }
    except HTTPError as e:
        elapsed = time.time() - start
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        return {
            "response": "",
            "session_id": "",
            "latency": round(elapsed, 2),
            "error": f"HTTP {e.code}: {body or e.reason}",
        }
    except (URLError, Exception) as e:
        elapsed = time.time() - start
        return {
            "response": "",
            "session_id": "",
            "latency": round(elapsed, 2),
            "error": str(e),
        }


def api_chat_stream(server: str, token: str, agent_id: str, query: str, model: str = None, timeout: int = 120) -> dict:
    """Send a query via SSE streaming and collect the full response + badge."""
    params = {
        "agent_id": agent_id,
        "message": query,
        "token": token,
    }
    if model:
        params["model"] = model

    url = f"{server}/api/chat/stream?{urlencode(params)}"

    start = time.time()
    try:
        req = Request(url)
        with urlopen(req, timeout=timeout) as resp:
            full_response = ""
            badge = ""
            session_id = ""

            for raw_line in resp:
                line = raw_line.decode("utf-8").strip()
                if line.startswith("data: "):
                    data = line[6:].replace("\\n", "\n")
                    full_response += data
                elif line.startswith("event: badge"):
                    pass  # next data line will be the badge
                elif line.startswith("event: session"):
                    pass
                elif line == "event: done":
                    break
                elif "event: badge" in line:
                    pass
                # Capture badge from data after badge event
                if badge == "" and "Reliability score" in full_response:
                    badge_match = re.search(r'(Reliability score:.*?)(?:<br>|$)', full_response)
                    if badge_match:
                        badge = badge_match.group(1)

        elapsed = time.time() - start
        return {
            "response": full_response,
            "badge": badge,
            "session_id": session_id,
            "latency": round(elapsed, 2),
            "error": None,
        }
    except (HTTPError, URLError, Exception) as e:
        elapsed = time.time() - start
        return {
            "response": "",
            "badge": "",
            "session_id": "",
            "latency": round(elapsed, 2),
            "error": str(e),
        }


# ============================================================
# Analysis
# ============================================================

def analyse_response(query_def: dict, response: str) -> dict:
    """Analyse a response against expected outcomes."""
    resp_lower = response.lower()
    expected = query_def["expected_behavior"]

    result = {
        "responded": bool(response and len(response) > 20),
        "refused": False,
        "correct_behavior": False,
        "content_match": False,
        "reliability_label": "",
        "confidence": None,
        "llm_pct": None,
        "hallucination_warnings": 0,
    }

    # Detect refusal
    refusal_patterns = [
        "outside my scope", "outside my domain", "cannot help",
        "not within my", "i'm afraid", "i am afraid",
        "not related to", "beyond my scope", "outside the scope",
    ]
    result["refused"] = any(p in resp_lower for p in refusal_patterns)

    # Check correct behavior
    if expected == "refuse":
        result["correct_behavior"] = result["refused"]
    elif expected == "refuse_or_redirect":
        result["correct_behavior"] = result["refused"] or not result["responded"]
    elif expected == "partial":
        result["correct_behavior"] = result["responded"]
    elif expected == "answer":
        result["correct_behavior"] = result["responded"] and not result["refused"]

    # Check content match
    expected_content = query_def.get("expected_content")
    if expected_content and result["responded"]:
        result["content_match"] = expected_content.lower() in resp_lower

    # Extract reliability info from badge
    reliability_match = re.search(r'(High|Good|Poor)', response)
    if reliability_match:
        result["reliability_label"] = reliability_match.group(1)

    confidence_match = re.search(r'(\d+)%', response[:500])
    if confidence_match:
        result["confidence"] = int(confidence_match.group(1))

    # Count hallucination warnings
    result["hallucination_warnings"] = response.count("not found in database") + response.count("hallucin")

    return result


# ============================================================
# Excel Report
# ============================================================

def generate_report(results: list, output_path: str, models: list):
    """Generate Excel benchmark report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_header(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ---- Sheet 1: Summary per model ----
    ws1 = wb.active
    ws1.title = "Summary"
    write_header(ws1, ["Metric"] + models)

    categories = sorted(set(r["category"] for r in results))

    metrics = []
    for model in models:
        model_results = [r for r in results if r["model"] == model]
        total = len(model_results)
        correct = sum(1 for r in model_results if r["correct_behavior"])
        content = sum(1 for r in model_results if r["content_match"])
        answered = sum(1 for r in model_results if r["responded"] and r["category"] not in ("Unrelated", "Scope Boundary"))
        refused_ok = sum(1 for r in model_results if r["correct_behavior"] and r["category"] == "Unrelated")
        hallucinations = sum(r["hallucination_warnings"] for r in model_results)
        avg_latency = sum(r["latency"] for r in model_results) / max(total, 1)

        # Reliability distribution
        high = sum(1 for r in model_results if r["reliability_label"] == "High")
        good = sum(1 for r in model_results if r["reliability_label"] == "Good")
        poor = sum(1 for r in model_results if r["reliability_label"] == "Poor")

        metrics.append({
            "total": total,
            "correct_behavior": f"{correct}/{total} ({100*correct//max(total,1)}%)",
            "content_match": f"{content}/{total}",
            "scope_refusal": f"{refused_ok}/5",
            "hallucinations": hallucinations,
            "avg_latency": f"{avg_latency:.1f}s",
            "reliability_high": high,
            "reliability_good": good,
            "reliability_poor": poor,
        })

    metric_rows = [
        ("Total queries", "total"),
        ("Correct behavior", "correct_behavior"),
        ("Content match", "content_match"),
        ("Scope refusal (out of 5)", "scope_refusal"),
        ("Hallucination warnings", "hallucinations"),
        ("Average latency", "avg_latency"),
        ("Reliability: High", "reliability_high"),
        ("Reliability: Good", "reliability_good"),
        ("Reliability: Poor", "reliability_poor"),
    ]

    for row_idx, (label, key) in enumerate(metric_rows, 2):
        ws1.cell(row=row_idx, column=1, value=label).border = thin_border
        ws1.cell(row=row_idx, column=1).font = Font(bold=True)
        for mi, m in enumerate(metrics):
            ws1.cell(row=row_idx, column=2 + mi, value=str(m[key])).border = thin_border

    for c in range(1, 2 + len(models)):
        ws1.column_dimensions[chr(64 + c)].width = 22

    # ---- Sheet 2: Per-category breakdown ----
    ws2 = wb.create_sheet("By Category")
    write_header(ws2, ["Category", "Model", "Queries", "Correct", "Content Match", "Avg Latency"])

    row = 2
    for cat in categories:
        for model in models:
            cat_results = [r for r in results if r["category"] == cat and r["model"] == model]
            total = len(cat_results)
            correct = sum(1 for r in cat_results if r["correct_behavior"])
            content = sum(1 for r in cat_results if r["content_match"])
            avg_lat = sum(r["latency"] for r in cat_results) / max(total, 1)

            ws2.cell(row=row, column=1, value=cat).border = thin_border
            ws2.cell(row=row, column=2, value=model).border = thin_border
            ws2.cell(row=row, column=3, value=total).border = thin_border
            ws2.cell(row=row, column=4, value=f"{correct}/{total}").border = thin_border
            ws2.cell(row=row, column=5, value=f"{content}/{total}").border = thin_border
            ws2.cell(row=row, column=6, value=f"{avg_lat:.1f}s").border = thin_border

            fill = pass_fill if correct == total else (fail_fill if correct == 0 else warn_fill)
            ws2.cell(row=row, column=4).fill = fill
            row += 1

    for c in "ABCDEF":
        ws2.column_dimensions[c].width = 18

    # ---- Sheet 3: Detailed results ----
    ws3 = wb.create_sheet("Detailed Results")
    write_header(ws3, ["ID", "Category", "Query", "Model", "Correct?", "Content?",
                        "Reliability", "Confidence", "Hallucinations", "Latency",
                        "Response (first 300 chars)", "Error"])

    row = 2
    for r in sorted(results, key=lambda x: (x["model"], x["id"])):
        ws3.cell(row=row, column=1, value=r["id"]).border = thin_border
        ws3.cell(row=row, column=2, value=r["category"]).border = thin_border
        ws3.cell(row=row, column=3, value=r["query"]).border = thin_border
        ws3.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws3.cell(row=row, column=4, value=r["model"]).border = thin_border
        ws3.cell(row=row, column=5, value="Yes" if r["correct_behavior"] else "No").border = thin_border
        ws3.cell(row=row, column=5).fill = pass_fill if r["correct_behavior"] else fail_fill
        ws3.cell(row=row, column=6, value="Yes" if r["content_match"] else "No").border = thin_border
        ws3.cell(row=row, column=7, value=r["reliability_label"]).border = thin_border
        ws3.cell(row=row, column=8, value=r["confidence"]).border = thin_border
        ws3.cell(row=row, column=9, value=r["hallucination_warnings"]).border = thin_border
        if r["hallucination_warnings"] > 0:
            ws3.cell(row=row, column=9).fill = fail_fill
        ws3.cell(row=row, column=10, value=f"{r['latency']}s").border = thin_border
        resp_preview = r.get("response_text", "")[:300].replace("\n", " ")
        ws3.cell(row=row, column=11, value=resp_preview).border = thin_border
        ws3.cell(row=row, column=11).alignment = Alignment(wrap_text=True)
        ws3.cell(row=row, column=12, value=r.get("error", "")).border = thin_border
        row += 1

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 40
    ws3.column_dimensions["D"].width = 20
    ws3.column_dimensions["E"].width = 9
    ws3.column_dimensions["F"].width = 9
    ws3.column_dimensions["G"].width = 12
    ws3.column_dimensions["H"].width = 12
    ws3.column_dimensions["I"].width = 14
    ws3.column_dimensions["J"].width = 10
    ws3.column_dimensions["K"].width = 50
    ws3.column_dimensions["L"].width = 30

    wb.save(output_path)
    print(f"\nReport saved: {output_path}")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Benchmark Responsible AI2 Agent")
    parser.add_argument("--server", required=True, help="Server URL (e.g., http://localhost:8000)")
    parser.add_argument("--token", required=True, help="Authentication token")
    parser.add_argument("--agent", default="responsible_ai3", help="Agent ID")
    parser.add_argument("--models", default="mistral-small-latest", help="Comma-separated list of models to test")
    parser.add_argument("--output", default=None, help="Output Excel path")
    parser.add_argument("--timeout", type=int, default=120, help="Timeout per query (seconds)")
    args = parser.parse_args()

    models = [m.strip() for m in args.models.split(",")]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or f"benchmark_{args.agent}_{timestamp}.xlsx"

    print(f"Benchmark: {args.agent}")
    print(f"Server: {args.server}")
    print(f"Models: {', '.join(models)}")
    print(f"Queries: {len(BENCHMARK_QUERIES)}")
    print(f"Total runs: {len(BENCHMARK_QUERIES) * len(models)}")
    print()

    all_results = []

    for model in models:
        print(f"=== Model: {model} ===")
        for i, q in enumerate(BENCHMARK_QUERIES):
            print(f"  [{i+1}/{len(BENCHMARK_QUERIES)}] {q['id']} ({q['category']}): {q['query'][:50]}...", end=" ", flush=True)

            result = api_chat(args.server, args.token, args.agent, q["query"], model=model, timeout=args.timeout)

            analysis = analyse_response(q, result["response"])

            all_results.append({
                "id": q["id"],
                "category": q["category"],
                "query": q["query"],
                "model": model,
                "response_text": result["response"],
                "latency": result["latency"],
                "error": result["error"],
                **analysis,
            })

            status = "OK" if analysis["correct_behavior"] else "FAIL"
            error_info = f" | ERROR: {result['error']}" if result["error"] else ""
            print(f"{status} ({result['latency']}s){error_info}")

            # Small delay to avoid overwhelming the server
            time.sleep(1)

        print()

    generate_report(all_results, output, models)

    # Print summary
    print("\n=== SUMMARY ===")
    for model in models:
        model_results = [r for r in all_results if r["model"] == model]
        total = len(model_results)
        correct = sum(1 for r in model_results if r["correct_behavior"])
        print(f"  {model}: {correct}/{total} correct ({100*correct//max(total,1)}%)")


if __name__ == "__main__":
    main()
