"""
Vanilla RAG Agent — NotebookLM-equivalent baseline for comparison benchmark.

A minimal Retrieval-Augmented Generation agent that uses the SAME knowledge
base as Tommi3/Responsible AI3 but with a standard RAG architecture:

  1. Chunk all documents (papers, researchers, projects, glossary)
  2. Embed chunks via Mistral embeddings API
  3. On query: embed query → cosine similarity → top-K chunks → LLM generates

NO classification, NO routing, NO post-processing, NO reliability cues.
This is intentionally simple — it represents what a generic RAG tool
(like NotebookLM) would do with the same data.

Usage:
    # First time: build the index
    python vanilla_rag.py --build-index

    # Query interactively
    python vanilla_rag.py --query "What is responsible AI?"

    # Run the comparison benchmark queries
    python vanilla_rag.py --run-benchmark

    # Run benchmark and fill the Excel
    python vanilla_rag.py --run-benchmark --fill-excel comparison_tommi3_vs_notebooklm_20260604.xlsx
"""

import argparse
import json
import math
import os
import pickle
import sys
import time
from pathlib import Path

# Load .env
_env_path = Path(__file__).resolve().parent.parent.parent.parent / "web" / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())

MISTRAL_API_KEY = os.environ.get("MISTRAL_API_KEY", "")
MISTRAL_MODEL = os.environ.get("MISTRAL_MODEL", "mistral-small-latest")
EMBEDDING_MODEL = "mistral-embed"

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
INDEX_PATH = Path(__file__).resolve().parent / "vanilla_rag_index.pkl"

# ============================================================
# Mistral API helpers (using urllib — no extra deps)
# ============================================================

from urllib.request import Request, urlopen
import urllib.error


def mistral_embeddings(texts: list[str], batch_size: int = 25) -> list[list[float]]:
    """Get embeddings from Mistral API in batches."""
    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        payload = json.dumps({
            "model": EMBEDDING_MODEL,
            "input": batch,
        }).encode()
        req = Request(
            "https://api.mistral.ai/v1/embeddings",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {MISTRAL_API_KEY}",
            },
        )
        with urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read())
        for item in sorted(result["data"], key=lambda x: x["index"]):
            all_embeddings.append(item["embedding"])
        if i + batch_size < len(texts):
            time.sleep(0.3)  # Rate limiting
    return all_embeddings


def mistral_chat(messages: list[dict], temperature: float = 0.3,
                 max_tokens: int = 2000) -> str:
    """Call Mistral chat API."""
    payload = json.dumps({
        "model": MISTRAL_MODEL,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }).encode()
    req = Request(
        "https://api.mistral.ai/v1/chat/completions",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {MISTRAL_API_KEY}",
        },
    )
    with urlopen(req, timeout=120) as resp:
        result = json.loads(resp.read())
    return result["choices"][0]["message"]["content"]


def cosine_similarity(a: list[float], b: list[float]) -> float:
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(x * x for x in b))
    if na == 0 or nb == 0:
        return 0.0
    return dot / (na * nb)


# ============================================================
# Document chunking
# ============================================================

def chunk_text(text: str, source: str, chunk_size: int = 800,
               overlap: int = 100) -> list[dict]:
    """Split text into overlapping chunks."""
    chunks = []
    words = text.split()
    i = 0
    while i < len(words):
        chunk_words = words[i:i + chunk_size]
        chunk_text = " ".join(chunk_words)
        if len(chunk_text.strip()) > 50:  # Skip tiny chunks
            chunks.append({
                "text": chunk_text,
                "source": source,
            })
        i += chunk_size - overlap
    return chunks


def load_and_chunk_all() -> list[dict]:
    """Load all knowledge base documents and chunk them."""
    chunks = []

    # 1. Glossary
    glossary_path = DATA_DIR / "docs" / "Glossary_Responsible_AI.md"
    if glossary_path.exists():
        text = glossary_path.read_text(encoding="utf-8")
        # Split by glossary entries (## headings)
        entries = text.split("\n## ")
        for entry in entries:
            entry = entry.strip()
            if len(entry) > 50:
                chunks.append({
                    "text": "## " + entry if not entry.startswith("#") else entry,
                    "source": "Glossary",
                })
        print(f"  Glossary: {len(entries)} entries")

    # 2. Papers (structured data → convert to readable text)
    papers_path = DATA_DIR / "papers.json"
    if papers_path.exists():
        with open(papers_path, encoding="utf-8") as f:
            papers_data = json.load(f)
        paper_count = 0
        for uni_code, uni_info in papers_data.get("universities", {}).items():
            uni_name = uni_info.get("name", uni_code)
            for paper in uni_info.get("papers", []):
                text_parts = [
                    f"Paper: {paper.get('title', 'Unknown')}",
                    f"University: {uni_name} ({uni_code})",
                    f"Year: {paper.get('year', 'Unknown')}",
                    f"DOI: {paper.get('doi', 'N/A')}",
                ]
                if paper.get("abstract"):
                    text_parts.append(f"Abstract: {paper['abstract']}")
                topics = paper.get("topics", [])
                if topics:
                    text_parts.append(f"Topics: {', '.join(topics)}")
                authors = paper.get("authors", [])
                if authors:
                    author_strs = []
                    for a in authors:
                        name = a.get("name", "")
                        inst = a.get("institution", "")
                        author_strs.append(f"{name} ({inst})" if inst else name)
                    text_parts.append(f"Authors: {', '.join(author_strs)}")
                cited = paper.get("cited_by_count", 0)
                if cited:
                    text_parts.append(f"Cited by: {cited}")

                paper_text = "\n".join(text_parts)
                # Most papers fit in one chunk; split only if very long
                for c in chunk_text(paper_text, f"Paper/{uni_code}/{paper.get('id', '')}"):
                    chunks.append(c)
                paper_count += 1
        print(f"  Papers: {paper_count} papers")

    # 3. Researchers (structured → text)
    researchers_path = DATA_DIR / "researchers.json"
    if researchers_path.exists():
        with open(researchers_path, encoding="utf-8") as f:
            researchers_data = json.load(f)
        researcher_count = 0
        for uni_code, researchers_list in researchers_data.items():
            for r in researchers_list:
                text_parts = [
                    f"Researcher: {r.get('name', 'Unknown')}",
                    f"University: {uni_code}",
                    f"Paper count: {r.get('paper_count', 0)}",
                ]
                topics = r.get("topics", [])
                if topics:
                    text_parts.append(f"Research topics: {', '.join(topics)}")
                papers = r.get("papers", [])
                if papers:
                    paper_strs = [f"- {p.get('title', '')} ({p.get('year', '')})"
                                  for p in papers[:10]]
                    text_parts.append("Publications:\n" + "\n".join(paper_strs))
                chunks.append({
                    "text": "\n".join(text_parts),
                    "source": f"Researcher/{uni_code}/{r.get('name', '')}",
                })
                researcher_count += 1
        print(f"  Researchers: {researcher_count}")

    # 4. Project docs
    project_dir = DATA_DIR / "project_docs"
    if project_dir.exists():
        project_count = 0
        for md_file in sorted(project_dir.glob("*.md")):
            text = md_file.read_text(encoding="utf-8")
            for c in chunk_text(text, f"Project/{md_file.stem}"):
                chunks.append(c)
            project_count += 1
        print(f"  Project docs: {project_count}")

    print(f"  TOTAL CHUNKS: {len(chunks)}")
    return chunks


# ============================================================
# Index: build and load
# ============================================================

def build_index():
    """Build the vector index from all documents."""
    print("Loading and chunking documents...")
    chunks = load_and_chunk_all()

    print(f"\nEmbedding {len(chunks)} chunks (this may take a minute)...")
    texts = [c["text"] for c in chunks]
    embeddings = mistral_embeddings(texts)

    index = {
        "chunks": chunks,
        "embeddings": embeddings,
    }
    with open(INDEX_PATH, "wb") as f:
        pickle.dump(index, f)
    print(f"Index saved to {INDEX_PATH} ({len(chunks)} chunks)")
    return index


def load_index():
    """Load the pre-built index."""
    if not INDEX_PATH.exists():
        print("Index not found. Building...")
        return build_index()
    with open(INDEX_PATH, "rb") as f:
        return pickle.load(f)


# ============================================================
# RAG query pipeline
# ============================================================

SYSTEM_PROMPT = """You are a research assistant with access to a knowledge base about Responsible AI research from the UNINOVIS European university alliance. The alliance includes 8 universities: USPN (France), UDCLV (Italy), UMA (Spain), KK (Lithuania), UT (Albania), THWS (Germany), TAMK (Finland), THUAS (Netherlands).

Your knowledge base contains:
- Research papers published by UNINOVIS partners
- Researcher profiles and their topics
- Research project descriptions
- A glossary of Responsible AI terms

Answer questions based on the provided context. If the context doesn't contain enough information to answer, say so honestly. Always cite your sources when possible.

If the question is clearly outside the scope of Responsible AI research and the UNINOVIS alliance (e.g., weather, sports, booking flights), politely explain that you can only help with Responsible AI topics."""

def retrieve(query: str, index: dict, top_k: int = 8) -> list[dict]:
    """Retrieve the top-K most similar chunks for a query."""
    query_emb = mistral_embeddings([query])[0]
    scored = []
    for i, emb in enumerate(index["embeddings"]):
        sim = cosine_similarity(query_emb, emb)
        scored.append((sim, i))
    scored.sort(reverse=True)
    results = []
    for sim, idx in scored[:top_k]:
        chunk = index["chunks"][idx].copy()
        chunk["similarity"] = round(sim, 4)
        results.append(chunk)
    return results


def query_rag(query: str, index: dict, top_k: int = 8,
              temperature: float = 0.3) -> dict:
    """Full RAG pipeline: retrieve → augment → generate."""
    start = time.time()

    # Retrieve
    retrieved = retrieve(query, index, top_k=top_k)

    # Build context
    context_parts = []
    for i, chunk in enumerate(retrieved):
        context_parts.append(
            f"[Source {i+1}: {chunk['source']} (relevance: {chunk['similarity']})]\n"
            f"{chunk['text']}"
        )
    context = "\n\n---\n\n".join(context_parts)

    # Generate
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"Context from knowledge base:\n\n{context}\n\n---\n\nQuestion: {query}"},
    ]
    response = mistral_chat(messages, temperature=temperature)
    elapsed = time.time() - start

    return {
        "response": response,
        "sources": [{"source": c["source"], "similarity": c["similarity"]}
                    for c in retrieved],
        "latency": round(elapsed, 2),
        "chunks_used": len(retrieved),
    }


# ============================================================
# Benchmark runner
# ============================================================

def run_benchmark(index: dict, fill_excel: str = None):
    """Run all comparison benchmark queries and optionally fill the Excel."""
    sys.path.insert(0, str(Path(__file__).parent))
    from comparison_benchmark import QUERIES

    results = []
    for i, q in enumerate(QUERIES):
        print(f"  [{i+1}/{len(QUERIES)}] {q['id']}: {q['query'][:60]}...", end="", flush=True)
        try:
            result = query_rag(q["query"], index)
            results.append({
                "id": q["id"],
                "query": q["query"],
                "response": result["response"],
                "sources": result["sources"],
                "latency": result["latency"],
                "error": None,
            })
            print(f" {result['latency']}s ({len(result['response'])} chars)")
        except Exception as e:
            results.append({
                "id": q["id"],
                "query": q["query"],
                "response": "",
                "sources": [],
                "latency": 0,
                "error": str(e),
            })
            print(f" ERROR: {e}")
        time.sleep(0.5)  # Rate limiting

    # Save JSON backup
    json_path = Path(__file__).parent / "comparison_responses_vanilla_rag.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nSaved {len(results)} responses to {json_path}")

    # Fill Excel if requested
    if fill_excel:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fill_excel)
            ws = wb["Responses"]
            by_id = {r["id"]: r for r in results}
            for row in range(2, ws.max_row + 1):
                qid = ws.cell(row=row, column=1).value
                if qid and qid in by_id:
                    r = by_id[qid]
                    ws.cell(row=row, column=6, value=r["response"][:32000])  # NLM Response
                    sources = ", ".join(s["source"] for s in r["sources"][:5])
                    ws.cell(row=row, column=7, value=f"Yes ({len(r['sources'])} chunks)")  # Citations
                    ws.cell(row=row, column=8, value=f"Latency: {r['latency']}s | Sources: {sources}"
                            + (f" | Error: {r['error']}" if r["error"] else ""))
            wb.save(fill_excel)
            print(f"Excel updated: {fill_excel}")
        except ImportError:
            print("openpyxl not installed — Excel not updated")

    return results


# ============================================================
# Interactive mode
# ============================================================

def interactive(index: dict):
    """Interactive query loop."""
    print("\nVanilla RAG Agent (type 'quit' to exit)")
    print("=" * 50)
    while True:
        try:
            query = input("\nQuery: ").strip()
        except (EOFError, KeyboardInterrupt):
            break
        if not query or query.lower() in ("quit", "exit", "q"):
            break
        result = query_rag(query, index)
        print(f"\n{result['response']}")
        print(f"\n--- Sources ({result['latency']}s) ---")
        for s in result["sources"][:5]:
            print(f"  {s['source']} (sim: {s['similarity']})")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Vanilla RAG Agent — NotebookLM baseline")
    parser.add_argument("--build-index", action="store_true", help="Build the vector index")
    parser.add_argument("--query", "-q", type=str, help="Single query")
    parser.add_argument("--run-benchmark", action="store_true", help="Run comparison benchmark")
    parser.add_argument("--fill-excel", type=str, help="Excel file to fill with responses")
    parser.add_argument("--interactive", "-i", action="store_true", help="Interactive mode")
    args = parser.parse_args()

    if not MISTRAL_API_KEY:
        print("ERROR: MISTRAL_API_KEY not set. Check web/.env")
        sys.exit(1)

    if args.build_index:
        build_index()
        return

    index = load_index()
    print(f"Index loaded: {len(index['chunks'])} chunks")

    if args.query:
        result = query_rag(args.query, index)
        print(f"\n{result['response']}")
        print(f"\n--- Sources ({result['latency']}s) ---")
        for s in result["sources"][:5]:
            print(f"  {s['source']} (sim: {s['similarity']})")
    elif args.run_benchmark:
        run_benchmark(index, fill_excel=args.fill_excel)
    elif args.interactive:
        interactive(index)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
