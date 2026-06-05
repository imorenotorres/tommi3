"""
Agentic RAG Agent — Tool-using agent for comparison benchmark.

A RAG agent that uses the SAME knowledge base as Tommi3/Responsible AI3
but lets the LLM decide which tools to call for each query, rather than
pre-engineering routing rules.

Architecture:
  1. The LLM receives the query + descriptions of available tools
  2. It decides which tool(s) to call (search_papers, lookup_researcher, etc.)
  3. Tool results are returned to the LLM
  4. The LLM generates a final answer grounded in the tool results
  5. Optionally: a verification step checks cited facts

Available tools:
  - search_papers(query, university, year) — search by title/abstract/topics
  - lookup_researcher(name) — find researcher by name
  - list_researchers(university, topic) — list researchers by affiliation/topic
  - lookup_glossary(term) — look up a Responsible AI term
  - search_projects(query) — search project descriptions
  - search_by_embedding(query) — fallback: standard RAG embedding search
  - verify_paper_title(title) — check if a paper exists in the database

The LLM orchestrates tool use — NO pre-engineered routing.
Same LLM (Mistral), same data, same temperature as the other two agents.

Usage:
    python agentic_rag.py --query "What is responsible AI?"
    python agentic_rag.py --run-benchmark
    python agentic_rag.py --run-benchmark --fill-excel comparison_tommi3_vs_notebooklm_20260604.xlsx
    python agentic_rag.py --interactive
"""

import argparse
import json
import math
import os
import pickle
import re
import sys
import time
from pathlib import Path
from urllib.request import Request, urlopen
import urllib.error

# Load .env
_env_path = Path(__file__).resolve().parent.parent.parent.parent / "web" / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())

MISTRAL_API_KEY = os.environ.get("MISTRAL_API_KEY", "")
MISTRAL_MODEL = os.environ.get("MISTRAL_MODEL", "mistral-small-latest")
EMBEDDING_MODEL = "mistral-embed"

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
INDEX_PATH = Path(__file__).resolve().parent / "vanilla_rag_index.pkl"

MAX_TOOL_ROUNDS = 3  # Maximum tool-calling rounds per query


# ============================================================
# Mistral API
# ============================================================

def mistral_api(messages, tools=None, temperature=0.3, max_tokens=2000):
    """Call Mistral chat API with optional tool definitions. Returns full response object."""
    payload = {
        "model": MISTRAL_MODEL,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    if tools:
        payload["tools"] = tools
        payload["tool_choice"] = "auto"

    req = Request(
        "https://api.mistral.ai/v1/chat/completions",
        data=json.dumps(payload).encode(),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {MISTRAL_API_KEY}",
        },
    )
    with urlopen(req, timeout=120) as resp:
        return json.loads(resp.read())


def mistral_embeddings(texts, batch_size=25):
    """Get embeddings from Mistral API."""
    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        payload = json.dumps({"model": EMBEDDING_MODEL, "input": batch}).encode()
        req = Request("https://api.mistral.ai/v1/embeddings", data=payload,
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {MISTRAL_API_KEY}"})
        with urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read())
        for item in sorted(result["data"], key=lambda x: x["index"]):
            all_embeddings.append(item["embedding"])
        if i + batch_size < len(texts):
            time.sleep(0.3)
    return all_embeddings


def cosine_similarity(a, b):
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(x * x for x in b))
    return dot / (na * nb) if na and nb else 0.0


# ============================================================
# Knowledge base loaders (loaded once)
# ============================================================

_papers_data = None
_researchers_data = None
_glossary_entries = None
_project_docs = None
_embedding_index = None


def get_papers():
    global _papers_data
    if _papers_data is None:
        with open(DATA_DIR / "papers.json", encoding="utf-8") as f:
            _papers_data = json.load(f)
    return _papers_data


def get_researchers():
    global _researchers_data
    if _researchers_data is None:
        with open(DATA_DIR / "researchers.json", encoding="utf-8") as f:
            _researchers_data = json.load(f)
    return _researchers_data


def get_glossary():
    global _glossary_entries
    if _glossary_entries is None:
        text = (DATA_DIR / "docs" / "Glossary_Responsible_AI.md").read_text(encoding="utf-8")
        _glossary_entries = {}
        for entry in text.split("\n## "):
            entry = entry.strip()
            if not entry:
                continue
            lines = entry.split("\n", 1)
            term = lines[0].strip().lstrip("#").strip()
            body = lines[1].strip() if len(lines) > 1 else ""
            if term:
                _glossary_entries[term.lower()] = {"term": term, "definition": body}
    return _glossary_entries


def get_projects():
    global _project_docs
    if _project_docs is None:
        _project_docs = {}
        pdir = DATA_DIR / "project_docs"
        for md in sorted(pdir.glob("*.md")):
            text = md.read_text(encoding="utf-8")
            name = md.stem
            # Try to extract project name from filename or first heading
            match = re.search(r'_([A-Za-z][\w-]+)$', name)
            proj_name = match.group(1) if match else name
            _project_docs[proj_name.lower()] = {"name": proj_name, "filename": md.name, "content": text}
    return _project_docs


def get_embedding_index():
    global _embedding_index
    if _embedding_index is None:
        if not INDEX_PATH.exists():
            print("WARNING: Embedding index not found. Run vanilla_rag.py --build-index first.")
            _embedding_index = {"chunks": [], "embeddings": []}
        else:
            with open(INDEX_PATH, "rb") as f:
                _embedding_index = pickle.load(f)
    return _embedding_index


# ============================================================
# Tool implementations
# ============================================================

def tool_search_papers(query="", university="", year=""):
    """Search papers by title, abstract, concepts. Optional filters by university and year."""
    papers = get_papers()
    query_lower = query.lower() if query else ""
    results = []

    for uni_code, uni_info in papers.get("universities", {}).items():
        if university and university.upper() != uni_code:
            continue
        for p in uni_info.get("papers", []):
            # Match against title, abstract, concepts
            title = (p.get("title") or "").lower()
            abstract = (p.get("abstract") or "").lower()
            concepts = " ".join(c.get("display_name", "") for c in p.get("concepts", [])).lower()
            searchable = f"{title} {abstract} {concepts}"

            if query_lower and query_lower not in searchable:
                # Try individual words
                words = query_lower.split()
                if not any(w in searchable for w in words if len(w) > 3):
                    continue

            if year and str(p.get("publication_year", "")) != str(year):
                continue

            authors = [a.get("name", "") for a in p.get("authors", [])]
            results.append({
                "title": p.get("title", ""),
                "university": uni_code,
                "year": p.get("publication_year", ""),
                "authors": authors[:5],
                "cited_by": p.get("cited_by_count", 0),
                "doi": p.get("doi", ""),
                "abstract": (p.get("abstract") or "")[:300],
            })

    # Sort by relevance (cited_by as proxy) and limit
    results.sort(key=lambda x: x.get("cited_by", 0), reverse=True)
    if len(results) > 15:
        results = results[:15]

    return json.dumps({
        "count": len(results),
        "papers": results,
        "note": f"Found {len(results)} papers" + (f" matching '{query}'" if query else "")
            + (f" from {university}" if university else "")
            + (f" in {year}" if year else ""),
    }, ensure_ascii=False)


def tool_lookup_researcher(name=""):
    """Look up a specific researcher by name."""
    researchers = get_researchers()
    name_lower = name.lower()
    results = []

    for uni_code, r_list in researchers.items():
        for r in r_list:
            r_name = r.get("name", "").lower()
            if name_lower in r_name or r_name in name_lower:
                papers = [{"title": p.get("title", ""), "year": p.get("year", "")}
                          for p in r.get("papers", [])[:10]]
                results.append({
                    "name": r.get("name", ""),
                    "university": uni_code,
                    "paper_count": r.get("paper_count", 0),
                    "topics": r.get("topics", []),
                    "papers": papers,
                })

    return json.dumps({
        "count": len(results),
        "researchers": results,
        "note": f"Found {len(results)} researcher(s) matching '{name}'",
    }, ensure_ascii=False)


def tool_list_researchers(university="", topic=""):
    """List researchers, optionally filtered by university or research topic."""
    researchers = get_researchers()
    results = []

    for uni_code, r_list in researchers.items():
        if university and university.upper() != uni_code:
            continue
        for r in r_list:
            if topic:
                topics_lower = [t.lower() for t in r.get("topics", [])]
                if not any(topic.lower() in t for t in topics_lower):
                    continue
            results.append({
                "name": r.get("name", ""),
                "university": uni_code,
                "paper_count": r.get("paper_count", 0),
                "topics": r.get("topics", [])[:8],
            })

    return json.dumps({
        "count": len(results),
        "researchers": results,
        "note": f"Found {len(results)} researchers"
            + (f" from {university}" if university else "")
            + (f" on topic '{topic}'" if topic else ""),
    }, ensure_ascii=False)


def tool_lookup_glossary(term=""):
    """Look up a term in the Responsible AI glossary."""
    glossary = get_glossary()
    term_lower = term.lower().strip()

    # Exact match
    if term_lower in glossary:
        entry = glossary[term_lower]
        return json.dumps({"found": True, "term": entry["term"],
                           "definition": entry["definition"]}, ensure_ascii=False)

    # Partial match
    matches = []
    for key, entry in glossary.items():
        if term_lower in key or key in term_lower:
            matches.append({"term": entry["term"], "definition": entry["definition"][:300]})

    if matches:
        return json.dumps({"found": True, "partial_match": True,
                           "matches": matches[:3]}, ensure_ascii=False)

    # List available terms
    all_terms = [e["term"] for e in glossary.values()]
    return json.dumps({"found": False, "note": f"'{term}' not found in glossary.",
                       "available_terms": all_terms}, ensure_ascii=False)


def tool_search_projects(query=""):
    """Search research project descriptions."""
    projects = get_projects()
    query_lower = query.lower()
    results = []

    for key, proj in projects.items():
        if query_lower in proj["content"].lower() or query_lower in key:
            # Extract first paragraph as summary
            lines = proj["content"].strip().split("\n")
            summary = ""
            for line in lines:
                line = line.strip()
                if line and not line.startswith("#"):
                    summary = line[:500]
                    break
            results.append({
                "project": proj["name"],
                "summary": summary,
                "content_preview": proj["content"][:800],
            })

    return json.dumps({
        "count": len(results),
        "projects": results,
        "note": f"Found {len(results)} projects matching '{query}'",
    }, ensure_ascii=False)


def tool_search_by_embedding(query=""):
    """Semantic search across all documents using embedding similarity."""
    index = get_embedding_index()
    if not index["chunks"]:
        return json.dumps({"error": "Embedding index not available"})

    query_emb = mistral_embeddings([query])[0]
    scored = []
    for i, emb in enumerate(index["embeddings"]):
        sim = cosine_similarity(query_emb, emb)
        scored.append((sim, i))
    scored.sort(reverse=True)

    results = []
    for sim, idx in scored[:6]:
        chunk = index["chunks"][idx]
        results.append({
            "source": chunk["source"],
            "text": chunk["text"][:500],
            "similarity": round(sim, 4),
        })

    return json.dumps({
        "count": len(results),
        "chunks": results,
        "note": f"Top {len(results)} chunks by semantic similarity to '{query}'",
    }, ensure_ascii=False)


def tool_verify_paper_title(title=""):
    """Check if a paper title exists in the database. Use after generating an answer to verify cited papers."""
    papers = get_papers()
    title_lower = title.lower().strip()

    for uni_code, uni_info in papers.get("universities", {}).items():
        for p in uni_info.get("papers", []):
            if p.get("title", "").lower().strip() == title_lower:
                return json.dumps({
                    "exists": True,
                    "title": p["title"],
                    "university": uni_code,
                    "year": p.get("publication_year", ""),
                    "doi": p.get("doi", ""),
                })

    # Partial match
    for uni_code, uni_info in papers.get("universities", {}).items():
        for p in uni_info.get("papers", []):
            if title_lower in p.get("title", "").lower():
                return json.dumps({
                    "exists": True,
                    "partial_match": True,
                    "title": p["title"],
                    "university": uni_code,
                })

    return json.dumps({"exists": False, "note": f"Paper '{title}' not found in database"})


# ============================================================
# Tool registry
# ============================================================

TOOL_FUNCTIONS = {
    "search_papers": tool_search_papers,
    "lookup_researcher": tool_lookup_researcher,
    "list_researchers": tool_list_researchers,
    "lookup_glossary": tool_lookup_glossary,
    "search_projects": tool_search_projects,
    "search_by_embedding": tool_search_by_embedding,
    "verify_paper_title": tool_verify_paper_title,
}

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "search_papers",
            "description": "Search research papers in the UNINOVIS database by topic, title keywords, or abstract content. Can filter by university and year. Use this for questions about publications, research topics, or counting papers.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Search terms (topic, keywords from title/abstract)"},
                    "university": {"type": "string", "description": "Filter by university code: KK, TAMK, THUAS, THWS, UDCLV, UMA, USPN, UT"},
                    "year": {"type": "string", "description": "Filter by publication year (e.g. '2024')"},
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "lookup_researcher",
            "description": "Look up a specific researcher by name. Returns their university, topics, and publications. Use when the user asks about a specific person.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "Researcher name (or partial name)"},
                },
                "required": ["name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_researchers",
            "description": "List researchers, optionally filtered by university or research topic. Use for questions like 'who works on X' or 'researchers from Y'.",
            "parameters": {
                "type": "object",
                "properties": {
                    "university": {"type": "string", "description": "Filter by university code: KK, TAMK, THUAS, THWS, UDCLV, UMA, USPN, UT"},
                    "topic": {"type": "string", "description": "Filter by research topic keyword"},
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "lookup_glossary",
            "description": "Look up a term in the Responsible AI glossary. Use for concept definition questions like 'What is X?' where X is an AI/ethics concept.",
            "parameters": {
                "type": "object",
                "properties": {
                    "term": {"type": "string", "description": "The term to look up (e.g. 'explainable AI', 'fairness', 'EU AI Act')"},
                },
                "required": ["term"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_projects",
            "description": "Search EU-funded research project descriptions. Use for questions about specific projects (TAILOR, DUCA, MENHIR, etc.) or project-related queries.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Project name or topic to search for"},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_by_embedding",
            "description": "Semantic search across ALL documents using embedding similarity. Use as a fallback when other tools don't find relevant results, or for broad/vague queries.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "The search query"},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "verify_paper_title",
            "description": "Verify that a specific paper title exists in the database. Use AFTER generating an answer to check that any cited paper titles are real and not hallucinated.",
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {"type": "string", "description": "The exact paper title to verify"},
                },
                "required": ["title"],
            },
        },
    },
]


# ============================================================
# System prompt
# ============================================================

SYSTEM_PROMPT = """You are a research assistant for the UNINOVIS European university alliance, specialising in Responsible AI research. The alliance includes 8 universities: USPN (France), UDCLV (Italy), UMA (Spain), KK (Lithuania), UT (Albania), THWS (Germany), TAMK (Finland), THUAS (Netherlands).

You have access to tools that query a knowledge base containing:
- Research papers published by UNINOVIS partners
- Researcher profiles with their topics and publications
- EU-funded research project descriptions
- A glossary of Responsible AI terms

INSTRUCTIONS:
1. For each user query, decide which tool(s) to call to find relevant information.
2. Use the tool results to construct your answer. Only state facts that are supported by tool results.
3. If the query is clearly outside the scope of Responsible AI and UNINOVIS (weather, sports, booking flights, writing essays, translations), politely refuse without calling any tools.
4. If a tool returns no results, say so honestly rather than inventing information.
5. When citing specific paper titles, use verify_paper_title to confirm they exist.
6. For concept/definition questions, try lookup_glossary first. If not found, use search_by_embedding.
7. You may call multiple tools if needed to fully answer a question."""


# ============================================================
# Agentic query loop
# ============================================================

def query_agentic(query, verbose=False):
    """Run the agentic RAG loop: LLM decides tools → execute → LLM answers."""
    start = time.time()
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": query},
    ]

    tool_calls_log = []

    for round_num in range(MAX_TOOL_ROUNDS):
        result = mistral_api(messages, tools=TOOL_DEFINITIONS)
        choice = result["choices"][0]
        assistant_msg = choice["message"]

        # If no tool calls, we have the final answer
        tool_calls = assistant_msg.get("tool_calls")
        if not tool_calls:
            elapsed = time.time() - start
            return {
                "response": assistant_msg.get("content", ""),
                "tool_calls": tool_calls_log,
                "rounds": round_num + 1,
                "latency": round(elapsed, 2),
                "error": None,
            }

        # Execute tool calls
        messages.append(assistant_msg)

        for tc in tool_calls:
            func_name = tc["function"]["name"]
            try:
                args = json.loads(tc["function"]["arguments"])
            except json.JSONDecodeError:
                args = {}

            tool_calls_log.append({"tool": func_name, "args": args, "round": round_num + 1})

            if verbose:
                print(f"    -> {func_name}({args})")

            func = TOOL_FUNCTIONS.get(func_name)
            if func:
                try:
                    tool_result = func(**args)
                except Exception as e:
                    tool_result = json.dumps({"error": str(e)})
            else:
                tool_result = json.dumps({"error": f"Unknown tool: {func_name}"})

            messages.append({
                "role": "tool",
                "name": func_name,
                "content": tool_result,
                "tool_call_id": tc.get("id", ""),
            })

    # If we exhausted rounds, do one final call without tools
    result = mistral_api(messages, tools=None)
    elapsed = time.time() - start
    return {
        "response": result["choices"][0]["message"].get("content", ""),
        "tool_calls": tool_calls_log,
        "rounds": MAX_TOOL_ROUNDS,
        "latency": round(elapsed, 2),
        "error": None,
    }


# ============================================================
# Benchmark runner
# ============================================================

def run_benchmark(fill_excel=None):
    """Run all comparison benchmark queries."""
    sys.path.insert(0, str(Path(__file__).parent))
    from comparison_benchmark import QUERIES

    results = []
    for i, q in enumerate(QUERIES):
        print(f"  [{i+1}/{len(QUERIES)}] {q['id']}: {q['query'][:60]}...", end="", flush=True)
        try:
            result = query_agentic(q["query"])
            tools_used = list(set(tc["tool"] for tc in result["tool_calls"]))
            results.append({
                "id": q["id"],
                "query": q["query"],
                "response": result["response"],
                "tool_calls": result["tool_calls"],
                "tools_used": tools_used,
                "rounds": result["rounds"],
                "latency": result["latency"],
                "error": None,
            })
            print(f" {result['latency']}s | tools: {tools_used} | rounds: {result['rounds']} ({len(result['response'])} chars)")
        except Exception as e:
            results.append({
                "id": q["id"],
                "query": q["query"],
                "response": "",
                "tool_calls": [],
                "tools_used": [],
                "rounds": 0,
                "latency": 0,
                "error": str(e),
            })
            print(f" ERROR: {e}")
        time.sleep(0.5)

    # Save JSON
    json_path = Path(__file__).parent / "comparison_responses_agentic_rag.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nSaved {len(results)} responses to {json_path}")

    # Fill Excel if requested
    if fill_excel:
        fill_excel_file(fill_excel, results)

    return results


def fill_excel_file(excel_path, results):
    """Add agentic RAG results to the Excel workbook as a new sheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — Excel not updated")
        return

    wb = openpyxl.load_workbook(excel_path)

    # Create or replace AgenticRAG sheet
    sheet_name = "AgenticRAG Responses"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "7B1FA2"

    headers = ["ID", "Query", "Agentic RAG Response", "Tools Used", "Rounds", "Latency", "Tool Call Details"]
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    by_id = {r["id"]: r for r in results}
    row = 2
    # Use query order from comparison_benchmark
    sys.path.insert(0, str(Path(__file__).parent))
    from comparison_benchmark import QUERIES
    for q in QUERIES:
        qid = q["id"]
        if qid not in by_id:
            continue
        r = by_id[qid]
        ws.cell(row=row, column=1, value=qid)
        ws.cell(row=row, column=2, value=q["query"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=(r["response"] or "")[:32000]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=4, value=", ".join(r.get("tools_used", [])))
        ws.cell(row=row, column=5, value=r.get("rounds", 0))
        ws.cell(row=row, column=6, value=f"{r.get('latency', 0)}s")
        details = "; ".join(f"{tc['tool']}({tc['args']})" for tc in r.get("tool_calls", []))
        ws.cell(row=row, column=7, value=details[:32000]).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 60

    # Also fill NLM columns in Responses sheet with agentic data for comparison
    # (We reuse the NLM columns since the vanilla RAG is already saved separately)

    wb.save(excel_path)
    print(f"Excel updated with AgenticRAG sheet: {excel_path}")


# ============================================================
# Interactive mode
# ============================================================

def interactive():
    """Interactive query loop."""
    print("\nAgentic RAG Agent (type 'quit' to exit)")
    print("=" * 50)
    while True:
        try:
            query = input("\nQuery: ").strip()
        except (EOFError, KeyboardInterrupt):
            break
        if not query or query.lower() in ("quit", "exit", "q"):
            break
        result = query_agentic(query, verbose=True)
        print(f"\n{result['response']}")
        print(f"\n--- Tools: {[tc['tool'] for tc in result['tool_calls']]} | Rounds: {result['rounds']} | {result['latency']}s ---")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Agentic RAG Agent — tool-using baseline")
    parser.add_argument("--query", "-q", type=str, help="Single query")
    parser.add_argument("--run-benchmark", action="store_true", help="Run comparison benchmark")
    parser.add_argument("--fill-excel", type=str, help="Excel file to update")
    parser.add_argument("--interactive", "-i", action="store_true", help="Interactive mode")
    args = parser.parse_args()

    if not MISTRAL_API_KEY:
        print("ERROR: MISTRAL_API_KEY not set. Check web/.env")
        sys.exit(1)

    if args.query:
        result = query_agentic(args.query, verbose=True)
        print(f"\n{result['response']}")
        print(f"\n--- Tools: {[tc['tool'] for tc in result['tool_calls']]} | Rounds: {result['rounds']} | {result['latency']}s ---")
    elif args.run_benchmark:
        run_benchmark(fill_excel=args.fill_excel)
    elif args.interactive:
        interactive()
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
