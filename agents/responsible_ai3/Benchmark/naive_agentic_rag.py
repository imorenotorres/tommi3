"""
Naive Agentic RAG (Agent B') — Agentic agent with generic tools only.

Unlike the domain-informed agentic RAG (Agent B), this agent has NO
domain-specific tool design. Its tools are what anyone could build in
an hour without prior knowledge of the query taxonomy:

  1. search(query) — search everything (papers, researchers, glossary, projects)
  2. verify_fact(claim) — check if a specific claim is supported by the data
  3. count(query) — count matching items

NO separate tools for papers vs researchers vs glossary.
NO domain-informed parameter design.
Same LLM, same data, same temperature.

This tests: how much value does domain knowledge add to tool design?

Usage:
    python naive_agentic_rag.py --query "What is responsible AI?"
    python naive_agentic_rag.py --run-benchmark
    python naive_agentic_rag.py --run-benchmark --fill-excel comparison_tommi3_vs_notebooklm_20260604.xlsx
"""

import argparse
import json
import math
import os
import pickle
import re
import sys
import time
from pathlib import Path
from urllib.request import Request, urlopen

# Load .env
_env_path = Path(__file__).resolve().parent.parent.parent.parent / "web" / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())

MISTRAL_API_KEY = os.environ.get("MISTRAL_API_KEY", "")
MISTRAL_MODEL = os.environ.get("MISTRAL_MODEL", "mistral-small-latest")
EMBEDDING_MODEL = "mistral-embed"

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
INDEX_PATH = Path(__file__).resolve().parent / "vanilla_rag_index.pkl"

MAX_TOOL_ROUNDS = 3


# ============================================================
# Mistral API
# ============================================================

def mistral_api(messages, tools=None, temperature=0.3, max_tokens=2000):
    payload = {
        "model": MISTRAL_MODEL,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    if tools:
        payload["tools"] = tools
        payload["tool_choice"] = "auto"

    req = Request(
        "https://api.mistral.ai/v1/chat/completions",
        data=json.dumps(payload).encode(),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {MISTRAL_API_KEY}",
        },
    )
    with urlopen(req, timeout=120) as resp:
        return json.loads(resp.read())


def mistral_embeddings(texts, batch_size=25):
    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        payload = json.dumps({"model": EMBEDDING_MODEL, "input": batch}).encode()
        req = Request("https://api.mistral.ai/v1/embeddings", data=payload,
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {MISTRAL_API_KEY}"})
        with urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read())
        for item in sorted(result["data"], key=lambda x: x["index"]):
            all_embeddings.append(item["embedding"])
        if i + batch_size < len(texts):
            time.sleep(0.3)
    return all_embeddings


def cosine_similarity(a, b):
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(x * x for x in b))
    return dot / (na * nb) if na and nb else 0.0


# ============================================================
# Knowledge base — loaded as one flat collection
# ============================================================

_all_items = None
_embedding_index = None


def get_all_items():
    """Load ALL data into a single flat list of searchable items.
    This is what someone would build without domain analysis."""
    global _all_items
    if _all_items is not None:
        return _all_items

    _all_items = []

    # Papers
    with open(DATA_DIR / "papers.json", encoding="utf-8") as f:
        papers_data = json.load(f)
    for uni_code, uni_info in papers_data.get("universities", {}).items():
        for p in uni_info.get("papers", []):
            authors = [a.get("name", "") for a in p.get("authors", [])]
            _all_items.append({
                "type": "paper",
                "title": p.get("title", ""),
                "text": f"{p.get('title', '')} {p.get('abstract', '')}",
                "university": uni_code,
                "year": str(p.get("publication_year", "")),
                "authors": authors,
                "cited_by": p.get("cited_by_count", 0),
                "doi": p.get("doi", ""),
            })

    # Researchers
    with open(DATA_DIR / "researchers.json", encoding="utf-8") as f:
        researchers_data = json.load(f)
    for uni_code, r_list in researchers_data.items():
        for r in r_list:
            paper_titles = [p.get("title", "") for p in r.get("papers", [])]
            _all_items.append({
                "type": "researcher",
                "title": r.get("name", ""),
                "text": f"{r.get('name', '')} {uni_code} {' '.join(r.get('topics', []))} {' '.join(paper_titles)}",
                "university": uni_code,
                "topics": r.get("topics", []),
                "paper_count": r.get("paper_count", 0),
                "papers": [{"title": p.get("title", ""), "year": p.get("year", "")}
                           for p in r.get("papers", [])[:10]],
            })

    # Glossary
    glossary_text = (DATA_DIR / "docs" / "Glossary_Responsible_AI.md").read_text(encoding="utf-8")
    for entry in glossary_text.split("\n## "):
        entry = entry.strip()
        if not entry:
            continue
        lines = entry.split("\n", 1)
        term = lines[0].strip().lstrip("#").strip()
        body = lines[1].strip() if len(lines) > 1 else ""
        if term:
            _all_items.append({
                "type": "glossary",
                "title": term,
                "text": f"{term} {body}",
                "definition": body,
            })

    # Project docs
    pdir = DATA_DIR / "project_docs"
    for md in sorted(pdir.glob("*.md")):
        content = md.read_text(encoding="utf-8")
        match = re.search(r'_([A-Za-z][\w-]+)$', md.stem)
        proj_name = match.group(1) if match else md.stem
        _all_items.append({
            "type": "project",
            "title": proj_name,
            "text": f"{proj_name} {content}",
            "content": content[:2000],
        })

    return _all_items


def get_embedding_index():
    global _embedding_index
    if _embedding_index is None:
        if not INDEX_PATH.exists():
            _embedding_index = {"chunks": [], "embeddings": []}
        else:
            with open(INDEX_PATH, "rb") as f:
                _embedding_index = pickle.load(f)
    return _embedding_index


# ============================================================
# Generic tools — no domain knowledge
# ============================================================

def tool_search(query=""):
    """Search across all data: papers, researchers, glossary, projects.
    Returns the most relevant items matching the query."""
    items = get_all_items()
    query_lower = query.lower()
    query_words = [w for w in query_lower.split() if len(w) > 2]

    # Score each item by keyword overlap
    scored = []
    for item in items:
        text_lower = item["text"].lower()
        # Full phrase match gets highest score
        if query_lower in text_lower:
            score = 10
        else:
            # Word-level matching
            score = sum(1 for w in query_words if w in text_lower)
        if score > 0:
            scored.append((score, item))

    scored.sort(key=lambda x: x[0], reverse=True)

    # If keyword search found nothing, fall back to embedding search
    if not scored:
        index = get_embedding_index()
        if index["chunks"]:
            query_emb = mistral_embeddings([query])[0]
            emb_scored = []
            for i, emb in enumerate(index["embeddings"]):
                sim = cosine_similarity(query_emb, emb)
                emb_scored.append((sim, i))
            emb_scored.sort(reverse=True)
            results = []
            for sim, idx in emb_scored[:8]:
                chunk = index["chunks"][idx]
                results.append({
                    "source": chunk["source"],
                    "text": chunk["text"][:400],
                    "relevance": round(sim, 4),
                })
            return json.dumps({
                "count": len(results),
                "method": "embedding_fallback",
                "results": results,
            }, ensure_ascii=False)

    # Format top results
    results = []
    for score, item in scored[:15]:
        result = {"type": item["type"], "title": item["title"], "relevance": score}
        if item["type"] == "paper":
            result["university"] = item.get("university", "")
            result["year"] = item.get("year", "")
            result["authors"] = item.get("authors", [])[:5]
            result["cited_by"] = item.get("cited_by", 0)
            result["doi"] = item.get("doi", "")
            result["abstract"] = item.get("text", "")[:300]
        elif item["type"] == "researcher":
            result["university"] = item.get("university", "")
            result["topics"] = item.get("topics", [])[:8]
            result["paper_count"] = item.get("paper_count", 0)
            result["papers"] = item.get("papers", [])[:5]
        elif item["type"] == "glossary":
            result["definition"] = item.get("definition", "")[:500]
        elif item["type"] == "project":
            result["content"] = item.get("content", "")[:500]
        results.append(result)

    return json.dumps({
        "count": len(results),
        "method": "keyword",
        "results": results,
    }, ensure_ascii=False)


def tool_verify_fact(claim=""):
    """Check if a specific factual claim is supported by the data.
    Use for verifying paper titles, researcher names, affiliations, counts."""
    items = get_all_items()
    claim_lower = claim.lower().strip()

    # Check paper titles
    for item in items:
        if item["type"] == "paper":
            if claim_lower in item["title"].lower() or item["title"].lower() in claim_lower:
                return json.dumps({
                    "verified": True,
                    "match_type": "paper_title",
                    "title": item["title"],
                    "university": item.get("university", ""),
                    "year": item.get("year", ""),
                })

    # Check researcher names
    for item in items:
        if item["type"] == "researcher":
            if claim_lower in item["title"].lower() or item["title"].lower() in claim_lower:
                return json.dumps({
                    "verified": True,
                    "match_type": "researcher",
                    "name": item["title"],
                    "university": item.get("university", ""),
                })

    # Check glossary terms
    for item in items:
        if item["type"] == "glossary":
            if claim_lower in item["title"].lower() or item["title"].lower() in claim_lower:
                return json.dumps({
                    "verified": True,
                    "match_type": "glossary_term",
                    "term": item["title"],
                })

    return json.dumps({
        "verified": False,
        "note": f"Could not verify: '{claim}'",
    })


def tool_count(query=""):
    """Count items matching a query. Returns counts by type and by university."""
    items = get_all_items()
    query_lower = query.lower()
    query_words = [w for w in query_lower.split() if len(w) > 2]

    matching = []
    for item in items:
        text_lower = item["text"].lower()
        if query_lower in text_lower or any(w in text_lower for w in query_words):
            matching.append(item)

    # Count by type
    by_type = {}
    for item in matching:
        t = item["type"]
        by_type[t] = by_type.get(t, 0) + 1

    # Count by university (for papers and researchers)
    by_uni = {}
    for item in matching:
        uni = item.get("university", "")
        if uni:
            by_uni[uni] = by_uni.get(uni, 0) + 1

    # Total items by type (for context)
    total_by_type = {}
    for item in items:
        t = item["type"]
        total_by_type[t] = total_by_type.get(t, 0) + 1

    return json.dumps({
        "query": query,
        "matching_count": len(matching),
        "by_type": by_type,
        "by_university": by_uni,
        "totals": total_by_type,
    }, ensure_ascii=False)


# ============================================================
# Tool registry — intentionally generic
# ============================================================

TOOL_FUNCTIONS = {
    "search": tool_search,
    "verify_fact": tool_verify_fact,
    "count": tool_count,
}

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "search",
            "description": "Search across the entire knowledge base (papers, researchers, glossary terms, research projects). Returns the most relevant items matching the query. Use this for any question that requires looking up information.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "What to search for"},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "verify_fact",
            "description": "Check if a specific factual claim is supported by the data. Use to verify paper titles, researcher names, or other specific facts before including them in your answer.",
            "parameters": {
                "type": "object",
                "properties": {
                    "claim": {"type": "string", "description": "The specific claim to verify (e.g. a paper title, researcher name)"},
                },
                "required": ["claim"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "count",
            "description": "Count items in the knowledge base matching a query. Returns counts broken down by type (papers, researchers, etc.) and by university. Use for 'how many' questions.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "What to count (e.g. 'papers on AI ethics', 'researchers from UMA', 'papers 2025')"},
                },
                "required": ["query"],
            },
        },
    },
]


# ============================================================
# System prompt — intentionally generic
# ============================================================

SYSTEM_PROMPT = """You are a research assistant with access to a knowledge base about Responsible AI research from the UNINOVIS European university alliance. The alliance includes 8 universities: USPN (France), UDCLV (Italy), UMA (Spain), KK (Lithuania), UT (Albania), THWS (Germany), TAMK (Finland), THUAS (Netherlands).

You have tools to search, count, and verify facts in the knowledge base. Use them to answer questions accurately.

INSTRUCTIONS:
1. Use the search tool to find relevant information before answering.
2. Use the count tool for "how many" questions.
3. Use verify_fact to check specific claims (paper titles, names) before stating them.
4. Only state facts supported by tool results.
5. If the question is clearly outside the scope of Responsible AI and UNINOVIS (weather, sports, flights, essay writing, translations), politely refuse without calling any tools.
6. If tools return no results, say so honestly."""


# ============================================================
# Agentic query loop
# ============================================================

def query_agentic(query, verbose=False):
    start = time.time()
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": query},
    ]

    tool_calls_log = []

    for round_num in range(MAX_TOOL_ROUNDS):
        result = mistral_api(messages, tools=TOOL_DEFINITIONS)
        choice = result["choices"][0]
        assistant_msg = choice["message"]

        tool_calls = assistant_msg.get("tool_calls")
        if not tool_calls:
            elapsed = time.time() - start
            return {
                "response": assistant_msg.get("content", ""),
                "tool_calls": tool_calls_log,
                "rounds": round_num + 1,
                "latency": round(elapsed, 2),
                "error": None,
            }

        messages.append(assistant_msg)

        for tc in tool_calls:
            func_name = tc["function"]["name"]
            try:
                args = json.loads(tc["function"]["arguments"])
            except json.JSONDecodeError:
                args = {}

            tool_calls_log.append({"tool": func_name, "args": args, "round": round_num + 1})

            if verbose:
                print(f"    -> {func_name}({args})")

            func = TOOL_FUNCTIONS.get(func_name)
            if func:
                try:
                    tool_result = func(**args)
                except Exception as e:
                    tool_result = json.dumps({"error": str(e)})
            else:
                tool_result = json.dumps({"error": f"Unknown tool: {func_name}"})

            messages.append({
                "role": "tool",
                "name": func_name,
                "content": tool_result,
                "tool_call_id": tc.get("id", ""),
            })

    # Final call without tools
    result = mistral_api(messages, tools=None)
    elapsed = time.time() - start
    return {
        "response": result["choices"][0]["message"].get("content", ""),
        "tool_calls": tool_calls_log,
        "rounds": MAX_TOOL_ROUNDS,
        "latency": round(elapsed, 2),
        "error": None,
    }


# ============================================================
# Benchmark runner
# ============================================================

def run_benchmark(fill_excel=None):
    sys.path.insert(0, str(Path(__file__).parent))
    from comparison_benchmark import QUERIES

    results = []
    for i, q in enumerate(QUERIES):
        print(f"  [{i+1}/{len(QUERIES)}] {q['id']}: {q['query'][:60]}...", end="", flush=True)
        try:
            result = query_agentic(q["query"])
            tools_used = list(set(tc["tool"] for tc in result["tool_calls"]))
            results.append({
                "id": q["id"],
                "query": q["query"],
                "response": result["response"],
                "tool_calls": result["tool_calls"],
                "tools_used": tools_used,
                "rounds": result["rounds"],
                "latency": result["latency"],
                "error": None,
            })
            print(f" {result['latency']}s | tools: {tools_used} | rounds: {result['rounds']} ({len(result['response'])} chars)")
        except Exception as e:
            results.append({
                "id": q["id"],
                "query": q["query"],
                "response": "",
                "tool_calls": [],
                "tools_used": [],
                "rounds": 0,
                "latency": 0,
                "error": str(e),
            })
            print(f" ERROR: {e}")
        time.sleep(0.5)

    # Save JSON
    json_path = Path(__file__).parent / "comparison_responses_naive_agentic_rag.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nSaved {len(results)} responses to {json_path}")

    if fill_excel:
        fill_excel_file(fill_excel, results)

    return results


def fill_excel_file(excel_path, results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — Excel not updated")
        return

    wb = openpyxl.load_workbook(excel_path)

    sheet_name = "NaiveAgentic Responses"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "E65100"

    headers = ["ID", "Query", "Naive Agentic Response", "Tools Used", "Rounds", "Latency", "Tool Call Details"]
    header_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    sys.path.insert(0, str(Path(__file__).parent))
    from comparison_benchmark import QUERIES

    by_id = {r["id"]: r for r in results}
    row = 2
    for q in QUERIES:
        qid = q["id"]
        if qid not in by_id:
            continue
        r = by_id[qid]
        ws.cell(row=row, column=1, value=qid)
        ws.cell(row=row, column=2, value=q["query"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=3, value=(r["response"] or "")[:32000]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=4, value=", ".join(r.get("tools_used", [])))
        ws.cell(row=row, column=5, value=r.get("rounds", 0))
        ws.cell(row=row, column=6, value=f"{r.get('latency', 0)}s")
        details = "; ".join(f"{tc['tool']}({tc['args']})" for tc in r.get("tool_calls", []))
        ws.cell(row=row, column=7, value=details[:32000]).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 60

    wb.save(excel_path)
    print(f"Excel updated with NaiveAgentic sheet: {excel_path}")


def main():
    parser = argparse.ArgumentParser(description="Naive Agentic RAG — generic tools only")
    parser.add_argument("--query", "-q", type=str, help="Single query")
    parser.add_argument("--run-benchmark", action="store_true", help="Run comparison benchmark")
    parser.add_argument("--fill-excel", type=str, help="Excel file to update")
    parser.add_argument("--interactive", "-i", action="store_true", help="Interactive mode")
    args = parser.parse_args()

    if not MISTRAL_API_KEY:
        print("ERROR: MISTRAL_API_KEY not set")
        sys.exit(1)

    if args.query:
        result = query_agentic(args.query, verbose=True)
        print(f"\n{result['response']}")
        print(f"\n--- Tools: {[tc['tool'] for tc in result['tool_calls']]} | Rounds: {result['rounds']} | {result['latency']}s ---")
    elif args.run_benchmark:
        run_benchmark(fill_excel=args.fill_excel)
    elif args.interactive:
        print("\nNaive Agentic RAG (type 'quit' to exit)")
        print("=" * 50)
        while True:
            try:
                query = input("\nQuery: ").strip()
            except (EOFError, KeyboardInterrupt):
                break
            if not query or query.lower() in ("quit", "exit", "q"):
                break
            result = query_agentic(query, verbose=True)
            print(f"\n{result['response']}")
            print(f"\n--- Tools: {[tc['tool'] for tc in result['tool_calls']]} | Rounds: {result['rounds']} | {result['latency']}s ---")
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
