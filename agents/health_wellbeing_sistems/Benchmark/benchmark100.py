"""
Benchmark 100 queries for the Health and Wellbeing Systems Agent.

Runs 100 queries (organized by user profile) against the agent API via SSE
streaming, captures responses and procedural banners, evaluates correctness,
and generates an Excel report.

Authentication: The script auto-logs in using credentials from the .env file
(BENCHMARK_USER / BENCHMARK_PASSWORD) or from command-line arguments.
Default credentials: admin / admin.

Usage:
    python benchmark100.py --server http://localhost:8000
    python benchmark100.py --server http://localhost:8000 --models mistral-small-latest,mistral-large-latest
    python benchmark100.py --server http://localhost:8000 --categories Students,Professors
    python benchmark100.py --server http://localhost:8000 --ids S01,P05,A03
    python benchmark100.py --server http://localhost:8000 --user admin --password admin
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import urlencode
from urllib.error import HTTPError, URLError

# Load .env from web/ directory (where auth credentials live)
_env_path = Path(__file__).resolve().parent.parent.parent / "web" / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip())


def login(server: str, username: str, password: str) -> str:
    """Authenticate against the TOMMI server and return a session token."""
    url = f"{server}/api/auth/login"
    data = json.dumps({"username": username, "password": password}).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    try:
        req = Request(url, data=data, headers=headers, method="POST")
        with urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode("utf-8"))
            token = result.get("token")
            if not token:
                print(f"ERROR: Login succeeded but no token returned. Response: {result}")
                sys.exit(1)
            return token
    except HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        print(f"ERROR: Login failed (HTTP {e.code}): {body}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not connect to {server}: {e}")
        sys.exit(1)


# ============================================================
# Banner constants (procedural transparency)
# ============================================================

BANNER_GREEN = "green"    # Direct database/metadata output, no LLM interpretation (figures, paper lists, structured data)
BANNER_YELLOW = "yellow"  # LLM interprets database content (glossary answers, projects, conceptual questions, researcher lookups)
BANNER_RED = "red"        # Unverified / off-topic / speculation / no supporting data


# ============================================================
# 100 Benchmark queries with expected outcomes
# ============================================================

BENCHMARK_QUERIES = [
    # ----------------------------------------------------------
    # A. Students with minimal experience
    # ----------------------------------------------------------

    # Understanding the tool
    {"id": "S01", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "What can you do?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["research", "health"]},

    {"id": "S02", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "How does this work?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["question", "database"]},

    {"id": "S03", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "What kind of questions can I ask you?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["research", "paper"]},

    {"id": "S04", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "What is UNINOVIS?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["alliance", "universit"]},

    {"id": "S05", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "Which universities are in UNINOVIS?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["UMA", "THUAS", "UDCLV"]},

    {"id": "S06", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "Can you search the internet?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": []},

    {"id": "S07", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "Why did you say you can't answer my question?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["scope", "database"]},

    {"id": "S08", "profile": "Students", "subcategory": "Understanding the tool",
     "query": "What do the green and yellow banners mean?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["reliab", "verif"]},

    # Basic concepts
    {"id": "S09", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What is artificial intelligence?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["artificial intelligence"]},

    {"id": "S10", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What are wellbeing systems?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["wellbeing", "health", "well-being"]},

    {"id": "S11", "profile": "Students", "subcategory": "Basic concepts",
     "query": "Is AI dangerous in healthcare?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["risk", "harm", "healthcare"]},

    {"id": "S12", "profile": "Students", "subcategory": "Basic concepts",
     "query": "Can AI be trusted for medical decisions?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["trust", "medical", "clinical"]},

    {"id": "S13", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What is a language model?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["language model", "LLM"]},

    {"id": "S14", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What is digital health?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["digital health", "telemedicine", "mHealth"]},

    {"id": "S15", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What is bias in AI? Can you give me a healthcare example?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["bias", "discriminat", "health"]},

    {"id": "S16", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What does transparency mean in the context of health AI?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["transparency", "explain"]},

    {"id": "S17", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What is the EU AI Act?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["EU AI Act", "regulation"]},

    {"id": "S18", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What is explainable AI?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["explainable", "XAI"]},

    {"id": "S19", "profile": "Students", "subcategory": "Basic concepts",
     "query": "What is the difference between AI in healthcare and digital health?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["healthcare", "digital"]},

    {"id": "S20", "profile": "Students", "subcategory": "Basic concepts",
     "query": "Can AI make clinical decisions on its own?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["autonom", "decision", "clinical"]},

    # Exploring the topic for coursework
    {"id": "S21", "profile": "Students", "subcategory": "Coursework",
     "query": "What is fairness in AI?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["fairness", "bias"]},

    {"id": "S22", "profile": "Students", "subcategory": "Coursework",
     "query": "What is precision medicine?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["precision", "personali"]},

    {"id": "S23", "profile": "Students", "subcategory": "Coursework",
     "query": "What is the difference between telemedicine and telehealth?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["telemedicine", "telehealth"]},

    {"id": "S24", "profile": "Students", "subcategory": "Coursework",
     "query": "Is XAI related to clinical decision support?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["XAI", "clinical"]},

    {"id": "S25", "profile": "Students", "subcategory": "Coursework",
     "query": "What is health informatics?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["health informatics", "biomedical"]},

    {"id": "S26", "profile": "Students", "subcategory": "Coursework",
     "query": "What is ambient assisted living?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["ambient", "assisted"]},

    {"id": "S27", "profile": "Students", "subcategory": "Coursework",
     "query": "What are wearable health technologies?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["wearable", "sensor"]},

    {"id": "S28", "profile": "Students", "subcategory": "Coursework",
     "query": "What is frailty syndrome?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["frailty", "vulnerab"]},

    {"id": "S29", "profile": "Students", "subcategory": "Coursework",
     "query": "How is AI used in cancer detection?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["cancer", "detect"]},

    {"id": "S30", "profile": "Students", "subcategory": "Coursework",
     "query": "What is human-centred AI?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["human"]},

    {"id": "S31", "profile": "Students", "subcategory": "Coursework",
     "query": "What is AI ethics in healthcare?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["ethics", "healthcare"]},

    {"id": "S32", "profile": "Students", "subcategory": "Coursework",
     "query": "What is privacy and data governance in health AI?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["privacy", "data", "health"]},

    {"id": "S33", "profile": "Students", "subcategory": "Coursework",
     "query": "Can you help me with my homework on AI and aging?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["aging"]},

    {"id": "S34", "profile": "Students", "subcategory": "Coursework",
     "query": "Write me a summary of health and wellbeing systems for my essay",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["health", "wellbeing"]},

    {"id": "S35", "profile": "Students", "subcategory": "Coursework",
     "query": "What are the main health and AI subtopics?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["topic"]},

    # Off-topic / misuse
    {"id": "S36", "profile": "Students", "subcategory": "Off-topic",
     "query": "Write an essay for me about philosophy",
     "expected_behavior": "refuse", "expected_banner": BANNER_RED,
     "expected_keywords": []},

    {"id": "S37", "profile": "Students", "subcategory": "Off-topic",
     "query": "What is the weather today?",
     "expected_behavior": "refuse", "expected_banner": BANNER_RED,
     "expected_keywords": []},

    {"id": "S38", "profile": "Students", "subcategory": "Off-topic",
     "query": "Who won the last World Cup?",
     "expected_behavior": "refuse", "expected_banner": BANNER_RED,
     "expected_keywords": []},

    {"id": "S39", "profile": "Students", "subcategory": "Off-topic",
     "query": "Can you book me a flight?",
     "expected_behavior": "refuse", "expected_banner": BANNER_RED,
     "expected_keywords": []},

    {"id": "S40", "profile": "Students", "subcategory": "Off-topic",
     "query": 'Translate this text to French: "Health AI is important"',
     "expected_behavior": "refuse", "expected_banner": BANNER_RED,
     "expected_keywords": []},

    # ----------------------------------------------------------
    # B. Professors with large experience
    # ----------------------------------------------------------

    # Searching specific publications
    {"id": "P01", "profile": "Professors", "subcategory": "Publications",
     "query": "List all papers from UDCLV on AI in healthcare",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["UDCLV", "healthcare"]},

    {"id": "P02", "profile": "Professors", "subcategory": "Publications",
     "query": "Papers on cancer detection published by UNINOVIS partners",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["cancer"]},

    {"id": "P03", "profile": "Professors", "subcategory": "Publications",
     "query": "How many papers have been published by UNINOVIS partners in 2024?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["2024"]},

    {"id": "P04", "profile": "Professors", "subcategory": "Publications",
     "query": "List papers from UMA on fall detection",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["UMA"]},

    {"id": "P05", "profile": "Professors", "subcategory": "Publications",
     "query": "Papers about AI and mental health",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["mental health"]},

    {"id": "P06", "profile": "Professors", "subcategory": "Publications",
     "query": "Research on AI and aging within UNINOVIS",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["aging"]},

    {"id": "P07", "profile": "Professors", "subcategory": "Publications",
     "query": "Studies on clinical decision support systems",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["decision", "clinical"]},

    {"id": "P08", "profile": "Professors", "subcategory": "Publications",
     "query": "Papers on wearable sensors and health monitoring",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["wearable", "sensor"]},

    {"id": "P09", "profile": "Professors", "subcategory": "Publications",
     "query": "Research on neurodegenerative disorders and AI",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["neurodegen"]},

    {"id": "P10", "profile": "Professors", "subcategory": "Publications",
     "query": "Which are the most cited papers in the database?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["cited"]},

    # Researcher lookup
    {"id": "P11", "profile": "Professors", "subcategory": "Researchers",
     "query": "List all researchers from USPN",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["USPN"]},

    {"id": "P12", "profile": "Professors", "subcategory": "Researchers",
     "query": "List researchers working on AI and healthcare",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["healthcare"]},

    {"id": "P13", "profile": "Professors", "subcategory": "Researchers",
     "query": "Which researchers work on genomics?",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["genomic"]},

    {"id": "P14", "profile": "Professors", "subcategory": "Researchers",
     "query": "What topics does Brigitte Séroussi work on?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["Séroussi"]},

    {"id": "P15", "profile": "Professors", "subcategory": "Researchers",
     "query": "What has S. Marrone published?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["Marrone"]},

    {"id": "P16", "profile": "Professors", "subcategory": "Researchers",
     "query": "Papers by E. Casilari",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["Casilari"]},

    {"id": "P17", "profile": "Professors", "subcategory": "Researchers",
     "query": "What are the research interests of Fabrizio Esposito?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["Esposito"]},

    {"id": "P18", "profile": "Professors", "subcategory": "Researchers",
     "query": "Who works on AI and health at UMA?",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["UMA"]},

    {"id": "P19", "profile": "Professors", "subcategory": "Researchers",
     "query": "Which researchers at USPN work on health informatics?",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["USPN"]},

    {"id": "P20", "profile": "Professors", "subcategory": "Researchers",
     "query": "List researchers from the University of Tirana",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["Tirana"]},

    # Research projects
    {"id": "P21", "profile": "Professors", "subcategory": "Projects",
     "query": "What is the TAILOR project about?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["TAILOR"]},

    {"id": "P22", "profile": "Professors", "subcategory": "Projects",
     "query": "Describe the IntelliMan project",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["IntelliMan"]},

    {"id": "P23", "profile": "Professors", "subcategory": "Projects",
     "query": "What does the DUCA project propose about data governance?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["DUCA", "data"]},

    {"id": "P24", "profile": "Professors", "subcategory": "Projects",
     "query": "What does the dAIbetes project address?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["dAIbetes"]},

    {"id": "P25", "profile": "Professors", "subcategory": "Projects",
     "query": "Describe the contributions of the MENHIR project to mental health",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["MENHIR", "mental"]},

    {"id": "P26", "profile": "Professors", "subcategory": "Projects",
     "query": "What is the AIAS project about?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["AIAS"]},

    {"id": "P27", "profile": "Professors", "subcategory": "Projects",
     "query": "What does the InnoGuard project address?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["InnoGuard"]},

    {"id": "P28", "profile": "Professors", "subcategory": "Projects",
     "query": "Describe the AI-MAPS project",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["AI-MAPS"]},

    # Collaboration and gap analysis
    {"id": "P29", "profile": "Professors", "subcategory": "Collaboration",
     "query": "Which universities have collaborated on publications?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["collaborat"]},

    {"id": "P30", "profile": "Professors", "subcategory": "Collaboration",
     "query": "Show collaborations between UMA and UDCLV",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["UMA", "UDCLV"]},

    {"id": "P31", "profile": "Professors", "subcategory": "Collaboration",
     "query": "Which partners have joint publications on AI and health?",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["health"]},

    {"id": "P32", "profile": "Professors", "subcategory": "Gap analysis",
     "query": "What health and wellbeing topics have not been studied in UNINOVIS?",
     "expected_behavior": "answer", "expected_banner": BANNER_RED,
     "expected_keywords": ["gap", "not"]},

    {"id": "P33", "profile": "Professors", "subcategory": "Gap analysis",
     "query": "Are there gaps in UNINOVIS research on AI and aging?",
     "expected_behavior": "answer", "expected_banner": BANNER_RED,
     "expected_keywords": ["aging"]},

    {"id": "P34", "profile": "Professors", "subcategory": "Gap analysis",
     "query": "Which health and AI subtopics are least studied?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["least", "topic"]},

    {"id": "P35", "profile": "Professors", "subcategory": "Gap analysis",
     "query": "Is there any research on AI and frailty syndrome in UNINOVIS?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["frailty"]},

    {"id": "P36", "profile": "Professors", "subcategory": "Collaboration",
     "query": "Which topics are shared between UMA and USPN?",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["UMA", "USPN"]},

    {"id": "P37", "profile": "Professors", "subcategory": "Collaboration",
     "query": "Compare the research focus of UMA and UDCLV",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["UMA", "UDCLV"]},

    # Figures and maps
    {"id": "P38", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a figure with all the publications per partner",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["figure", "iframe"]},

    {"id": "P39", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a figure of studies on the topic AI and Healthcare",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["figure", "iframe"]},

    {"id": "P40", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a map with the number of research projects per partner",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["map", "iframe"]},

    {"id": "P41", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a figure of the collaborations among the partners",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["collaborat", "iframe"]},

    {"id": "P42", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a figure of the collaborations in the year 2025",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["2025", "iframe"]},

    {"id": "P43", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a map of publications on wellbeing systems",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["map", "iframe"]},

    {"id": "P44", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a figure of papers by year",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["figure", "iframe"]},

    {"id": "P45", "profile": "Professors", "subcategory": "Figures",
     "query": "Show a figure of publications on cancer detection",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["cancer", "iframe"]},

    # Advanced / boundary topics
    {"id": "P46", "profile": "Professors", "subcategory": "Advanced",
     "query": "What is the microbiome and how is AI used in microbiome research?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["microbiome"]},

    {"id": "P47", "profile": "Professors", "subcategory": "Advanced",
     "query": "What are the ethical challenges of predictive diagnosis?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["predictive", "diagnosis", "ethical"]},

    {"id": "P48", "profile": "Professors", "subcategory": "Advanced",
     "query": "Explain pharmacovigilance and its relation to AI",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["pharmacovigilance"]},

    {"id": "P49", "profile": "Professors", "subcategory": "Advanced",
     "query": "What is the role of NLP in healthcare?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["NLP", "natural language"]},

    {"id": "P50", "profile": "Professors", "subcategory": "Advanced",
     "query": "What are the WHO principles for AI in health?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["WHO", "principle"]},

    {"id": "P51", "profile": "Professors", "subcategory": "Advanced",
     "query": "What is federated learning in the context of health data?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["federated"]},

    {"id": "P52", "profile": "Professors", "subcategory": "Advanced",
     "query": "What is radiomics?",
     "expected_behavior": "answer", "expected_banner": BANNER_YELLOW,
     "expected_keywords": ["radiomics"]},

    # ----------------------------------------------------------
    # C. Admin staff
    # ----------------------------------------------------------

    {"id": "A01", "profile": "Admin", "subcategory": "Institutional overview",
     "query": "How many papers have UNINOVIS partners published in total?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["paper", "total"]},

    {"id": "A02", "profile": "Admin", "subcategory": "Institutional overview",
     "query": "Show all publications per partner",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["UMA", "THUAS"]},

    {"id": "A03", "profile": "Admin", "subcategory": "Institutional overview",
     "query": "How many researchers are there per university?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["researcher"]},

    {"id": "A04", "profile": "Admin", "subcategory": "Institutional overview",
     "query": "How many papers were published in 2025?",
     "expected_behavior": "answer", "expected_banner": None,
     "expected_keywords": ["2025"]},

    {"id": "A05", "profile": "Admin", "subcategory": "Institutional overview",
     "query": "List research projects involving UNINOVIS partners",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["project"]},

    {"id": "A06", "profile": "Admin", "subcategory": "Reporting",
     "query": "Show a figure of the number of papers per topic",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["figure", "iframe"]},

    {"id": "A07", "profile": "Admin", "subcategory": "Reporting",
     "query": "Show a map of researchers working on AI and healthcare",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["map", "iframe"]},

    {"id": "A08", "profile": "Admin", "subcategory": "Reporting",
     "query": "Show a figure of the collaborations in the year 2025",
     "expected_behavior": "answer", "expected_banner": BANNER_GREEN,
     "expected_keywords": ["collaborat", "iframe"]},
]

assert len(BENCHMARK_QUERIES) == 100, f"Expected 100 queries, got {len(BENCHMARK_QUERIES)}"


# ============================================================
# SSE streaming client
# ============================================================

def api_chat_stream(server: str, token: str, agent_id: str, query: str,
                    model: str = None, prompt_level: str = "stringent",
                    transparency_level: str = "scaffolded",
                    timeout: int = 180) -> dict:
    """Send a query via SSE streaming and collect the full response + banner."""
    params = {
        "agent_id": agent_id,
        "message": query,
        "token": token,
        "prompt_level": prompt_level,
        "transparency_level": transparency_level,
    }
    if model:
        params["model"] = model

    url = f"{server}/api/chat/stream?{urlencode(params)}"

    start = time.time()
    try:
        req = Request(url)
        with urlopen(req, timeout=timeout) as resp:
            full_response = ""
            for raw_line in resp:
                line = raw_line.decode("utf-8").strip()
                if line.startswith("data: "):
                    data = line[6:].replace("\\n", "\n")
                    full_response += data
                elif line == "event: done":
                    break

        elapsed = time.time() - start
        return {
            "response": full_response,
            "latency": round(elapsed, 2),
            "error": None,
        }
    except (HTTPError, URLError, Exception) as e:
        elapsed = time.time() - start
        error_msg = str(e)
        if isinstance(e, HTTPError):
            try:
                error_msg = f"HTTP {e.code}: {e.read().decode('utf-8', errors='replace')}"
            except Exception:
                error_msg = f"HTTP {e.code}"
        return {
            "response": "",
            "latency": round(elapsed, 2),
            "error": error_msg,
        }


# ============================================================
# Response analysis
# ============================================================

def detect_banner(response: str) -> str:
    """Detect the procedural banner colour from the response HTML."""
    resp = response[:1500]

    # Green banners
    green_patterns = [
        "verified data", "datos verificados", "directly from",
        "background-color:#d4edda", "background:#d4edda",
        "#28a745",
    ]
    # Red banners
    red_patterns = [
        "unverified", "no verificad", "speculation",
        "outside", "off-topic", "fuera del",
        "background-color:#f8d7da", "background:#f8d7da",
        "#dc3545",
    ]
    # Yellow banners
    yellow_patterns = [
        "ai interpretation", "ai commentary", "interpretación",
        "on-topic, undefined", "undefined topic",
        "background-color:#fff3cd", "background:#fff3cd",
        "#ffc107",
    ]

    resp_lower = resp.lower()

    if any(p in resp_lower for p in green_patterns):
        return BANNER_GREEN
    if any(p in resp_lower for p in red_patterns):
        return BANNER_RED
    if any(p in resp_lower for p in yellow_patterns):
        return BANNER_YELLOW

    return "none"


def analyse_response(query_def: dict, response: str) -> dict:
    """Analyse a response against expected outcomes."""
    resp_lower = response.lower()
    expected = query_def["expected_behavior"]

    result = {
        "responded": bool(response and len(response.strip()) > 20),
        "refused": False,
        "correct_behavior": False,
        "keyword_matches": 0,
        "keyword_total": len(query_def.get("expected_keywords", [])),
        "detected_banner": detect_banner(response),
        "banner_correct": None,
        "hallucination_warnings": 0,
        "response_length": len(response),
    }

    # Detect refusal
    refusal_patterns = [
        "outside my scope", "outside my domain", "cannot help",
        "not within my", "i'm afraid", "i am afraid",
        "not related to", "beyond my scope", "outside the scope",
        "fuera de mi", "no puedo ayudar",
    ]
    result["refused"] = any(p in resp_lower for p in refusal_patterns)

    # Check correct behavior
    if expected == "refuse":
        result["correct_behavior"] = result["refused"] or result["detected_banner"] == BANNER_RED
    elif expected == "answer":
        result["correct_behavior"] = result["responded"] and not result["refused"]

    # Keyword matches
    keywords = query_def.get("expected_keywords", [])
    if keywords and result["responded"]:
        result["keyword_matches"] = sum(1 for kw in keywords if kw.lower() in resp_lower)

    # Banner correctness
    expected_banner = query_def.get("expected_banner")
    if expected_banner:
        result["banner_correct"] = result["detected_banner"] == expected_banner

    # Hallucination warnings
    result["hallucination_warnings"] = (
        response.count("not found in database")
        + response.count("hallucin")
        + response.count("could not verify")
        + response.count("no encontr")
    )

    return result


# ============================================================
# Excel report
# ============================================================

def generate_report(results: list, output_path: str, models: list):
    """Generate Excel benchmark report with summary, per-profile, and detailed sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\nERROR: openpyxl is required for Excel reports. Install with: pip install openpyxl")
        # Fallback: save as JSON
        json_path = output_path.replace(".xlsx", ".json")
        with open(json_path, "w") as f:
            json.dump(results, f, indent=2, default=str)
        print(f"Results saved as JSON: {json_path}")
        return

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_cell(ws, row, col, value, fill=None, bold=False, wrap=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if bold:
            cell.font = Font(bold=True)
        if wrap:
            cell.alignment = Alignment(wrap_text=True)
        return cell

    def write_header(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "Summary"
    write_header(ws1, ["Metric"] + models)

    for model_idx, model in enumerate(models):
        mr = [r for r in results if r["model"] == model]
        total = len(mr)
        correct = sum(1 for r in mr if r["correct_behavior"])
        banner_tested = [r for r in mr if r["banner_correct"] is not None]
        banner_ok = sum(1 for r in banner_tested if r["banner_correct"])
        kw_total = sum(r["keyword_total"] for r in mr)
        kw_match = sum(r["keyword_matches"] for r in mr)
        halluc = sum(r["hallucination_warnings"] for r in mr)
        avg_lat = sum(r["latency"] for r in mr) / max(total, 1)
        errors = sum(1 for r in mr if r["error"])

        # Banner distribution
        b_green = sum(1 for r in mr if r["detected_banner"] == BANNER_GREEN)
        b_yellow = sum(1 for r in mr if r["detected_banner"] == BANNER_YELLOW)
        b_red = sum(1 for r in mr if r["detected_banner"] == BANNER_RED)
        b_none = sum(1 for r in mr if r["detected_banner"] == "none")

        metrics = [
            ("Total queries", total),
            ("Correct behavior", f"{correct}/{total} ({100*correct//max(total,1)}%)"),
            ("Banner accuracy", f"{banner_ok}/{len(banner_tested)} ({100*banner_ok//max(len(banner_tested),1)}%)"),
            ("Keyword coverage", f"{kw_match}/{kw_total} ({100*kw_match//max(kw_total,1)}%)"),
            ("Hallucination warnings", halluc),
            ("Average latency", f"{avg_lat:.1f}s"),
            ("Errors", errors),
            ("", ""),
            ("Banners: Green", b_green),
            ("Banners: Yellow", b_yellow),
            ("Banners: Red", b_red),
            ("Banners: None detected", b_none),
        ]

        for row_idx, (label, value) in enumerate(metrics, 2):
            style_cell(ws1, row_idx, 1, label, bold=True)
            style_cell(ws1, row_idx, 2 + model_idx, str(value))

    ws1.column_dimensions["A"].width = 25
    for i in range(len(models)):
        ws1.column_dimensions[chr(66 + i)].width = 25

    # ---- Sheet 2: By Profile ----
    ws2 = wb.create_sheet("By Profile")
    write_header(ws2, ["Profile", "Subcategory", "Model", "Queries", "Correct",
                        "Banner OK", "Keywords", "Avg Latency"])

    profiles = ["Students", "Professors", "Admin"]
    row = 2
    for profile in profiles:
        subcats = sorted(set(r["subcategory"] for r in results if r["profile"] == profile))
        for subcat in subcats:
            for model in models:
                sr = [r for r in results
                      if r["profile"] == profile and r["subcategory"] == subcat and r["model"] == model]
                total = len(sr)
                correct = sum(1 for r in sr if r["correct_behavior"])
                bt = [r for r in sr if r["banner_correct"] is not None]
                bok = sum(1 for r in bt if r["banner_correct"])
                kw = sum(r["keyword_matches"] for r in sr)
                kwt = sum(r["keyword_total"] for r in sr)
                avg_lat = sum(r["latency"] for r in sr) / max(total, 1)

                fill = green_fill if correct == total else (red_fill if correct == 0 else yellow_fill)

                style_cell(ws2, row, 1, profile)
                style_cell(ws2, row, 2, subcat)
                style_cell(ws2, row, 3, model)
                style_cell(ws2, row, 4, total)
                style_cell(ws2, row, 5, f"{correct}/{total}", fill=fill)
                style_cell(ws2, row, 6, f"{bok}/{len(bt)}" if bt else "n/a")
                style_cell(ws2, row, 7, f"{kw}/{kwt}" if kwt else "n/a")
                style_cell(ws2, row, 8, f"{avg_lat:.1f}s")
                row += 1

    for c, w in zip("ABCDEFGH", [14, 22, 22, 9, 10, 10, 10, 12]):
        ws2.column_dimensions[c].width = w

    # ---- Sheet 3: Detailed Results ----
    ws3 = wb.create_sheet("Detailed Results")
    write_header(ws3, ["ID", "Profile", "Subcategory", "Query", "Model",
                        "Correct?", "Banner Expected", "Banner Detected", "Banner OK?",
                        "Keywords", "Halluc.", "Latency",
                        "Response (first 500 chars)", "Error"])

    row = 2
    for r in sorted(results, key=lambda x: (x["model"], x["id"])):
        style_cell(ws3, row, 1, r["id"])
        style_cell(ws3, row, 2, r["profile"])
        style_cell(ws3, row, 3, r["subcategory"])
        style_cell(ws3, row, 4, r["query"], wrap=True)
        style_cell(ws3, row, 5, r["model"])

        c_fill = green_fill if r["correct_behavior"] else red_fill
        style_cell(ws3, row, 6, "Yes" if r["correct_behavior"] else "No", fill=c_fill)

        style_cell(ws3, row, 7, r.get("expected_banner") or "any")
        style_cell(ws3, row, 8, r["detected_banner"])

        if r["banner_correct"] is not None:
            b_fill = green_fill if r["banner_correct"] else red_fill
            style_cell(ws3, row, 9, "Yes" if r["banner_correct"] else "No", fill=b_fill)
        else:
            style_cell(ws3, row, 9, "n/a")

        style_cell(ws3, row, 10, f"{r['keyword_matches']}/{r['keyword_total']}")

        h_fill = red_fill if r["hallucination_warnings"] > 0 else None
        style_cell(ws3, row, 11, r["hallucination_warnings"], fill=h_fill)
        style_cell(ws3, row, 12, f"{r['latency']}s")

        preview = r.get("response_text", "")[:500].replace("\n", " ")
        style_cell(ws3, row, 13, preview, wrap=True)
        style_cell(ws3, row, 14, r.get("error") or "")
        row += 1

    for c, w in zip("ABCDEFGHIJKLMN", [6, 12, 18, 45, 22, 9, 14, 14, 10, 10, 8, 9, 55, 30]):
        ws3.column_dimensions[c].width = w

    wb.save(output_path)
    print(f"\nReport saved: {output_path}")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Benchmark 100 queries for the Health and Wellbeing Systems Agent",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  python benchmark100.py --server http://localhost:8000
  python benchmark100.py --server http://localhost:8000 --user admin --password admin
  python benchmark100.py --server http://localhost:8000 --models mistral-small-latest,mistral-large-latest
  python benchmark100.py --server http://localhost:8000 --categories Students
  python benchmark100.py --server http://localhost:8000 --ids S10,P01,A01
""")
    parser.add_argument("--server", default="http://localhost:8000", help="Server URL (default: http://localhost:8000)")
    parser.add_argument("--user", default=None, help="Login username (default: from BENCHMARK_USER env or 'admin')")
    parser.add_argument("--password", default=None, help="Login password (default: from BENCHMARK_PASSWORD env or 'admin')")
    parser.add_argument("--token", default=None, help="Auth token (skip login if provided)")
    parser.add_argument("--agent", default="health_wellbeing_sistems", help="Agent ID (default: health_wellbeing_sistems)")
    parser.add_argument("--models", default="mistral-small-latest",
                        help="Comma-separated list of models to test (default: mistral-small-latest)")
    parser.add_argument("--categories", default=None,
                        help="Comma-separated profiles to run: Students,Professors,Admin (default: all)")
    parser.add_argument("--ids", default=None,
                        help="Comma-separated query IDs to run (e.g., S01,P05,A03)")
    parser.add_argument("--output", default=None, help="Output Excel path")
    parser.add_argument("--timeout", type=int, default=180, help="Timeout per query in seconds (default: 180)")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay between queries in seconds (default: 2)")
    parser.add_argument("--prompt-level", default="stringent",
                        choices=["stringent", "tolerant", "lax"], help="Prompt level (default: stringent)")
    parser.add_argument("--transparency", default="scaffolded",
                        choices=["scaffolded", "unscaffolded"], help="Transparency level (default: scaffolded)")
    args = parser.parse_args()

    # Authenticate
    if args.token:
        token = args.token
        print(f"Using provided token.")
    else:
        username = args.user or os.environ.get("BENCHMARK_USER", "admin")
        password = args.password or os.environ.get("BENCHMARK_PASSWORD", "admin")
        print(f"Logging in as '{username}'...", end=" ", flush=True)
        token = login(args.server, username, password)
        print("OK")

    models = [m.strip() for m in args.models.split(",")]

    # Filter queries
    queries = BENCHMARK_QUERIES
    if args.ids:
        selected_ids = {x.strip() for x in args.ids.split(",")}
        queries = [q for q in queries if q["id"] in selected_ids]
    elif args.categories:
        selected_cats = {x.strip() for x in args.categories.split(",")}
        queries = [q for q in queries if q["profile"] in selected_cats]

    if not queries:
        print("ERROR: No queries matched the filter criteria.")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or f"benchmark100_{args.agent}_{timestamp}.xlsx"

    print(f"Benchmark: {args.agent}")
    print(f"Server:    {args.server}")
    print(f"Models:    {', '.join(models)}")
    print(f"Queries:   {len(queries)}")
    print(f"Prompt:    {args.prompt_level}")
    print(f"Transp.:   {args.transparency}")
    print(f"Total runs: {len(queries) * len(models)}")
    print()

    all_results = []

    for model in models:
        print(f"{'='*60}")
        print(f"  Model: {model}")
        print(f"{'='*60}")

        current_profile = None
        for i, q in enumerate(queries):
            if q["profile"] != current_profile:
                current_profile = q["profile"]
                print(f"\n  --- {current_profile} ---")

            label = f"  [{i+1}/{len(queries)}] {q['id']} ({q['subcategory']})"
            query_preview = q["query"][:50]
            print(f"{label}: {query_preview}...", end=" ", flush=True)

            result = api_chat_stream(
                args.server, token, args.agent, q["query"],
                model=model, prompt_level=args.prompt_level,
                transparency_level=args.transparency,
                timeout=args.timeout,
            )

            analysis = analyse_response(q, result["response"])

            all_results.append({
                "id": q["id"],
                "profile": q["profile"],
                "subcategory": q["subcategory"],
                "query": q["query"],
                "model": model,
                "expected_banner": q.get("expected_banner"),
                "response_text": result["response"],
                "latency": result["latency"],
                "error": result["error"],
                **analysis,
            })

            # Status line
            status = "OK" if analysis["correct_behavior"] else "FAIL"
            banner_info = f" [{analysis['detected_banner']}]"
            error_info = f" ERROR: {result['error']}" if result["error"] else ""
            print(f"{status}{banner_info} ({result['latency']}s){error_info}")

            time.sleep(args.delay)

        print()

    # Generate report
    generate_report(all_results, output, models)

    # Print summary
    print(f"\n{'='*60}")
    print("  SUMMARY")
    print(f"{'='*60}")
    for model in models:
        mr = [r for r in all_results if r["model"] == model]
        total = len(mr)
        correct = sum(1 for r in mr if r["correct_behavior"])
        bt = [r for r in mr if r["banner_correct"] is not None]
        bok = sum(1 for r in bt if r["banner_correct"])
        avg_lat = sum(r["latency"] for r in mr) / max(total, 1)

        print(f"\n  {model}:")
        print(f"    Correct behavior: {correct}/{total} ({100*correct//max(total,1)}%)")
        print(f"    Banner accuracy:  {bok}/{len(bt)} ({100*bok//max(len(bt),1)}%)")
        print(f"    Avg latency:      {avg_lat:.1f}s")

        for profile in ["Students", "Professors", "Admin"]:
            pr = [r for r in mr if r["profile"] == profile]
            if pr:
                pc = sum(1 for r in pr if r["correct_behavior"])
                print(f"      {profile}: {pc}/{len(pr)} correct")


if __name__ == "__main__":
    main()
